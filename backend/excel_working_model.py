"""
Financial Timeline Engine
Sprint 10 - Professional Excel Working Model

Builds a 6-sheet Excel workbook from the deterministic student workspace:

  Sheet 1 - Financial Data        (Metric, Canonical, Period, Value, Unit,
                                   Currency, Source, Page, Evidence,
                                   Provenance, Status)
  Sheet 2 - Ratio Analysis        (REAL Excel formulas referencing the
                                   Financial Data sheet — e.g.
                                   =ROUND('Financial Data'!E5/'Financial Data'!E9,4)
                                   never a pasted result)
  Sheet 3 - External Variables    (Variable, Value, Unit, Period/Date,
                                   Origin, Source, Status, Student Input)
  Sheet 4 - Comparison            (Canonical Metric, Company A, Company B,
                                   Period, Difference, Status, Evidence)
  Sheet 5 - Driver Analysis       (evidence-backed observations + sources)
  Sheet 6 - Assignment Requirements (Requirement, Status, Result,
                                   Evidence/Source, Review/Blocked reason)

Formatting is deliberately restrained and professional: navy headers,
white bold header text, muted borders, frozen header rows, sensible
column widths, and per-column number formats (currency / percentage /
ratio). No neon, no decorative dashboards.

PURE module — no Streamlit, no AI. Deterministic.
"""

from __future__ import annotations

import io
import re
import zipfile
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.formula_engine import FORMULA_REGISTRY

# ---------------------------------------------------------------------------
# Professional style tokens
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")     # navy
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_BODY_FONT = Font(name="Calibri", size=11)
_TITLE_FONT = Font(name="Calibri", size=13, bold=True, color="1F3864")
_NOTE_FONT = Font(name="Calibri", size=10, italic=True, color="595959")

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_CENTER = Alignment(horizontal="center", vertical="center")


def _style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        c.border = _BORDER


def _style_body(ws, first_row: int, last_row: int, ncols: int,
                number_formats: Optional[Dict[int, str]] = None) -> None:
    number_formats = number_formats or {}
    for r in range(first_row, last_row + 1):
        for col in range(1, ncols + 1):
            c = ws.cell(row=r, column=col)
            c.font = _BODY_FONT
            c.border = _BORDER
            if col in number_formats and c.value is not None:
                c.number_format = number_formats[col]
            c.alignment = _RIGHT if col in number_formats else _LEFT


def _set_widths(ws, widths: List[float]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _currency_fmt(currency: str) -> str:
    """Professional currency number format with parentheses negatives."""
    if str(currency or "").strip().upper() in ("USD", "INR", "EUR", "GBP", ""):
        return '#,##0.00_);(#,##0.00)'
    return '#,##0.00_);(#,##0.00)'


# ---------------------------------------------------------------------------
# Number-format detection per canonical metric
# ---------------------------------------------------------------------------

_PERCENT_METRICS = {
    "ROE", "ROA", "Profit Margin", "Operating Margin",
    "Revenue Growth", "EPS Growth", "CAGR",
}
_RATIO_METRICS = {"Current Ratio", "Debt to Equity"}


def _fmt_for_metric(metric: str) -> str:
    if metric in _PERCENT_METRICS:
        return "0.00%"
    if metric in _RATIO_METRICS:
        return "0.00"
    return '#,##0.00_);(#,##0.00)'


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------


def _financial_data_sheet(ws, workspace: Dict[str, Any]) -> Dict[str, int]:
    """Sheet 1 — Financial Data. Returns {metric_key: excel_row} so the
    Ratio Analysis sheet can build REAL cell references."""
    headers = ["Metric", "Canonical Metric", "Period", "Value", "Unit",
               "Currency", "Source", "Page", "Evidence", "Provenance", "Status"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    row_index: Dict[str, int] = {}
    facts = workspace.get("normalized_facts") or []
    # Only numeric facts belong on the data sheet.
    r = 2
    for f in facts:
        value = f.get("value")
        if value is None:
            continue
        metric = str(f.get("metric") or "")
        ws.append([
            metric,
            str(f.get("canonical") or "—"),
            str(f.get("period") or "—"),
            float(value),
            str(f.get("unit") or "—"),
            str(f.get("currency") or "—"),
            str(f.get("source") or "—"),
            str(f.get("page") or "—"),
            str(f.get("evidence") or ""),
            str(f.get("provenance_tier") or "—"),
            str(f.get("normalization_status") or "—"),
        ])
        if metric:
            row_index.setdefault(metric, r)
        r += 1
    _style_body(ws, 2, r - 1, len(headers), number_formats={4: '#,##0.00_);(#,##0.00)'})
    _set_widths(ws, [26, 18, 12, 16, 10, 10, 30, 10, 46, 16, 16])
    return row_index


def _ratio_analysis_sheet(ws, workspace: Dict[str, Any],
                          row_index: Dict[str, int]) -> None:
    """Sheet 2 — Ratio Analysis with REAL Excel formulas. The formula cell
    references the Financial Data sheet (e.g. =ROUND('Financial Data'!E5/
    'Financial Data'!E9,4)); the calculated result is never pasted."""
    headers = ["Metric", "Formula", "Excel Formula", "Status", "Notes"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    facts = {(f.get("metric") or ""): f for f in (workspace.get("normalized_facts") or [])}
    r = 2
    for key, reg in FORMULA_REGISTRY.items():
        inputs = reg.required_inputs
        metric = reg.display_name
        if not all(inp in row_index for inp in inputs):
            ws.append([metric, reg.formula, "—", "BLOCKED",
                       "One or more required inputs are not disclosed in the Financial Data sheet."])
            _style_body(ws, r, r, len(headers))
            r += 1
            continue
        refs = [f"'Financial Data'!E{row_index[inp]}" for inp in inputs]
        if len(refs) == 2:
            formula = f"=ROUND({refs[0]}/{refs[1]},4)"
        elif len(refs) == 3:
            # Growth-style: (cur - prev) / prev
            formula = f"=ROUND(({refs[0]}-{refs[1]})/{refs[2]},4)"
        else:
            continue
        status = "DERIVED"
        notes = "Live Excel formula — recalculates if the Financial Data sheet changes."
        if reg.kind == "percent":
            notes += " Result shown as a percentage."
        elif reg.kind == "ratio":
            notes += " Result shown as a ratio."
        ws.append([metric, reg.formula, formula, status, notes])
        c = ws.cell(row=r, column=3)
        c.number_format = _fmt_for_metric(metric)
        _style_body(ws, r, r, len(headers))
        r += 1

    _set_widths(ws, [20, 40, 46, 14, 58])


def _external_variables_sheet(ws, workspace: Dict[str, Any]) -> None:
    """Sheet 3 — External Variables (student-entered; never document data)."""
    headers = ["Variable", "Value", "Unit", "Period/Date", "Origin",
               "Source", "Status", "Student Input"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    r = 2
    for var in workspace.get("external_variables") or []:
        ws.append([
            var.get("name") or "—",
            var.get("value"),
            var.get("unit") or "—",
            var.get("period") or "—",
            var.get("origin") or "—",
            var.get("source") or "—",
            var.get("status_label") or "—",
            "Yes" if var.get("student_entered") else "No",
        ])
        r += 1
    _style_body(ws, 2, r - 1, len(headers), number_formats={2: '0.00'})
    _set_widths(ws, [22, 14, 10, 14, 18, 24, 18, 12])


def _comparison_sheet(ws, workspace: Dict[str, Any]) -> None:
    """Sheet 4 — Comparison (two companies aligned on canonical metrics)."""
    headers = ["Canonical Metric", "Company A", "Company B", "Period",
               "Difference", "Comparison Status", "Evidence References"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    comparison = workspace.get("comparison") or {}
    rows = list(comparison.get("rows") or []) + list(comparison.get("review_rows") or [])
    r = 2
    for row in rows:
        ev = " | ".join(
            e for e in (str(row.get("evidence_a") or ""), str(row.get("evidence_b") or ""))
            if e
        )
        ws.append([
            row.get("canonical") or "—",
            str(row.get("value_a") or "—"),
            str(row.get("value_b") or "—"),
            str(row.get("period") or "—"),
            row.get("difference") or "—",
            row.get("status_label") or "—",
            ev or "—",
        ])
        r += 1
    if r == 2:
        ws.append(["No comparison data.", "", "", "", "", "", ""])
        r += 1
    _style_body(ws, 2, r - 1, len(headers))
    _set_widths(ws, [22, 18, 18, 12, 14, 24, 60])


def _driver_analysis_sheet(ws, workspace: Dict[str, Any]) -> None:
    """Sheet 5 — Driver Analysis (evidence-backed observations)."""
    headers = ["Metric", "From", "To", "From Value", "To Value",
               "Change", "Direction", "Evidence/Source"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    driver = workspace.get("driver_analysis") or {}
    r = 2
    for obs in driver.get("observations") or []:
        ws.append([
            obs.get("metric") or "—",
            obs.get("from") or "—",
            obs.get("to") or "—",
            obs.get("from_value") or "—",
            obs.get("to_value") or "—",
            obs.get("change_display") or "—",
            obs.get("direction") or "—",
            f"{driver.get('company') or 'Company'} verified period data.",
        ])
        r += 1
    for cause in driver.get("causes") or []:
        ws.append(["Cause", cause.get("period_from") or "—", cause.get("period_to") or "—",
                   "", "", "", "", cause.get("statement") or ""])
        r += 1
    if r == 2:
        ws.append(["No period-over-period data available.", "", "", "", "", "", "",
                   "Cause not established from available evidence."])
        r += 1
    _style_body(ws, 2, r - 1, len(headers))
    _set_widths(ws, [22, 12, 12, 14, 14, 14, 14, 80])


def _requirements_sheet(ws, workspace: Dict[str, Any]) -> None:
    """Sheet 6 — Assignment Requirements checklist."""
    headers = ["Requirement", "Status", "Result", "Evidence/Source",
               "Review/Blocked Reason"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    r = 2
    for req in workspace.get("requirements") or []:
        reason = req.get("detail") or ""
        if req.get("status") in ("REVIEW_REQUIRED", "BLOCKED", "UNANALYZED"):
            reason = reason or "No verified evidence available."
        ws.append([
            req.get("requirement") or "—",
            req.get("status_label") or "—",
            req.get("result") or "—",
            req.get("evidence") or "—",
            reason,
        ])
        r += 1
    if r == 2:
        ws.append(["No requirements defined.", "—", "—", "—", "—"])
        r += 1
    _style_body(ws, 2, r - 1, len(headers))
    _set_widths(ws, [24, 22, 16, 60, 44])


def _qualitative_drivers_sheet(ws, workspace: Dict[str, Any]) -> None:
    """Sheet 7 — Qualitative Drivers & Catalysts (Sprint 11). Evidence-
    classified driver-catalyst relationships only; never fabricated."""
    headers = ["Metric", "Period", "Prior Value", "Current Value", "Change",
               "Numerical Driver", "Catalyst", "Relationship Status",
               "Evidence", "Source", "Page", "Section", "Confidence",
               "Student Interpretation"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    qual = workspace.get("qualitative_drivers") or {}
    rows = list(qual.get("rows") or [])
    r = 2
    for q in rows:
        ws.append([
            q.get("metric") or "—",
            f"{q.get('period_from') or '—'} → {q.get('period_to') or '—'}",
            q.get("from_value") or "—",
            q.get("to_value") or "—",
            q.get("change_display") or "—",
            f"{q.get('numerical_driver') or '—'} ({q.get('driver_change') or '—'})",
            q.get("catalyst") or "—",
            q.get("relationship_label") or "—",
            q.get("evidence") or "—",
            q.get("source") or "—",
            str(q.get("page") if q.get("page") is not None else "—"),
            q.get("section") or "—",
            q.get("confidence") or "—",
            q.get("student_explanation") or "—",
        ])
        r += 1
    if r == 2:
        ws.append(["No qualitative catalyst evidence available."] + [""] * 13)
        r += 1
    _style_body(ws, 2, r - 1, len(headers))
    _set_widths(ws, [16, 14, 12, 12, 10, 22, 22, 22, 46, 26, 8, 26, 12, 60])


# Deterministic serialization: openpyxl stamps the current UTC time into
# docProps/core.xml and into every ZIP entry header, which makes the byte
# stream vary from run to run. Pin a fixed timestamp and re-serialize the
# archive with stable entry times so generation is byte-identical. Sheet
# content, formulas, and values are untouched.
_FIXED_TS = (1980, 1, 1, 0, 0, 0)
_FIXED_ISO = "1980-01-01T00:00:00Z"
# Matches the W3CDTF timestamps openpyxl writes into docProps/core.xml.
_CORE_TS_RE = re.compile(rb">\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z</dcterms:(created|modified)>")


def _pin_document_properties(wb: Workbook) -> None:
    """Pin docProps created to a fixed timestamp (deterministic). openpyxl's
    writer re-stamps `modified` at save time, so that field is normalized in
    the archive rewrite below."""
    wb.properties.created = datetime(*_FIXED_TS)


def _deterministic_zip(data: bytes) -> bytes:
    """Re-serialize the xlsx container with fixed per-entry timestamps and a
    fixed docProps/core.xml created/modified value so the byte stream is
    identical across runs."""
    out_bio = io.BytesIO()
    with zipfile.ZipFile(out_bio, "w", zipfile.ZIP_DEFLATED) as out:
        with zipfile.ZipFile(io.BytesIO(data), "r") as src:
            for info in src.infolist():
                raw = src.read(info.filename)
                if info.filename == "docProps/core.xml":
                    raw = _CORE_TS_RE.sub(
                        lambda m: b">" + _FIXED_ISO.encode("utf-8") + b"</dcterms:" + m.group(1) + b">",
                        raw,
                    )
                zinfo = zipfile.ZipInfo(info.filename, date_time=_FIXED_TS)
                zinfo.compress_type = zipfile.ZIP_DEFLATED
                zinfo.external_attr = info.external_attr
                out.writestr(zinfo, raw)
    return out_bio.getvalue()


def build_excel_working_model(workspace: Dict[str, Any]) -> bytes:
    """Build the complete 7-sheet working model and return the raw .xlsx
    bytes (safe for st.download_button / file write)."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Financial Data"
    row_index = _financial_data_sheet(ws1, workspace)

    ws2 = wb.create_sheet("Ratio Analysis")
    _ratio_analysis_sheet(ws2, workspace, row_index)

    ws3 = wb.create_sheet("External Variables")
    _external_variables_sheet(ws3, workspace)

    ws4 = wb.create_sheet("Comparison")
    _comparison_sheet(ws4, workspace)

    ws5 = wb.create_sheet("Driver Analysis")
    _driver_analysis_sheet(ws5, workspace)

    ws6 = wb.create_sheet("Assignment Requirements")
    _requirements_sheet(ws6, workspace)

    ws7 = wb.create_sheet("Qualitative Drivers")
    _qualitative_drivers_sheet(ws7, workspace)

    _pin_document_properties(wb)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return _deterministic_zip(bio.getvalue())
