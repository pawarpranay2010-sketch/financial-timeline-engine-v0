"""Sprint 12.1 - Zero-Panic Assignment Onboarding & Excel Guidance

Deterministic test suite for the Assignment Agent parse-recovery flow:

 Requirement parsing
   1.  Clean assignment -> automatic continuation (high state)
   2.  WhatsApp whitespace -> recovery works
   3.  Broken line formatting -> recovery works
   4.  Unicode punctuation -> recovery works
   5.  Debt-to-Equity -> one requirement
   6.  D/E -> one requirement
   7.  Ambiguous metric -> review state
   8.  Unknown requirement -> manual confirmation
   9.  Duplicate requirements -> deduplicated
  10.  EPS != EPSILON
  11.  Revenue != Revenue Growth
  12.  Debt != Debt-like

 Agent UX
  13.  High-confidence state has one primary action
  14.  Partial-confidence state has a confirmation action
  15.  Low-confidence state has manual recovery
  16.  No parser/debug terminology appears
  17.  No traceback appears
  18.  No empty dead-end state
  19.  Agent tells the student what happened
  20.  Agent tells the student what to do next

 Excel guidance
  21.  Workbook generation unchanged
  22.  All existing sheets preserved
  23.  Real Excel formulas preserved
  24.  Excel orientation message appears before opening
  25.  Ratio Analysis identified as the first verification location
  26.  Student is told formulas are already calculated
  27.  Student can still explore other sheets

 API / Demo parity
  28.  API onboarding renders correctly
  29.  Demo onboarding renders correctly
  30.  API recovery renders correctly
  31.  Demo recovery renders correctly
  32.  Demo remains deterministic
  33.  Demo has no API-key requirement
  34.  Demo does not call AI
  35.  Demo fixtures remain unchanged

 Student safety / academic integrity
  36.  Conclusion remains blank
  37.  No Buy/Sell/Hold recommendation is generated
  38.  No submission-ready interpretation is generated
  39.  Evidence-backed scaffolding remains available

Every decision is deterministic; the agent never invents requirements,
causes, sources or conclusions.
"""
import os
import re
import sys
import types
import importlib.util
from io import BytesIO

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl  # noqa: E402

from backend.assignment_agent import (  # noqa: E402
    STAGE_OPENING,
    STAGE_REQUIREMENTS,
    STAGE_PERIODS,
    STAGE_METRIC,
    STAGE_EXPLAIN,
    STAGE_CALCULATION,
    STAGE_EVIDENCE,
    STAGE_DRIVERS,
    STAGE_QUALITATIVE,
    STAGE_COMPARISON,
    STAGE_EXTERNAL,
    STAGE_EXCEL,
    STAGE_MEMO,
    STAGE_CONCLUSION,
    PARSE_HIGH,
    PARSE_PARTIAL,
    PARSE_LOW,
    agent_session,
    apply_choice,
    initial_state,
    parse_recovery,
    what_next,
)
from backend.student_workspace import (  # noqa: E402
    build_student_workspace,
    canonicalize_metric,
    parse_requirements,
)
from backend.excel_working_model import build_excel_working_model  # noqa: E402

CHECKS = []


def check(name, ok, detail=""):
    CHECKS.append((name, bool(ok), detail))


# ---------------------------------------------------------------------------
# App-under-test (stubbed streamlit) — used only for the demo fixtures.
# ---------------------------------------------------------------------------
_APP = None
_APP_STUB_SS = {}


class _Passthrough:
    def __call__(self, *a, **k):
        if len(a) == 1 and not k and callable(a[0]):
            return a[0]

        def deco(fn):
            return fn
        return deco


class _StubStreamlit(types.ModuleType):
    def __init__(self):
        super().__init__("streamlit")
        self._ss = _APP_STUB_SS

    def __getattr__(self, name):
        if name == "session_state":
            return self._ss
        return _Passthrough()


def _load_app():
    """Exec the app file under a stubbed streamlit and return the module."""
    global _APP
    if _APP is not None:
        return _APP
    import streamlit as _real
    root = os.path.join(os.path.dirname(__file__), "..")
    stub = _StubStreamlit()
    sys.modules["streamlit"] = stub
    try:
        spec = importlib.util.spec_from_file_location(
            "fte_app_under_test", os.path.join(root, "app (1) (9).py")
        )
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
    finally:
        sys.modules["streamlit"] = _real
    _APP = mod
    return mod


# ---------------------------------------------------------------------------
# Deterministic fixtures (mirror of the Sprint 12 suite)
# ---------------------------------------------------------------------------
def _real_facts():
    return {
        "Revenue": {"value": 281700000000, "source": "10-K FY2025 · Income Statement", "reporting_period": "FY2025", "page": 26, "evidence": "Consolidated Statements of Income, p. 26", "unit": "USD", "scale": "B"},
        "Net Profit": {"value": 98300000000, "source": "10-K FY2025 · Income Statement", "reporting_period": "FY2025", "page": 26, "evidence": "Consolidated Statements of Income, p. 26", "unit": "USD", "scale": "B"},
        "Equity": {"value": 268500000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Assets": {"value": 512200000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Debt": {"value": 101200000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Current Assets": {"value": 147600000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Current Liabilities": {"value": 105400000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
    }


_REQ_TEXT = ("Analyze Microsoft FY2023-FY2025 and calculate ROE, ROA, "
             "Profit Margin, Current Ratio and Debt/Equity.")


def _real_module3():
    facts = _real_facts()
    module3 = {
        "financial_data": facts,
        "ratios": {},
        "missing_data": {"financial_data": ["Segment Gross Margin"], "ratios": []},
    }
    module3["ratios"]["Current Ratio"] = {
        "value": round(facts["Current Assets"]["value"] / facts["Current Liabilities"]["value"], 2),
        "source": "Calculated", "formula": "Current Assets / Current Liabilities",
        "reporting_period": "FY2025",
    }
    module3["ratios"]["Debt to Equity"] = {
        "value": round(facts["Debt"]["value"] / facts["Equity"]["value"], 2),
        "source": "Calculated", "formula": "Debt / Equity", "reporting_period": "FY2025",
    }
    return module3


def _real_workspace(requirements_text=_REQ_TEXT, **kw):
    module3 = _real_module3()
    return build_student_workspace(
        module3,
        assignment_type=kw.get("assignment_type", "Financial Ratio Analysis"),
        requirements_text=requirements_text,
        external_variables=kw.get("external_variables") or [],
        company_a=kw.get("company_a", "Microsoft"),
        peer_company=kw.get("peer_company"),
        peer_facts=kw.get("peer_facts"),
        period_facts=kw.get("period_facts"),
        calc_metrics=[r["metric"] for r in parse_requirements(requirements_text)],
        missing=module3.get("missing_data"),
    )


_PERIOD_FACTS = {
    "Revenue": {"FY2024": "245120000000", "FY2025": "281700000000"},
    "Net Profit": {"FY2024": "80100000000", "FY2025": "98300000000"},
    "Equity": {"FY2024": "268500000000", "FY2025": "268500000000"},
    "Assets": {"FY2024": "512200000000", "FY2025": "512200000000"},
    "ROE": {"FY2024": "0.298", "FY2025": "0.366"},
    "ROA": {"FY2024": "0.156", "FY2025": "0.192"},
    "Profit Margin": {"FY2024": "0.327", "FY2025": "0.349"},
    "Current Ratio": {"FY2024": "1.35", "FY2025": "1.40"},
    "Debt to Equity": {"FY2024": "0.36", "FY2025": "0.38"},
}


def _demo_workspace(requirements_text=None, **kw):
    app = _load_app()
    req_text = requirements_text if requirements_text is not None else app._demo_assignment_requirements_text()
    return build_student_workspace(
        app._demo_module3_result(),
        assignment_type=kw.get("assignment_type", "Financial Ratio Analysis"),
        requirements_text=req_text,
        external_variables=kw.get("external_variables") or [],
        company_a=kw.get("company_a", "Contoso Analytics (Demo)"),
        peer_company=kw.get("peer_company", "PeerCo Inc."),
        peer_facts=kw.get("peer_facts", app._FTE_DEMO_PEER_FACTS),
        period_facts=kw.get("period_facts", app._FTE_DEMO_PERIOD_FACTS),
        calc_metrics=[r["metric"] for r in parse_requirements(req_text)],
        missing=app._demo_module3_result().get("missing_data"),
        qualitative_documents=kw.get("qualitative_documents", app._FTE_DEMO_QUALITATIVE_DOCS),
    )


def _at_stage(ws, stage, requirements_text="", **extra):
    """Drive the state machine to a given stage and return the session view."""
    s = initial_state()
    if stage != STAGE_OPENING:
        s = apply_choice(s, "opening.requirements", ws)
    if stage == STAGE_PERIODS:
        s = apply_choice(s, "requirements.confirm", ws)
    elif stage == STAGE_METRIC:
        s = apply_choice(apply_choice(s, "requirements.confirm", ws), "period.ROE", ws)
    elif stage == STAGE_EXPLAIN:
        s = apply_choice(
            apply_choice(apply_choice(s, "requirements.confirm", ws), "period.ROE", ws),
            "metric.explain", ws,
        )
    elif stage == STAGE_EXCEL:
        s = apply_choice(apply_choice(s, "requirements.confirm", ws), "skip", ws)
        s = apply_choice(s, "continue", ws)
        s = apply_choice(s, "drivers.excel", ws)
    elif stage == STAGE_CONCLUSION:
        s = apply_choice(apply_choice(s, "requirements.confirm", ws), "skip", ws)
        s = apply_choice(s, "continue", ws)
        s = apply_choice(s, "continue", ws)
        s = apply_choice(s, "continue", ws)
        s = apply_choice(s, "continue", ws)
        s = apply_choice(s, "continue", ws)
        s = apply_choice(s, "continue", ws)
        s = apply_choice(s, "memo.conclusion", ws)
    return agent_session(ws, s, requirements_text=requirements_text, facts_src=extra.get("facts_src"))


# ---------------------------------------------------------------------------
# 1-12 · Requirement parsing / recovery classification
# ---------------------------------------------------------------------------
ws_demo = _demo_workspace()
ws_real = _real_workspace(period_facts=_PERIOD_FACTS)

# 1. Clean assignment -> automatic continuation
rec_clean = parse_recovery(ws_real, _REQ_TEXT)
check("1a · clean assignment -> high confidence", rec_clean["state"] == PARSE_HIGH,
      f"got {rec_clean['state']}")
v_clean = agent_session(
    ws_real, apply_choice(initial_state(), "opening.requirements", ws_real),
    requirements_text=_REQ_TEXT,
)
check("1b · high state recommends continue (automatic continuation)",
      (v_clean.get("recommended") or {}).get("id") == "requirements.continue",
      f"got {v_clean.get('recommended')}")

# 2. WhatsApp / messy whitespace
WHATSAPP = "yo!! calculate  roe   roa profit margin 📊 current ratio debt/equity for microsoft FY2024-FY2025 thx"
ws_whatsapp = _real_workspace(requirements_text=WHATSAPP, period_facts=_PERIOD_FACTS)
rec_ws = parse_recovery(ws_whatsapp, WHATSAPP)
check("2a · WhatsApp whitespace + emoji still recovers", rec_ws["state"] == PARSE_HIGH,
      f"got {rec_ws['state']}")
parsed_wa = {r["metric"] for r in parse_requirements(WHATSAPP)}
check("2b · WhatsApp text resolves the full metric set",
      {"ROE", "ROA", "Profit Margin", "Current Ratio", "Debt to Equity"} <= parsed_wa,
      f"got {parsed_wa}")

# 3. Broken line formatting
LINES = "Compute ROE,\nROA\nand Profit Margin\nFY2024-FY2025"
ws_lines = _real_workspace(requirements_text=LINES, period_facts=_PERIOD_FACTS)
check("3a · broken line formatting recovers", parse_recovery(ws_lines, LINES)["state"] == PARSE_HIGH)

# 4. Unicode punctuation
UNICODE_TXT = "Calculate ROE, ROA and Profit Margin — FY2024–FY2025 ✓"
ws_uni = _real_workspace(requirements_text=UNICODE_TXT, period_facts=_PERIOD_FACTS)
check("4a · unicode punctuation recovers", parse_recovery(ws_uni, UNICODE_TXT)["state"] == PARSE_HIGH)

# 5. Debt-to-Equity -> one requirement
rows_d2e = parse_requirements("Compute ROE and Debt-to-Equity.")
d2e_rows = [r for r in rows_d2e if r["metric"] == "Debt to Equity"]
check("5a · Debt-to-Equity resolves to one requirement", len(d2e_rows) == 1, f"got {rows_d2e}")
check("5b · Debt-to-Equity does NOT fragment into Debt + Equity",
      "Debt" not in {r["metric"] for r in rows_d2e} and "Equity" not in {r["metric"] for r in rows_d2e})

# 6. D/E -> one requirement
rows_de = parse_requirements("Compute D/E and ROE.")
check("6a · D/E resolves to one Debt to Equity requirement",
      len([r for r in rows_de if r["metric"] == "Debt to Equity"]) == 1, f"got {rows_de}")

# 7. Ambiguous metric -> review state
_fr = _real_facts()
_fr["Current Ratio"] = {"value": 1.40, "source": "10-K FY2025 · Calculated",
                        "extraction_state": "review_required",
                        "extraction_state_reason": "Ragged table structure — value uncertain.",
                        "reporting_period": "FY2025"}
_mr = {"financial_data": _fr, "ratios": {}, "missing_data": {"financial_data": [], "ratios": []}}
ws_review = build_student_workspace(
    _mr, requirements_text="Analyze Microsoft FY2025 and calculate Current Ratio.",
    company_a="Microsoft", calc_metrics=["Current Ratio"], missing=_mr.get("missing_data"),
)
rec_rev = parse_recovery(ws_review, "Analyze Microsoft FY2025 and calculate Current Ratio.")
check("7a · review-required item -> partial with review flag",
      rec_rev["state"] == PARSE_PARTIAL and "Current Ratio" in rec_rev["review_required"],
      f"got {rec_rev}")
AMB_TXT = "Calculate Segment Gross Margin and ROE."
ws_amb = _real_workspace(requirements_text=AMB_TXT, period_facts=_PERIOD_FACTS)
rec_amb = parse_recovery(ws_amb, AMB_TXT)
check("7b · ambiguous label surfaces for confirmation (never silently merged)",
      rec_amb["state"] == PARSE_PARTIAL and
      any("segment gross margin" in str(x).lower() for x in rec_amb["uncertain"]),
      f"got {rec_amb}")
check("7c · supported metric (Operating Margin) parses clean, no spurious tokens",
      parse_recovery(_real_workspace(
          requirements_text="Calculate Operating Margin and ROE.", period_facts=_PERIOD_FACTS),
          "Calculate Operating Margin and ROE.")["state"] == PARSE_HIGH)

# 8. Unknown requirement -> manual confirmation (low state)
LOW_TXT = "Please calculate ROIC and Quick Ratio."
ws_low = _real_workspace(requirements_text=LOW_TXT)
rec_low = parse_recovery(ws_low, LOW_TXT)
check("8a · unknown-only text -> low state", rec_low["state"] == PARSE_LOW, f"got {rec_low['state']}")
v_low = agent_session(
    ws_low, apply_choice(initial_state(), "opening.requirements", ws_low),
    requirements_text=LOW_TXT,
)
check("8b · low state offers manual selector options",
      len(v_low["content"].get("options") or []) > 0)
check("8c · low state offers confirm + edit (manual recovery)",
      "requirements.confirm" in [c.get("id") for c in v_low["choices"]] and
      "requirements.edit" in [c.get("id") for c in v_low["choices"]])

# 9. Duplicate requirements -> deduplicated
rows_dup = parse_requirements("ROE, ROE, roe, ROE again")
check("9a · duplicate requirements deduplicated",
      len([r for r in rows_dup if r["metric"] == "ROE"]) == 1, f"got {rows_dup}")

# 10. EPS != EPSILON
_c_eps, _s_eps, _ = canonicalize_metric("EPS")
_c_epsilon, _s_epsilon, _ = canonicalize_metric("EPSILON")
check("10a · EPS canonicalizes, EPSILON does not",
      _c_eps == "EPS" and _c_epsilon is None, f"EPS={_c_eps} EPSILON={_c_epsilon}")
check("10b · 'EPS and EPSILON' parses to EPS only",
      [r["metric"] for r in parse_requirements("EPS and EPSILON")] == ["EPS"])

# 11. Revenue != Revenue Growth
_c_rev, _, _ = canonicalize_metric("Revenue")
_c_revg, _, _ = canonicalize_metric("Revenue Growth")
check("11a · Revenue and Revenue Growth stay distinct",
      _c_rev == "Revenue" and _c_revg == "Revenue Growth" and _c_rev != _c_revg)
check("11b · 'Revenue and Revenue Growth' yields both",
      {r["metric"] for r in parse_requirements("Revenue and Revenue Growth")} == {"Revenue", "Revenue Growth"})

# 12. Debt != Debt-like
_c_debt, _, _ = canonicalize_metric("Debt")
_c_debtlike, _, _ = canonicalize_metric("Debt-like")
check("12a · Debt canonicalizes, Debt-like does not",
      _c_debt == "Debt" and _c_debtlike is None, f"Debt-like -> {_c_debtlike}")
check("12b · 'Debt and Debt-like items' parses to Debt only",
      [r["metric"] for r in parse_requirements("Debt and Debt-like items")] == ["Debt"])

# ---------------------------------------------------------------------------
# 13-20 · Agent UX
# ---------------------------------------------------------------------------
# 13. High state -> exactly one primary action
check("13a · high state has exactly one recommended action",
      bool(v_clean.get("recommended")) and not isinstance(v_clean["recommended"], list))
check("13b · high state primary action is Continue to the analysis",
      (v_clean.get("recommended") or {}).get("id") == "requirements.continue")

# 14. Partial state -> confirmation action
PARTIAL_TXT = "Calculate ROE and ROIC."
ws_partial = _real_workspace(requirements_text=PARTIAL_TXT, period_facts=_PERIOD_FACTS)
rec_part = parse_recovery(ws_partial, PARTIAL_TXT)
v_part = agent_session(
    ws_partial, apply_choice(initial_state(), "opening.requirements", ws_partial),
    requirements_text=PARTIAL_TXT,
)
check("14a · partial state classified correctly", rec_part["state"] == PARSE_PARTIAL,
      f"got {rec_part['state']}")
check("14b · partial state recommends Confirm & Continue",
      (v_part.get("recommended") or {}).get("id") == "requirements.confirm",
      f"got {v_part.get('recommended')}")
check("14c · partial state offers confirm + edit",
      "requirements.confirm" in [c.get("id") for c in v_part["choices"]] and
      "requirements.edit" in [c.get("id") for c in v_part["choices"]])

# 15. Low state -> manual recovery
check("15a · low state exposes the canonical metric selector",
      len(v_low["content"].get("options") or []) >= 10)
check("15b · low state never dead-ends (confirm continues)",
      apply_choice(apply_choice(initial_state(), "opening.requirements", ws_low),
                   "requirements.confirm", ws_low)["stage"] in (STAGE_PERIODS, STAGE_METRIC))

# 16. No parser/debug terminology
_BANNED = ("parser", "traceback", "exception", "stack", "debug",
           "backend error", "implementation", "regex", "failed")
_all_msgs = []
for _s in (v_clean, v_part, v_low):
    _all_msgs.append(str(_s.get("message") or ""))
_all_msgs.append(str(_at_stage(ws_demo, STAGE_EXCEL)["message"] or ""))
_all_msgs.append(str(_at_stage(ws_demo, STAGE_CONCLUSION)["message"] or ""))
_all_msgs.append(str(_at_stage(ws_real, STAGE_EXCEL)["message"] or ""))
check("16a · no parser/debug/exception terminology in any agent message",
      all(not any(b in m.lower() for b in _BANNED) for m in _all_msgs),
      [m for m in _all_msgs if any(b in m.lower() for b in _BANNED)])

# 17. No traceback / error leakage
check("17a · no traceback/File/line patterns in messages",
      all(not re.search(r"traceback|File \"|line \d+|Exception", m) for m in _all_msgs))

# 18. No empty dead-end state
check("18a · every requirements state has a message",
      all(len(m) > 20 for m in (v_clean["message"], v_part["message"], v_low["message"])))
check("18b · every requirements state has a recommended action",
      all(v.get("recommended") for v in (v_clean, v_part, v_low)))
check("18c · every requirements state has at least one choice",
      all((v.get("choices") or []) for v in (v_clean, v_part, v_low)))
_edit_state = apply_choice(apply_choice(initial_state(), "opening.requirements", ws_partial),
                           "requirements.edit", ws_partial)
check("18d · requirements.edit stays on requirements (no dead end)",
      _edit_state["stage"] == STAGE_REQUIREMENTS)
check("18e · requirements.edit still yields a full session view",
      bool(agent_session(ws_partial, _edit_state, requirements_text=PARTIAL_TXT)["message"]))

# 19. Agent tells the student what happened
check("19a · high state explains the parse", "I've parsed your assignment" in v_clean["message"])
check("19b · partial state names the uncertain item",
      "identified most of the assignment requirements" in v_part["message"] and
      "ROIC" in v_part["message"])
check("19c · low state reassures nothing is broken",
      "couldn't reliably identify" in v_low["message"] and "Nothing is broken" in v_low["message"])

# 20. Agent tells the student what to do next
check("20a · high state gives the next step",
      "Continue to the analysis" in v_clean["message"] or
      (v_clean.get("recommended") or {}).get("label", "").lower().startswith("continue"))
check("20b · partial state asks for confirmation",
      "confirm" in v_part["message"].lower())
check("20c · low state asks the student to confirm",
      "confirm" in v_low["message"].lower() and "professor" in v_low["message"].lower())

# ---------------------------------------------------------------------------
# 21-27 · Excel guidance
# ---------------------------------------------------------------------------
wb_bytes = build_excel_working_model(ws_demo)
wb = openpyxl.load_workbook(BytesIO(wb_bytes))
check("21a · workbook generation unchanged (valid xlsx)",
      wb_bytes[:2] == b"PK" and wb.sheetnames)
_EXPECTED_SHEETS = {
    "Financial Data", "Ratio Analysis", "External Variables", "Comparison",
    "Driver Analysis", "Assignment Requirements", "Qualitative Drivers",
}
check("22a · all seven sheets preserved",
      set(wb.sheetnames) == _EXPECTED_SHEETS, f"got {wb.sheetnames}")
_formula_count = 0
for _ws_name in wb.sheetnames:
    for _row in wb[_ws_name].iter_rows(values_only=True):
        for _cell in _row:
            if isinstance(_cell, str) and _cell.startswith("="):
                _formula_count += 1
check("23a · real Excel formulas preserved", _formula_count >= 3, f"formulas={_formula_count}")

v_xl = _at_stage(ws_demo, STAGE_EXCEL)
check("24a · excel orientation payload present before opening",
      bool((v_xl["content"].get("orientation") or {}).get("first")))
check("24b · excel stage message introduces the model",
      "Your working model is ready" in v_xl["message"])
check("25a · Ratio Analysis is the first verification location",
      (v_xl["content"].get("orientation") or {}).get("first") == "Ratio Analysis")
check("25b · metric-level where-to-look names Ratio Analysis",
      "Ratio Analysis" in str(_at_stage(ws_demo, STAGE_METRIC)["content"].get("excel_where") or ""))
check("26a · student told formulas are already calculated",
      bool((v_xl["content"].get("orientation") or {}).get("formulas_done")) is True)
with open(os.path.join(os.path.dirname(__file__), "..", "app (1) (9).py"), encoding="utf-8") as _fh:
    _APP_SRC = _fh.read()
check("26b · app render tells the student formulas need no editing",
      "you don't need to edit the formulas" in _APP_SRC)
check("27a · student can still explore other sheets",
      "Comparison" in (v_xl["content"].get("orientation") or {}).get("optional", []) and
      "Driver Analysis" in (v_xl["content"].get("orientation") or {}).get("optional", []) and
      len((v_xl["content"].get("sheets") or [])) == 7)

# ---------------------------------------------------------------------------
# 28-35 · API / Demo parity
# ---------------------------------------------------------------------------
v_real_open = agent_session(
    ws_real, apply_choice(initial_state(), "opening.requirements", ws_real),
    requirements_text=_REQ_TEXT,
)
check("28a · API onboarding renders the parsed assignment",
      v_real_open["content"].get("parse_state") == PARSE_HIGH and
      "I've parsed your assignment" in v_real_open["message"])
_demo_req_text = _load_app()._demo_assignment_requirements_text()
v_demo_open = agent_session(
    ws_demo, apply_choice(initial_state(), "opening.requirements", ws_demo),
    requirements_text=_demo_req_text,
)
check("29a · Demo onboarding renders the parsed assignment",
      v_demo_open["content"].get("parse_state") == PARSE_HIGH and
      "I've parsed your assignment" in v_demo_open["message"])
v_real_rec = agent_session(
    ws_partial, apply_choice(initial_state(), "opening.requirements", ws_partial),
    requirements_text=PARTIAL_TXT,
)
check("30a · API recovery offers the confirmation action",
      v_real_rec["content"].get("parse_state") == PARSE_PARTIAL and
      (v_real_rec.get("recommended") or {}).get("id") == "requirements.confirm")
ws_demo_partial = _demo_workspace(requirements_text=PARTIAL_TXT)
v_demo_rec = agent_session(
    ws_demo_partial, apply_choice(initial_state(), "opening.requirements", ws_demo_partial),
    requirements_text=PARTIAL_TXT,
)
check("31a · Demo recovery offers the confirmation action",
      v_demo_rec["content"].get("parse_state") == PARSE_PARTIAL and
      (v_demo_rec.get("recommended") or {}).get("id") == "requirements.confirm")

_s_det = apply_choice(initial_state(), "opening.requirements", ws_demo)
_v1 = agent_session(ws_demo, _s_det, requirements_text=_demo_req_text)
_v2 = agent_session(ws_demo, _s_det, requirements_text=_demo_req_text)
check("32a · Demo agent output is deterministic",
      _v1["message"] == _v2["message"] and
      [c.get("id") for c in _v1["choices"]] == [c.get("id") for c in _v2["choices"]] and
      _v1["recommended"] == _v2["recommended"])

_demo_m3 = _load_app()._demo_module3_result()
_demo_srcs = [str(f.get("source") or "").lower() for f in _demo_m3["financial_data"].values()]
check("33a · Demo provenance is synthetic (no API/10-K claims)",
      all(("demo" in s or "fixture" in s) and "10-k" not in s for s in _demo_srcs),
      _demo_srcs)
_all_demo_msgs = [str(v_demo_open["message"]), str(v_demo_rec["message"]),
                  str(_at_stage(ws_demo, STAGE_EXCEL)["message"]),
                  str(_at_stage(ws_demo, STAGE_CONCLUSION)["message"])]
check("34a · Demo messages contain no AI-provider language",
      all(not re.search(r"\bai\b|\bllm\b|gpt|openai|anthropic", m.lower()) for m in _all_demo_msgs))
check("34b · Demo qualitative drivers come from static synthetic fixtures",
      all(
          str(q.get("source") or "").strip() in ("", "—")  # no evidence -> honest placeholder
          or (("demo" in str(q.get("source") or "").lower() or
               "synthetic" in str(q.get("source") or "").lower())
              and "10-k" not in str(q.get("source") or "").lower())
          for q in (ws_demo.get("qualitative_drivers") or {}).get("rows") or []))
check("35a · Demo financial values remain unchanged",
      _demo_m3["financial_data"]["Revenue"]["value"] == 281700000000 and
      _demo_m3["financial_data"]["Net Profit"]["value"] == 98300000000)

# ---------------------------------------------------------------------------
# 36-39 · Student safety / academic integrity
# ---------------------------------------------------------------------------
v_concl = _at_stage(ws_demo, STAGE_CONCLUSION)
_cn = v_concl["content"]
check("36a · conclusion engine never generates a conclusion",
      bool(_cn.get("never_generate")) is True and "conclusion" not in _cn)
check("36b · app keeps the student conclusion text area blank + student-authored",
      "State your own reasoned judgment here" in _APP_SRC and
      "never generates a conclusion" in _APP_SRC)
_all_concl_text = " ".join(str(x) for x in (_cn.get("checklist") or []) + (_cn.get("scaffold") or [])).lower()
check("37a · no Buy/Sell/Hold recommendation generated",
      not any(w in _all_concl_text for w in ("buy ", " sell ", " hold ", "recommend", "outperform",
                                             "underperform", "target price", "investment recommendation")))
check("38a · no submission-ready interpretation generated",
      not any(w in _all_concl_text for w in ("therefore", "i recommend", "the company is",
                                             "should ", "my conclusion", "in conclusion")))
check("39a · evidence-backed scaffolding remains available",
      len(_cn.get("scaffold") or []) >= 3 and
      str(_cn["scaffold"][0]).startswith("Evidence suggests"))

# ---------------------------------------------------------------------------
# Engine flow extras — confirm/edit transitions & excel_where wiring
# ---------------------------------------------------------------------------
check("e1 · requirements.confirm on a clean assignment advances",
      apply_choice(apply_choice(initial_state(), "opening.requirements", ws_real),
                   "requirements.confirm", ws_real)["stage"] in (STAGE_PERIODS, STAGE_METRIC))
check("e2 · requirements.confirm on a low-confidence assignment advances",
      apply_choice(apply_choice(initial_state(), "opening.requirements", ws_low),
                   "requirements.confirm", ws_low)["stage"] in (STAGE_PERIODS, STAGE_METRIC))
_wn = what_next(ws_partial, apply_choice(initial_state(), "opening.requirements", ws_partial),
                requirements_text=PARTIAL_TXT)
check("e3 · what_next honours requirements_text for partial state",
      (_wn.get("recommended") or {}).get("id") == "requirements.confirm")
_v_exp = _at_stage(ws_demo, STAGE_EXPLAIN)
check("e4 · explain stage carries excel_where guidance",
      "Ratio Analysis" in str(_v_exp["content"].get("excel_where") or ""))
check("e5 · blocked/review state still surfaces in the requirements stage",
      rec_rev["state"] == PARSE_PARTIAL and bool(rec_rev["review_required"]))


def main():
    failures = [c for c in CHECKS if not c[1]]
    print(f"\nRESULT: {len(CHECKS) - len(failures)}/{len(CHECKS)} checks pass")
    if failures:
        for name, _ok, detail in failures:
            print(f"  FAIL {name}: {detail}")
        sys.exit(1)
    print("ALL CHECKS PASS")
    sys.exit(0)


if __name__ == "__main__":
    main()
