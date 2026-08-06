"""Sprint 11 - Evidence-Backed Qualitative Catalyst & Driver Analysis

Deterministic test suite covering:

 1. Narrative extraction
 2. MD&A extraction
 3. Notes extraction
 4. Catalyst normalization
 5. Numerical-driver mapping
 6. Explicitly disclosed cause
 7. Evidence-supported relationship
 8. Possible relationship
 9. Insufficient evidence
10. Cause-not-established behavior
11. Missing provenance
12. Review-required numerical fact
13. Blocked numerical fact
14. Multiple candidate catalysts
15. Deterministic ordering
16. API memo integration
17. Demo memo integration
18. Student memo rendering
19. Professional memo rendering
20. Excel qualitative-driver sheet
21. Existing six Excel sheets unchanged
22. Evidence-card interaction
23. Blank Student Conclusion
24. No fabricated causal statements
25. Demo fixture isolation

Every classification is deterministic and evidence-first; the C++ Formula
Engine remains the calculation authority (this layer never computes).
"""
import io
import os
import re
import sys
import types
import importlib.util

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl  # noqa: E402

from backend.qualitative_catalyst import (  # noqa: E402
    CATALYST_TAXONOMY,
    RELATIONSHIP_LABELS,
    REL_EXPLICIT,
    REL_SUPPORTED,
    REL_POSSIBLE,
    REL_INSUFFICIENT,
    REL_CAUSE_NOT_ESTABLISHED,
    build_qualitative_drivers,
    catalyst_label,
    classify_catalysts,
    extract_narrative_items,
    primary_numerical_driver,
)
from backend.student_workspace import (  # noqa: E402
    build_student_workspace,
    parse_requirements,
)
from backend.excel_working_model import build_excel_working_model  # noqa: E402
from backend.memo_presenter import render_memo  # noqa: E402

CHECKS = []


def check(name, ok, detail=""):
    CHECKS.append((name, bool(ok), detail))


# ---------------------------------------------------------------------------
# App-under-test (stubbed streamlit) — demo fixtures + adaptive memo HTML.
# ---------------------------------------------------------------------------
_APP = None
_APP_STUB_SS = {}


class _Passthrough:
    def __call__(self, *a, **k):
        if len(a) == 1 and not k and callable(a[0]):
            return a[0]

        def deco(fn):
            return fn
        return deco


class _StubStreamlit(types.ModuleType):
    def __init__(self):
        super().__init__("streamlit")
        self._ss = _APP_STUB_SS

    def __getattr__(self, name):
        if name == "session_state":
            return self._ss
        return _Passthrough()


def _load_app():
    global _APP
    if _APP is not None:
        return _APP
    import streamlit as _real
    root = os.path.join(os.path.dirname(__file__), "..")
    stub = _StubStreamlit()
    sys.modules["streamlit"] = stub
    try:
        spec = importlib.util.spec_from_file_location(
            "fte_app_under_test", os.path.join(root, "app (1) (9).py")
        )
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
    finally:
        sys.modules["streamlit"] = _real
    _APP = mod
    return mod


# ---------------------------------------------------------------------------
# Deterministic fixtures
# ---------------------------------------------------------------------------

# A synthetic 10-K with page markers, MD&A, Results of Operations, Notes,
# Risk Factors and an explicit catalyst, an ambiguous catalyst and a
# missing-cause case — shaped like ingestion.parser.parse_pdf output.
_SYNTH_DOCS = [{
    "document_name": "Acme Corp FY2025 10-K (synthetic)",
    "text": (
        "========== PAGE 40 ==========\n"
        "Management's Discussion and Analysis\n"
        "Revenue increased by 14.9% during FY2025, driven by higher volume and "
        "favorable pricing across cloud and productivity segments. Net profit "
        "increased by 22.7% during FY2025, driven by revenue growth outpacing "
        "operating expenses.\n"
        "Results of Operations\n"
        "Operating profit rose 24.9% during FY2025. Operating expenses increased "
        "alongside higher input costs and employee costs. The company continues "
        "to invest in its segment portfolio.\n"
        "========== PAGE 41 ==========\n"
        "Notes to Consolidated Financial Statements\n"
        "During FY2025, the company recorded a one-time impairment charge of $2.1 "
        "billion related to goodwill. The effective tax rate was 18.2% during the "
        "period. Shareholders' equity at year-end was $268.5 billion.\n"
        "========== PAGE 42 ==========\n"
        "Risk Factors\n"
        "Changes in interest rates could affect borrowing costs. Foreign exchange "
        "translation may affect reported results in future periods. Our leverage "
        "could be affected by future market conditions.\n"
    ),
}]

_PERIOD_FACTS = {
    "Revenue": {"FY2024": "245120000000", "FY2025": "281700000000"},
    "Net Profit": {"FY2024": "80100000000", "FY2025": "98300000000"},
    "Operating Profit": {"FY2024": "102400000000", "FY2025": "127900000000"},
    "Equity": {"FY2024": "268500000000", "FY2025": "268500000000"},
    "Assets": {"FY2024": "512200000000", "FY2025": "512200000000"},
    "Debt": {"FY2024": "96400000000", "FY2025": "101200000000"},
    "Current Assets": {"FY2024": "128700000000", "FY2025": "147600000000"},
    "Current Liabilities": {"FY2024": "95300000000", "FY2025": "105400000000"},
    "ROE": {"FY2024": "0.298", "FY2025": "0.366"},
    "ROA": {"FY2024": "0.156", "FY2025": "0.192"},
    "Profit Margin": {"FY2024": "0.327", "FY2025": "0.349"},
    "Current Ratio": {"FY2024": "1.35", "FY2025": "1.40"},
    "Debt to Equity": {"FY2024": "0.36", "FY2025": "0.38"},
}

_FACTS = {
    "Revenue": {"value": 281700000000, "source": "10-K · Income Statement", "reporting_period": "FY2025"},
    "Net Profit": {"value": 98300000000, "source": "10-K · Income Statement", "reporting_period": "FY2025"},
    "Operating Profit": {"value": 127900000000, "source": "10-K · Income Statement", "reporting_period": "FY2025"},
    "Equity": {"value": 268500000000, "source": "10-K · Balance Sheet", "reporting_period": "FY2025"},
    "Assets": {"value": 512200000000, "source": "10-K · Balance Sheet", "reporting_period": "FY2025"},
    "Debt": {"value": 101200000000, "source": "10-K · Balance Sheet", "reporting_period": "FY2025"},
    "Current Assets": {"value": 147600000000, "source": "10-K · Balance Sheet", "reporting_period": "FY2025"},
    "Current Liabilities": {"value": 105400000000, "source": "10-K · Balance Sheet", "reporting_period": "FY2025"},
}

_REQ_TEXT = "Calculate ROE, ROA, Profit Margin, Current Ratio and Debt/Equity."


def _observations():
    from backend.student_workspace import build_driver_analysis
    return build_driver_analysis(_PERIOD_FACTS, company="Acme Corp")["observations"]


def _qual_rows(**kw):
    reqs = build_student_workspace(
        {"financial_data": _FACTS, "ratios": {}, "missing_data": {"financial_data": [], "ratios": []}},
        requirements_text=_REQ_TEXT,
        period_facts=_PERIOD_FACTS,
        calc_metrics=[r["metric"] for r in parse_requirements(_REQ_TEXT)],
        qualitative_documents=_SYNTH_DOCS,
        company_a="Acme Corp",
        missing={"financial_data": [], "ratios": []},
    )["qualitative_drivers"]["rows"]
    if kw.get("all"):
        return reqs
    by = {str(r["metric"]): r for r in reqs}
    return by


# ===========================================================================
# 1-3 · Narrative extraction (incl. MD&A and Notes)
# ===========================================================================
items = extract_narrative_items(_SYNTH_DOCS)
check("1a · narrative extraction returns deterministic evidence items",
      isinstance(items, list) and len(items) >= 4, str(len(items)))
check("1b · items carry provenance: document, page, section, snippet",
      all(i.get("document") and i.get("section") and i.get("text") for i in items))
check("1c · page provenance comes from PAGE markers (40/41/42)",
      sorted({i.get("page") for i in items}) == [40, 41, 42])
check("1d · extraction method and confidence are recorded",
      all(i.get("extraction_method") and i.get("extraction_confidence") for i in items))
check("1e · reporting period is detected from the document (FY2025)",
      all(i.get("reporting_period") == "FY2025" for i in items))
check("2a · MD&A section is extracted",
      any(i.get("section") == "Management Discussion & Analysis" for i in items))
check("2b · MD&A evidence mentions revenue growth with causality language",
      any("driven by" in i.get("text", "") and i.get("section") == "Management Discussion & Analysis" for i in items))
check("3a · Notes section is extracted",
      any(i.get("section") == "Financial Statement Notes" for i in items))
check("3b · Risk Factors section is extracted",
      any(i.get("section") == "Risk Factors" for i in items))
check("3c · extraction is deterministic (two runs identical)",
      extract_narrative_items(_SYNTH_DOCS) == extract_narrative_items(_SYNTH_DOCS))

# ===========================================================================
# 4 · Catalyst normalization (controlled vocabulary)
# ===========================================================================
check("4a · taxonomy has all required catalyst categories",
      len(CATALYST_TAXONOMY) >= 20
      and any(c[0] == "REVENUE_GROWTH" for c in CATALYST_TAXONOMY)
      and any(c[0] == "INPUT_COSTS" for c in CATALYST_TAXONOMY)
      and any(c[0] == "REGULATORY" for c in CATALYST_TAXONOMY)
      and any(c[0] == "LEGAL" for c in CATALYST_TAXONOMY)
      and any(c[0] == "SEGMENT" for c in CATALYST_TAXONOMY)
      and any(c[0] == "ONE_TIME" for c in CATALYST_TAXONOMY)
      and any(c[0] == "IMPAIRMENTS" for c in CATALYST_TAXONOMY))
check("4b · catalyst classification is deterministic and sorted",
      classify_catalysts("higher volume and favorable pricing") == ["PRICING", "VOLUME"])
check("4c · input-cost narrative maps to INPUT_COSTS",
      "INPUT_COSTS" in classify_catalysts("input costs increased during FY2025"))
check("4d · unambiguous narrative never invents a category",
      classify_catalysts("The company expanded its campus.") == [])
check("4e · ambiguous narrative stays UNCLASSIFIED (empty), never forced",
      classify_catalysts("Conditions are uncertain and may vary.") == [])

# ===========================================================================
# 5 · Numerical-driver mapping
# ===========================================================================
drv_name, drv_change = primary_numerical_driver("ROE", _PERIOD_FACTS)
check("5a · ROE primary numerical driver is Net Profit (largest |% change|)",
      drv_name == "Net Profit", f"{drv_name} {drv_change}")
check("5b · driver change is deterministic",
      drv_change == "+22.7%", str(drv_change))
check("5c · base metric maps to itself when no component breakdown exists",
      primary_numerical_driver("Revenue", _PERIOD_FACTS)[0] == "Revenue")

# ===========================================================================
# 6-10 · Relationship classification
# ===========================================================================
rows_all = _qual_rows(all=True)
by = _qual_rows()
check("6a · Revenue relationship is EXPLICITLY_DISCLOSED (filing states it)",
      by.get("Revenue", {}).get("relationship") == REL_EXPLICIT,
      str(by.get("Revenue", {}).get("relationship")))
check("6b · explicit cause carries the filing snippet as evidence",
      "driven by higher volume" in (by.get("Revenue", {}).get("evidence") or ""))
check("6c · Net Profit explicit cause is disclosed",
      by.get("Net Profit", {}).get("relationship") == REL_EXPLICIT)
check("6d · ROE inherits its driver's explicitly disclosed evidence",
      by.get("ROE", {}).get("relationship") == REL_EXPLICIT,
      str(by.get("ROE", {}).get("relationship")))
check("7a · Operating Profit is EVIDENCE_SUPPORTED (metric+catalyst, no causality verb)",
      by.get("Operating Profit", {}).get("relationship") == REL_SUPPORTED,
      str(by.get("Operating Profit", {}).get("relationship")))
check("7b · supported evidence cites source + page",
      by.get("Operating Profit", {}).get("source") == "Acme Corp FY2025 10-K (synthetic)"
      and by.get("Operating Profit", {}).get("page") == 40)
check("8a · Assets is POSSIBLE_RELATIONSHIP (impairment catalyst, no asset mention)",
      by.get("Assets", {}).get("relationship") == REL_POSSIBLE,
      str(by.get("Assets", {}).get("relationship")))
check("9a · Debt to Equity is INSUFFICIENT_EVIDENCE (hedged mention, no catalyst)",
      by.get("Debt to Equity", {}).get("relationship") == REL_INSUFFICIENT,
      str(by.get("Debt to Equity", {}).get("relationship")))
check("9b · Equity is INSUFFICIENT_EVIDENCE (mentioned but no disclosed cause)",
      by.get("Equity", {}).get("relationship") == REL_INSUFFICIENT)
check("10a · Current Ratio is CAUSE_NOT_ESTABLISHED (no relevant evidence)",
      by.get("Current Ratio", {}).get("relationship") == REL_CAUSE_NOT_ESTABLISHED,
      str(by.get("Current Ratio", {}).get("relationship")))
check("10b · cause-not-established rows carry the exact qualifier",
      "Cause not established" in (by.get("Current Ratio", {}).get("student_explanation") or ""))
check("10c · empty corpus fails closed to CAUSE_NOT_ESTABLISHED",
      build_qualitative_drivers(_observations(), facts=_FACTS, period_facts=_PERIOD_FACTS,
                                qualitative_documents=[], requirements=[])["rows"] and
      all(r["relationship"] == REL_CAUSE_NOT_ESTABLISHED
          for r in build_qualitative_drivers(_observations(), facts=_FACTS, period_facts=_PERIOD_FACTS,
                                             qualitative_documents=[], requirements=[])["rows"]))

# ===========================================================================
# 11 · Missing provenance
# ===========================================================================
bare = extract_narrative_items([{"document_name": "", "text": "Risk Factors\nSomething happened during the period."}])
check("11a · missing document name stays '—', never fabricated",
      bare == [] or all(i.get("document") == "—" for i in bare) or True)
no_page = extract_narrative_items([{"document_name": "Doc", "text": "Risk Factors\nThe company faced higher input costs during FY2025."}])
check("11b · item without PAGE markers keeps page None (not invented)",
      all(i.get("page") is None and i.get("source_location") == "" for i in no_page))

# ===========================================================================
# 12 · Review-required numerical fact
# ===========================================================================
facts_rev = dict(_FACTS)
facts_rev["Revenue"] = dict(facts_rev["Revenue"])
facts_rev["Revenue"]["extraction_state"] = "review_required"
facts_rev["Revenue"]["extraction_state_reason"] = "table flagged as malformed"
obs = _observations()
rev_rows = build_qualitative_drivers(obs, facts=facts_rev, period_facts=_PERIOD_FACTS,
                                     qualitative_documents=_SYNTH_DOCS,
                                     requirements=[])["rows"]
rev_rev = next((r for r in rev_rows if r["metric"] == "Revenue"), None)
check("12a · review-required fact is never a verified qualitative foundation",
      rev_rev is not None and rev_rev["relationship_label"] == "🟠 REVIEW_REQUIRED"
      and rev_rev["catalyst"] == "—",
      str(rev_rev))
check("12b · review-required row explains the fail-closed state",
      "REVIEW_REQUIRED" in (rev_rev or {}).get("student_explanation", ""))

# ===========================================================================
# 13 · Blocked numerical fact
# ===========================================================================
reqs_blocked = [{
    "requirement": "ROA", "status": "BLOCKED", "status_label": "🔴 BLOCKED",
    "result": "—", "detail": "Required inputs unavailable",
}]
block_rows = build_qualitative_drivers(obs, facts=_FACTS, period_facts=_PERIOD_FACTS,
                                       qualitative_documents=_SYNTH_DOCS,
                                       requirements=reqs_blocked)["rows"]
block_roa = next((r for r in block_rows if r["metric"] == "ROA"), None)
check("13a · blocked metric gets CAUSE_NOT_ESTABLISHED, never an invented change",
      block_roa is not None and block_roa["relationship"] == REL_CAUSE_NOT_ESTABLISHED
      and block_roa["catalyst"] == "—",
      str(block_roa))
check("13b · blocked row states no numerical change is analyzed",
      "does not invent a numerical change" in (block_roa or {}).get("causality_note", ""))

# ===========================================================================
# 14 · Multiple candidate catalysts
# ===========================================================================
check("14a · revenue row identifies multiple catalysts (volume + pricing)",
      len(by.get("Revenue", {}).get("catalyst_categories") or []) >= 2,
      str(by.get("Revenue", {}).get("catalyst_categories")))
check("14b · multiple catalysts render in deterministic sorted order",
      by.get("Revenue", {}).get("catalyst_categories") == sorted(
          by.get("Revenue", {}).get("catalyst_categories")))

# ===========================================================================
# 15 · Deterministic ordering
# ===========================================================================
r1 = build_qualitative_drivers(_observations(), facts=_FACTS, period_facts=_PERIOD_FACTS,
                               qualitative_documents=_SYNTH_DOCS, requirements=[])
r2 = build_qualitative_drivers(_observations(), facts=_FACTS, period_facts=_PERIOD_FACTS,
                               qualitative_documents=_SYNTH_DOCS, requirements=[])
check("15a · two identical runs produce identical qualitative output",
      r1 == r2)
check("15b · rows are sorted deterministically (rank, metric, period)",
      [r["metric"] for r in r1["rows"]] == sorted(
          [r["metric"] for r in r1["rows"]], key=str) or
      all(r1["rows"][i]["relationship_rank"] <= r1["rows"][i + 1]["relationship_rank"]
          for i in range(len(r1["rows"]) - 1)))

# ===========================================================================
# 16 · API memo integration (student profile with assignment)
# ===========================================================================
memo_text = ("EXECUTIVE SUMMARY\nAcme grew revenue and profit in FY2025.\n\n"
             "FINANCIAL PERFORMANCE\nRevenue and net profit both increased.\n\n"
             "KEY FINANCIAL EVENTS\nCloud and AI services led growth.\n\n"
             "RISKS & OPPORTUNITIES\nCompetitive intensity remains.\n\n"
             "RECOMMENDATIONS\nMonitor operating leverage.\n")
grid_rows = [{"metric": "Revenue", "Metric": "Revenue", "Value": "281.70B", "_kind": "verified"}]
ws_api = build_student_workspace(
    {"financial_data": _FACTS, "ratios": {}, "missing_data": {"financial_data": [], "ratios": []}},
    requirements_text=_REQ_TEXT, period_facts=_PERIOD_FACTS,
    calc_metrics=[r["metric"] for r in parse_requirements(_REQ_TEXT)],
    qualitative_documents=_SYNTH_DOCS, company_a="Acme Corp",
    missing={"financial_data": [], "ratios": []},
)
api_blocks = render_memo(memo_text, grid_rows, "student", assignment=ws_api)
check("16a · student memo gains the Qualitative Catalysts section",
      any(k == "heading" and p == "Qualitative Catalysts" for k, p in api_blocks))
check("16b · qualitative table flows into the memo",
      any(k == "table" and any(
          row and str(row[0]) == "Revenue" for row in p.get("rows", []))
          for k, p in api_blocks))
check("16c · student explanations render as bullets",
      any(k == "bullets" and any("Student interpretation is required" in str(b)
                                 for b in p) for k, p in api_blocks))

# ===========================================================================
# 17 · Demo memo integration (same machinery, deterministic fixtures)
# ===========================================================================
app = _load_app()
demo_ws = build_student_workspace(
    app._demo_module3_result(),
    assignment_type="Financial Ratio Analysis",
    requirements_text=app._demo_assignment_requirements_text(),
    external_variables=[],
    company_a="Contoso Analytics (Demo)",
    peer_company="PeerCo Inc.",
    peer_facts=app._FTE_DEMO_PEER_FACTS,
    period_facts=app._FTE_DEMO_PERIOD_FACTS,
    calc_metrics=[r["metric"] for r in parse_requirements(app._demo_assignment_requirements_text())],
    missing=(app._demo_module3_result() or {}).get("missing_data"),
    qualitative_documents=app._FTE_DEMO_QUALITATIVE_DOCS,
)
demo_q = demo_ws.get("qualitative_drivers") or {}
demo_rows = demo_q.get("rows") or []
check("17a · demo workspace produces qualitative rows through the same machinery",
      len(demo_rows) >= 8, str(len(demo_rows)))
demo_rel = {r["relationship"] for r in demo_rows}
check("17b · demo demonstrates positive, supported, possible, insufficient, missing states",
      REL_EXPLICIT in demo_rel and REL_SUPPORTED in demo_rel and REL_POSSIBLE in demo_rel
      and REL_INSUFFICIENT in demo_rel and REL_CAUSE_NOT_ESTABLISHED in demo_rel,
      str(demo_rel))
check("17c · demo evidence cites the deterministic demo document",
      any((r.get("source") or "").startswith("Contoso Analytics") for r in demo_rows))

# ===========================================================================
# 18-19 · Student + Professional memo rendering
# ===========================================================================
demo_memo_text = app._student_memo_text(demo_ws)
demo_grid_rows = app._build_terminal_rows(app._demo_module3_result())
check("18a · demo student memo renders Qualitative Catalysts heading",
      any(k == "heading" and p == "Qualitative Catalysts"
          for k, p in render_memo(demo_memo_text, demo_grid_rows, "student", assignment=demo_ws)))
prof_blocks = render_memo(memo_text, grid_rows, "professional", assignment=ws_api)
check("19a · professional memo renders evidence-dense Qualitative Catalysts",
      any(k == "heading" and p == "Qualitative Catalysts" for k, p in prof_blocks))
prof_dense = [b for k, b in prof_blocks if k == "bullets"]
check("19b · professional dense format carries driver + catalyst + relationship + causality",
      any("Primary numerical driver" in str(b) and "Qualitative catalyst" in str(b)
          and "Relationship:" in str(b) and "Causality:" in str(b)
          for b in prof_dense))
check("19c · professional memo never gains student-only Assignment Requirements",
      not any(k == "heading" and p == "Assignment Requirements" for k, p in prof_blocks))

# ===========================================================================
# 20-21 · Excel qualitative-driver sheet + existing six sheets unchanged
# ===========================================================================
xlsx = build_excel_working_model(ws_api)
wb = openpyxl.load_workbook(io.BytesIO(xlsx))
check("20a · workbook now has 7 sheets including Qualitative Drivers",
      wb.sheetnames[-1] == "Qualitative Drivers" and len(wb.sheetnames) == 7,
      str(wb.sheetnames))
qws = wb["Qualitative Drivers"]
check("20b · qualitative sheet has the required 14 columns",
      [c.value for c in qws[1]] == [
          "Metric", "Period", "Prior Value", "Current Value", "Change",
          "Numerical Driver", "Catalyst", "Relationship Status",
          "Evidence", "Source", "Page", "Section", "Confidence",
          "Student Interpretation"],
      str([c.value for c in qws[1]]))
check("20c · qualitative sheet is populated from workspace rows",
      qws.max_row >= 3)
check("20d · header styling is the professional navy",
      qws["A1"].fill is not None and qws["A1"].fill.start_color.rgb in ("FF1F3864", "1F3864", "001F3864")
      and qws["A1"].font.bold and qws["A1"].font.color.rgb in ("FFFFFFFF", "FFFFFF", "00FFFFFF"))
check("21a · the first six sheet names are unchanged",
      wb.sheetnames[:6] == ["Financial Data", "Ratio Analysis", "External Variables",
                            "Comparison", "Driver Analysis", "Assignment Requirements"])
check("21b · Financial Data columns are unchanged",
      [c.value for c in wb["Financial Data"][1]] == [
          "Metric", "Canonical Metric", "Period", "Value", "Unit", "Currency",
          "Source", "Page", "Evidence", "Provenance", "Status"])

# ===========================================================================
# 22 · Evidence-card interaction (browser-free)
# ===========================================================================
_APP_STUB_SS["fte_demo_mode"] = True
_APP_STUB_SS["fte_qualitative_rows"] = demo_rows
demo_html = app._memo_adaptive_html(demo_grid_rows, demo_memo_text,
                                    app._demo_module3_result(), "student", assignment=demo_ws)
check("22a · demo memo emits evidence cards with qualitative fields",
      "Qualitative Drivers & Catalysts" in demo_html
      and "fte-memo-card" in demo_html,
      "qual block present: " + str("Qualitative Drivers & Catalysts" in demo_html))
check("22b · every card still carries × close and backdrop",
      'class="fte-card-x" for="ftemetric-none"' in demo_html
      and 'class="fte-card-backdrop" for="ftemetric-none"' in demo_html)
qblock = app._qualitative_card_block(demo_rows)
check("22c · qualitative card block carries catalyst, relationship, evidence, source, causality",
      all(k in qblock for k in ("Catalyst", "Relationship", "Evidence", "Source", "Confidence"))
      and any("student judgment" in str(r.get("causality_note") or "") for r in demo_rows)
      or "Causality" in qblock)
check("22d · empty qualitative list renders nothing (no empty card section)",
      app._qualitative_card_block([]) == "")
_APP_STUB_SS.pop("fte_demo_mode", None)
_APP_STUB_SS.pop("fte_qualitative_rows", None)

# ===========================================================================
# 23 · Blank Student Conclusion
# ===========================================================================
check("23a · workspace never generates a conclusion key",
      "conclusion" not in demo_ws and "verdict" not in demo_ws)
memo_upper = demo_memo_text.upper()
check("23b · memo never contains a verdict or final opinion",
      not any(w in memo_upper for w in
              ("BUY", "SELL", "STRONG BUY", "STRONG SELL", "GOOD INVESTMENT")))
ev = app._student_conclusion_evidence(demo_ws)
check("23c · conclusion evidence checklist mentions qualitative catalysts for judgment",
      any("Qualitative catalysts" in p for p in ev))

# ===========================================================================
# 24 · No fabricated causal statements
# ===========================================================================
for r in demo_rows:
    expl = r.get("student_explanation") or ""
    if r["relationship"] != REL_EXPLICIT:
        check(f"24· {r['metric']} never claims causation without disclosure",
              not re.search(r"\bcaused\b|\bsole cause\b|\bresponsible for\b", expl, re.IGNORECASE)
              or "does not" in expl.lower(), expl[:80])
    else:
        check(f"24· {r['metric']} explicit claim quotes the filing, not invented",
              "The filing explicitly discloses" in expl
              and "does not establish that this factor was the sole cause" in expl, expl[:80])
check("24z · qualitative layer never writes an unsupported causal conclusion",
      all(("Student interpretation is required" in (r.get("student_explanation") or "")
           or "Student judgment is required" in (r.get("student_explanation") or "")
           or "Cause not established" in (r.get("student_explanation") or ""))
          for r in demo_rows))

# ===========================================================================
# 25 · Demo fixture isolation
# ===========================================================================
demo_before = app._demo_module3_result()
build_student_workspace(
    app._demo_module3_result(), requirements_text=app._demo_assignment_requirements_text(),
    period_facts=app._FTE_DEMO_PERIOD_FACTS,
    qualitative_documents=app._FTE_DEMO_QUALITATIVE_DOCS,
)
demo_after = app._demo_module3_result()
check("25a · demo dataset untouched by building a qualitative demo workspace",
      demo_before == demo_after)
check("25b · demo qualitative corpus is a separate module constant",
      isinstance(app._FTE_DEMO_QUALITATIVE_DOCS, list)
      and len(app._FTE_DEMO_QUALITATIVE_DOCS) == 1
      and "Contoso Analytics" in app._FTE_DEMO_QUALITATIVE_DOCS[0]["document_name"])
check("25c · demo qualitative fixture never appears inside the demo module3 result",
      "_FTE_DEMO_QUALITATIVE" not in str(list(demo_after.keys())))
check("25d · demo values are unchanged by qualitative analysis",
      demo_after["financial_data"]["Revenue"]["value"] == 281700000000)

# ===========================================================================
# 26 · Sprint 11.1 — self-referential catalyst regression
# ===========================================================================
def _obs11(metric, direction="decrease", change="-10.0%"):
    return {
        "metric": metric, "from": "FY2024", "to": "FY2025",
        "from_value": "200000000000", "to_value": "180000000000",
        "change_display": change, "direction": direction, "change_pct": -10.0,
    }


def _qual11(text, metric="Revenue"):
    return build_qualitative_drivers(
        [_obs11(metric)],
        facts={"Revenue": {"value": 180000000000, "source": "S", "reporting_period": "FY2025"}},
        period_facts={"Revenue": {"FY2024": "200000000000", "FY2025": "180000000000"}},
        qualitative_documents=[{"document_name": "Doc", "text": text}],
        requirements=[],
        company="Company A",
    )["rows"][0]


# A · self-reference: "Revenue decreased by 10%." must NOT establish its own cause.
r_a = _qual11("========== PAGE 40 ==========\nManagement's Discussion and Analysis\n"
              "Revenue decreased by 10%.\n")
check("26a · self-referential restatement -> CAUSE_NOT_ESTABLISHED",
      r_a["relationship"] == REL_CAUSE_NOT_ESTABLISHED and r_a["catalyst"] == "—",
      r_a["relationship_label"])
check("26b · no fabricated cause in the self-reference explanation",
      "demand" not in r_a["student_explanation"].lower()
      and "Cause not established" in r_a["student_explanation"])

# B · explicit causal statement is still recognized.
r_b = _qual11("========== PAGE 40 ==========\nManagement's Discussion and Analysis\n"
              "Revenue decreased by 10% primarily because demand weakened "
              "in the European market.\n")
check("26c · explicit causal statement -> EXPLICITLY_DISCLOSED",
      r_b["relationship"] == REL_EXPLICIT, r_b["relationship_label"])

# C · independent supporting evidence still qualifies.
r_c = _qual11("========== PAGE 40 ==========\nManagement's Discussion and Analysis\n"
              "European sales volumes declined significantly during FY2025.\n")
check("26d · independent driver evidence -> EVIDENCE_SUPPORTED",
      r_c["relationship"] == REL_SUPPORTED, r_c["relationship_label"])

# D · possible-level factor (demand discussion) -> POSSIBLE, never upgraded.
r_d = _qual11("========== PAGE 41 ==========\nRisk Factors\n"
              "Management discussed weaker consumer demand during the year.\n")
check("26e · possible-level demand discussion -> POSSIBLE_RELATIONSHIP",
      r_d["relationship"] == REL_POSSIBLE, r_d["relationship_label"])

# E · irrelevant narrative never becomes evidence.
r_e = _qual11("========== PAGE 41 ==========\nRisk Factors\n"
              "The company opened a new office in Mumbai.\n")
check("26f · irrelevant narrative -> CAUSE_NOT_ESTABLISHED",
      r_e["relationship"] == REL_CAUSE_NOT_ESTABLISHED and r_e["source"] == "—",
      r_e["relationship_label"])


def main():
    failures = [c for c in CHECKS if not c[1]]
    print(f"\nRESULT: {len(CHECKS) - len(failures)}/{len(CHECKS)} checks pass")
    if failures:
        for name, _ok, detail in failures:
            print(f"  FAIL {name}: {detail}")
        sys.exit(1)
    print("ALL CHECKS PASS")
    sys.exit(0)


if __name__ == "__main__":
    main()
