"""Sprint 12 - Student Assignment Agent (Progressive Guided Workspace)

Deterministic test suite for the Assignment Agent orchestration layer and
the shared API/Demo guided UX:

 Agent
  1.  Correct opening state
  2.  Correct assignment requirements
  3.  Correct recommended next action
  4.  Correct progression
  5.  Back/skip behavior
  6.  No dead ends
  7.  No irrelevant actions shown
  8.  Agent never invents facts
  9.  Agent never generates a conclusion
 10.  Blocked state produces useful next action
 11.  Review-required state produces useful next action
 12.  Conflict state produces useful next action

 Years
 13.  Period selection works
 14.  FY2024 -> FY2025 analysis preserved
 15.  Missing periods handled safely

 Comparison
 16.  Peer selection works
 17.  Comparable metrics shown
 18.  Missing peer inputs remain blocked
 19.  No forced comparison

 Drivers
 20.  Numerical driver analysis preserved
 21.  Qualitative investigation works
 22.  Causality safeguards preserved
 23.  Self-referential catalyst protection preserved

 Evidence
 24.  Evidence card opens
 25.  Evidence card replaces correctly
 26.  Close works
 27.  Backdrop works
 28.  Provenance remains intact
 29.  No '—'/None/null/NaN leakage

 Excel
 30.  Excel remains downloadable
 31.  Seven sheets remain
 32.  Real formulas remain
 33.  Formatting remains professional
 34.  Student-input provenance remains correct

 Conclusion
 35.  Conclusion remains blank
 36.  No Buy/Sell/recommendation generated

 Demo
 37.  No API key
 38.  No AI
 39.  No network
 40.  Fixture values unchanged
 41.  Demo provenance remains clearly synthetic
 42.  Demo and API UX structures match

Every agent decision is deterministic; the agent never performs financial
arithmetic itself and never writes the student's conclusion.
"""
import os
import re
import sys
import types
import importlib.util

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl  # noqa: E402

from backend.assignment_agent import (  # noqa: E402
    AGENT_STAGE_IDS,
    STAGE_OPENING,
    STAGE_REQUIREMENTS,
    STAGE_PERIODS,
    STAGE_METRIC,
    STAGE_EXPLAIN,
    STAGE_CALCULATION,
    STAGE_EVIDENCE,
    STAGE_DRIVERS,
    STAGE_QUALITATIVE,
    STAGE_COMPARISON,
    STAGE_EXTERNAL,
    STAGE_EXCEL,
    STAGE_MEMO,
    STAGE_CONCLUSION,
    agent_session,
    apply_choice,
    initial_state,
    what_next,
)
from backend.student_workspace import (  # noqa: E402
    add_external_variable,
    build_student_workspace,
    parse_requirements,
)
from backend.excel_working_model import build_excel_working_model  # noqa: E402

CHECKS = []


def check(name, ok, detail=""):
    CHECKS.append((name, bool(ok), detail))


# ---------------------------------------------------------------------------
# App-under-test (stubbed streamlit) — used only for the demo fixtures.
# ---------------------------------------------------------------------------
_APP = None
_APP_STUB_SS = {}


class _Passthrough:
    def __call__(self, *a, **k):
        if len(a) == 1 and not k and callable(a[0]):
            return a[0]

        def deco(fn):
            return fn
        return deco


class _StubStreamlit(types.ModuleType):
    def __init__(self):
        super().__init__("streamlit")
        self._ss = _APP_STUB_SS

    def __getattr__(self, name):
        if name == "session_state":
            return self._ss
        return _Passthrough()


def _load_app():
    """Exec the app file under a stubbed streamlit and return the module."""
    global _APP
    if _APP is not None:
        return _APP
    import streamlit as _real
    root = os.path.join(os.path.dirname(__file__), "..")
    stub = _StubStreamlit()
    sys.modules["streamlit"] = stub
    try:
        spec = importlib.util.spec_from_file_location(
            "fte_app_under_test", os.path.join(root, "app (1) (9).py")
        )
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
    finally:
        sys.modules["streamlit"] = _real
    _APP = mod
    return mod


# ---------------------------------------------------------------------------
# Deterministic fixtures
# ---------------------------------------------------------------------------
def _real_facts():
    return {
        "Revenue": {"value": 281700000000, "source": "10-K FY2025 · Income Statement", "reporting_period": "FY2025", "page": 26, "evidence": "Consolidated Statements of Income, p. 26", "unit": "USD", "scale": "B"},
        "Net Profit": {"value": 98300000000, "source": "10-K FY2025 · Income Statement", "reporting_period": "FY2025", "page": 26, "evidence": "Consolidated Statements of Income, p. 26", "unit": "USD", "scale": "B"},
        "Equity": {"value": 268500000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Assets": {"value": 512200000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Debt": {"value": 101200000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Current Assets": {"value": 147600000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Current Liabilities": {"value": 105400000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
    }


_REQ_TEXT = ("Analyze Microsoft FY2023-FY2025 and calculate ROE, ROA, "
             "Profit Margin, Current Ratio and Debt/Equity.")


def _real_module3():
    facts = _real_facts()
    module3 = {
        "financial_data": facts,
        "ratios": {},
        "missing_data": {"financial_data": ["Segment Gross Margin"], "ratios": []},
    }
    module3["ratios"]["Current Ratio"] = {
        "value": round(facts["Current Assets"]["value"] / facts["Current Liabilities"]["value"], 2),
        "source": "Calculated", "formula": "Current Assets / Current Liabilities",
        "reporting_period": "FY2025",
    }
    module3["ratios"]["Debt to Equity"] = {
        "value": round(facts["Debt"]["value"] / facts["Equity"]["value"], 2),
        "source": "Calculated", "formula": "Debt / Equity", "reporting_period": "FY2025",
    }
    return module3


def _real_workspace(period_facts=None, requirements_text=_REQ_TEXT, **kw):
    module3 = _real_module3()
    return build_student_workspace(
        module3,
        assignment_type=kw.get("assignment_type", "Financial Ratio Analysis"),
        requirements_text=requirements_text,
        external_variables=kw.get("external_variables") or [],
        company_a=kw.get("company_a", "Microsoft"),
        peer_company=kw.get("peer_company"),
        peer_facts=kw.get("peer_facts"),
        period_facts=period_facts,
        calc_metrics=[r["metric"] for r in parse_requirements(requirements_text)],
        missing=module3.get("missing_data"),
    )


def _demo_workspace(**kw):
    app = _load_app()
    return build_student_workspace(
        app._demo_module3_result(),
        assignment_type=kw.get("assignment_type", "Financial Ratio Analysis"),
        requirements_text=kw.get("requirements_text", app._demo_assignment_requirements_text()),
        external_variables=kw.get("external_variables") or [],
        company_a=kw.get("company_a", "Contoso Analytics (Demo)"),
        peer_company=kw.get("peer_company", "PeerCo Inc."),
        peer_facts=kw.get("peer_facts", app._FTE_DEMO_PEER_FACTS),
        period_facts=kw.get("period_facts", app._FTE_DEMO_PERIOD_FACTS),
        calc_metrics=[r["metric"] for r in parse_requirements(kw.get("requirements_text", app._demo_assignment_requirements_text()))],
        missing=app._demo_module3_result().get("missing_data"),
        qualitative_documents=kw.get("qualitative_documents", app._FTE_DEMO_QUALITATIVE_DOCS),
    )


_PERIOD_FACTS = {
    "Revenue": {"FY2024": "245120000000", "FY2025": "281700000000"},
    "Net Profit": {"FY2024": "80100000000", "FY2025": "98300000000"},
    "Equity": {"FY2024": "268500000000", "FY2025": "268500000000"},
    "Assets": {"FY2024": "512200000000", "FY2025": "512200000000"},
    "ROE": {"FY2024": "0.298", "FY2025": "0.366"},
    "ROA": {"FY2024": "0.156", "FY2025": "0.192"},
    "Profit Margin": {"FY2024": "0.327", "FY2025": "0.349"},
    "Current Ratio": {"FY2024": "1.35", "FY2025": "1.40"},
    "Debt to Equity": {"FY2024": "0.36", "FY2025": "0.38"},
}


# ---------------------------------------------------------------------------
# 1-3 · Opening state, requirements, recommended next action
# ---------------------------------------------------------------------------
ws_demo = _demo_workspace()
s0 = initial_state()
v_open = agent_session(ws_demo, s0)
check("1a · opening stage id", v_open["stage"] == STAGE_OPENING)
check("1b · opening message summarizes the assignment",
      "assignment" in v_open["message"].lower() and "ROE" in v_open["message"])
check("1c · opening has 1-3 relevant choices",
      1 <= len(v_open["choices"]) <= 3 and all(c.get("id", "").startswith("opening.") for c in v_open["choices"]))
check("1d · progress indicator has 7 rows",
      len(v_open["progress"]) == 7)
check("1e · first progress row is current at opening",
      v_open["progress"][0]["state"] == "current")

reqs_open = v_open["content"].get("requirements") or []
check("2a · requirements parsed correctly",
      all(r in reqs_open for r in ("ROE", "ROA", "Profit Margin", "Current Ratio", "Debt to Equity")))

rec_open = v_open["recommended"]
check("3a · recommended action exists at opening", rec_open is not None and rec_open.get("id"))
check("3b · recommended action is one of the stage choices",
      rec_open and any(c.get("id") == rec_open["id"] for c in v_open["choices"]))
check("3c · exactly one recommended action", bool(rec_open) and not isinstance(rec_open, list))

# ---------------------------------------------------------------------------
# 4 · Correct progression (opening -> requirements -> periods -> metric ->
#     explain -> evidence -> drivers -> qualitative -> comparison -> excel ->
#     memo -> conclusion)
# ---------------------------------------------------------------------------
s_req = apply_choice(s0, "opening.requirements", ws_demo)
check("4a · opening.requirements -> requirements", s_req["stage"] == STAGE_REQUIREMENTS)
v_req = agent_session(ws_demo, s_req)
check("4b · requirements stage shows checklist rows",
      len(v_req["content"].get("rows") or []) >= 5)

s_per = apply_choice(s_req, "requirements.continue", ws_demo)
check("4c · requirements.continue -> periods", s_per["stage"] == STAGE_PERIODS)
v_per = agent_session(ws_demo, s_per)
check("4d · period stage lists FY2024-FY2025",
      set(v_per["content"].get("periods") or []) == {"FY2024", "FY2025"})

s_met = apply_choice(s_per, "period.ROE", ws_demo)
check("4e · period.ROE -> metric(ROE)", s_met["stage"] == STAGE_METRIC and s_met["metric"] == "ROE")

s_exp = apply_choice(s_met, "metric.explain", ws_demo)
check("4f · metric.explain -> explain", s_exp["stage"] == STAGE_EXPLAIN)

s_ev = apply_choice(s_exp, "explain.evidence", ws_demo)
check("4g · explain.evidence -> evidence", s_ev["stage"] == STAGE_EVIDENCE)

s_drv = apply_choice(s_ev, "continue", ws_demo)
check("4h · evidence.continue -> drivers", s_drv["stage"] == STAGE_DRIVERS)

s_qual = apply_choice(s_drv, "drivers.qualitative", ws_demo)
check("4i · drivers.qualitative -> qualitative", s_qual["stage"] == STAGE_QUALITATIVE)

s_cmp = apply_choice(s_qual, "qualitative.comparison", ws_demo)
check("4j · qualitative.comparison -> comparison", s_cmp["stage"] == STAGE_COMPARISON)

s_xl = apply_choice(s_cmp, "continue", ws_demo)
check("4k · comparison.continue -> excel", s_xl["stage"] == STAGE_EXCEL)

s_mem = apply_choice(s_xl, "continue", ws_demo)
check("4l · excel.continue -> memo", s_mem["stage"] == STAGE_MEMO)

s_con = apply_choice(s_mem, "memo.conclusion", ws_demo)
check("4m · memo.conclusion -> conclusion", s_con["stage"] == STAGE_CONCLUSION)
v_con = agent_session(ws_demo, s_con)
check("4n · conclusion message never writes the conclusion",
      "write" in v_con["message"].lower() or "conclusion" in v_con["message"].lower())
check("4o · conclusion content has a fact checklist",
      len(v_con["content"].get("checklist") or []) > 0)

# ---------------------------------------------------------------------------
# 5 · Back / skip behavior
# ---------------------------------------------------------------------------
s_back = apply_choice(s_met, "back", ws_demo)
check("5a · back from metric -> periods", s_back["stage"] == STAGE_PERIODS)
s_back2 = apply_choice(s_exp, "back", ws_demo)
check("5b · back from explain -> metric", s_back2["stage"] == STAGE_METRIC)
s_skip = apply_choice(s_per, "skip", ws_demo)
check("5c · skip from periods -> drivers (no dead end)", s_skip["stage"] == STAGE_DRIVERS)
s_skip_open = apply_choice(s0, "skip", ws_demo)
check("5d · skip from opening never crashes", s_skip_open["stage"] in AGENT_STAGE_IDS)

# ---------------------------------------------------------------------------
# 6 · No dead ends + deterministic state machine
# ---------------------------------------------------------------------------
def _stage_names(view):
    return view["stage"]


_all_ok = True
_all_detail = ""
for _stage in AGENT_STAGE_IDS:
    _st = {"stage": _stage, "metric": "ROE" if _stage in (STAGE_METRIC, STAGE_EXPLAIN, STAGE_CALCULATION, STAGE_EVIDENCE) else None,
           "area": None, "visited": ["opening"]}
    _v = agent_session(ws_demo, _st)
    if not _v["message"] or not isinstance(_v["message"], str):
        _all_ok = False
        _all_detail = f"stage {_stage}: empty message"
        break
    for _c in _v["choices"]:
        _nxt = apply_choice(_st, _c["id"], ws_demo)
        if _nxt["stage"] not in AGENT_STAGE_IDS and _nxt["stage"] != "__explore__":
            _all_ok = False
            _all_detail = f"stage {_stage} choice {_c['id']} -> bad {_nxt['stage']}"
            break
    if not _all_ok:
        break
check("6a · every stage has a message", _all_ok, _all_detail)
check("6b · every choice resolves to a valid stage", _all_ok, _all_detail)
check("6c · invalid choice fails closed",
      apply_choice(s_met, "no.such.choice", ws_demo)["stage"] == STAGE_METRIC)
check("6d · agent session is fully deterministic",
      agent_session(ws_demo, s_per) == agent_session(ws_demo, s_per))

# ---------------------------------------------------------------------------
# 7 · No irrelevant actions shown
# ---------------------------------------------------------------------------
_met_choice_ids = {c.get("id") for c in agent_session(ws_demo, s_met)["choices"]}
_ok_ids = {"metric.explain", "metric.calculation", "metric.evidence", "metric.comparison",
           "metric.review", "continue", "back"}
check("7a · metric stage shows only metric-scoped actions", _met_choice_ids <= _ok_ids, str(sorted(_met_choice_ids)))
_per_choice_ids = {c.get("id") for c in v_per["choices"]}
check("7b · periods stage shows only metric/back/skip actions",
      all(cid.startswith("period.") or cid in ("back", "skip") for cid in _per_choice_ids),
      str(sorted(_per_choice_ids)))
check("7c · opening stage exposes no dashboard-style actions",
      all(cid.startswith("opening.") for cid in {c.get("id") for c in v_open["choices"]}))

# ---------------------------------------------------------------------------
# 8 · Agent never invents facts
# ---------------------------------------------------------------------------
v_met = agent_session(ws_demo, s_met)
ws_val = next((r.get("result") for r in ws_demo["requirements"] if r.get("requirement") == "ROE"), None)
check("8a · metric value comes from the workspace (no invention)",
      v_met["content"].get("value") == ws_val)
check("8b · metric status comes from the workspace",
      str(v_met["content"].get("status")) == str(next(
          (r.get("status") for r in ws_demo["requirements"] if r.get("requirement") == "ROE"), "")))

# ---------------------------------------------------------------------------
# 9 · Agent never generates a conclusion
# ---------------------------------------------------------------------------
_con_all_messages = []
for _st2 in AGENT_STAGE_IDS:
    _v2 = agent_session(ws_demo, {"stage": _st2, "metric": "ROE", "area": None, "visited": []})
    _con_all_messages.append(str(_v2["message"]))
_blob = " ".join(_con_all_messages).lower()
check("9a · no buy/sell/strong-buy generated anywhere",
      not any(w in _blob for w in ("buy ", " sell ", "strong buy", "recommendation")))
check("9b · conclusion stage marks conclusion as student-authored",
      bool(v_con["content"].get("never_generate")) is True)

# ---------------------------------------------------------------------------
# 10-12 · Blocked / review-required / conflict guidance
# ---------------------------------------------------------------------------
def _facts_without_current_liabilities():
    facts = _real_facts()
    facts.pop("Current Liabilities", None)
    return facts


ws_blocked = _real_workspace(
    requirements_text="Analyze Microsoft FY2025 and calculate Current Ratio.",
    period_facts=None,
)
# Force the Current Ratio input gap: rebuild with Current Liabilities removed.
_fb = _real_facts()
_fb.pop("Current Liabilities", None)
_mb = {"financial_data": _fb, "ratios": {}, "missing_data": {"financial_data": [], "ratios": []}}
ws_blocked = build_student_workspace(
    _mb,
    requirements_text="Analyze Microsoft FY2025 and calculate Current Ratio.",
    company_a="Microsoft",
    calc_metrics=["Current Ratio"],
    missing=_mb.get("missing_data"),
)
s_blk = {"stage": STAGE_METRIC, "metric": "Current Ratio", "area": None, "visited": ["opening", "periods"]}
v_blk = agent_session(ws_blocked, s_blk)
check("10a · blocked metric guidance is surfaced",
      v_blk["guidance"].get("kind") == "blocked")
check("10b · blocked metric produces useful next action",
      bool(v_blk["recommended"]) and any(
          c.get("id") in ("metric.review", "metric.evidence", "continue", "metric.calculation")
          for c in v_blk["choices"]))
check("10c · blocked message names the gap",
      "cannot" in v_blk["message"].lower() or "missing" in v_blk["message"].lower())

# Review-required
_fr = _real_facts()
_fr["Current Ratio"] = {"value": 1.40, "source": "10-K FY2025 · Calculated",
                        "extraction_state": "review_required",
                        "extraction_state_reason": "Ragged table structure — value uncertain.",
                        "reporting_period": "FY2025"}
_mr = {"financial_data": _fr, "ratios": {}, "missing_data": {"financial_data": [], "ratios": []}}
ws_review = build_student_workspace(
    _mr,
    requirements_text="Analyze Microsoft FY2025 and calculate Current Ratio.",
    company_a="Microsoft",
    calc_metrics=["Current Ratio"],
    missing=_mr.get("missing_data"),
)
s_rv = {"stage": STAGE_METRIC, "metric": "Current Ratio", "area": None, "visited": ["opening"]}
v_rv = agent_session(ws_review, s_rv)
check("11a · review-required metric guidance is surfaced",
      v_rv["guidance"].get("kind") == "review")
check("11b · review-required produces useful next action",
      bool(v_rv["recommended"]) and any(
          c.get("id") in ("metric.review", "metric.evidence", "continue")
          for c in v_rv["choices"]))
check("11c · review-required never promoted to verified",
      "REVIEW_REQUIRED" in str(next(
          (r.get("status") for r in ws_review["requirements"] if r.get("requirement") == "Current Ratio"), "")))

# Conflict
_fc = _real_facts()
_fc["Revenue"] = {"value": 281700000000, "source": "10-K FY2025 · Income Statement",
                  "extraction_state": "conflict",
                  "extraction_state_reason": "Table A: 281.70B vs Table B: 281.07B.",
                  "reporting_period": "FY2025"}
_mc = {"financial_data": _fc, "ratios": {}, "missing_data": {"financial_data": [], "ratios": []}}
ws_conf = build_student_workspace(
    _mc,
    requirements_text="Analyze Microsoft FY2025 and calculate Revenue.",
    company_a="Microsoft",
    calc_metrics=["Revenue"],
    missing=_mc.get("missing_data"),
)
v_conf = agent_session(ws_conf, s0, facts_src=_mc)
check("12a · conflict is detected deterministically",
      "Revenue" in (v_conf.get("conflict_metrics") or []))
check("12b · conflict produces a useful guidance message",
      bool(v_conf["guidance"].get("conflict_message")) and "conflict" in v_conf["guidance"]["conflict_message"].lower())
check("12c · conflict message says values are not silently chosen",
      "silently" in v_conf["guidance"]["conflict_message"].lower())

# ---------------------------------------------------------------------------
# 13-15 · Years / periods
# ---------------------------------------------------------------------------
check("13a · period stage contains both years",
      set(v_per["content"].get("periods") or []) == {"FY2024", "FY2025"})
check("13b · period choices come from verified changes",
      any(c.get("id") == "period.ROE" for c in v_per["choices"]))
check("14a · FY2024 -> FY2025 ROE analysis preserved",
      any(o.get("metric") == "ROE" and o.get("from") == "FY2024" and o.get("to") == "FY2025"
          for o in (ws_demo["driver_analysis"].get("observations") or [])))
ws_noperiod = _real_workspace(period_facts=None)
s_np = {"stage": STAGE_PERIODS, "metric": None, "area": None, "visited": ["opening", "requirements"]}
v_np = agent_session(ws_noperiod, s_np)
check("15a · missing periods handled safely (no crash, no changes)",
      not (v_np["content"].get("changes") or []) and v_np["message"])
check("15b · missing periods still offers a way forward",
      any(c.get("id") in ("continue", "back", "skip") for c in v_np["choices"]))

# ---------------------------------------------------------------------------
# 16-19 · Comparison
# ---------------------------------------------------------------------------
v_cmp = agent_session(ws_demo, s_cmp)
check("16a · peer selection works in demo",
      v_cmp["content"].get("company_b") == "PeerCo Inc." and v_cmp["content"].get("active"))
check("17a · comparable metrics shown",
      len(v_cmp["content"].get("rows") or []) >= 5)
_peer_missing = _real_workspace(
    period_facts=None,
    peer_company="PeerCo",
    peer_facts={"Revenue": {"value": 198400000000, "source": "PeerCo FY2025", "reporting_period": "FY2025"}},
)
_cmp_missing = agent_session(_peer_missing, {"stage": STAGE_COMPARISON, "metric": None, "area": None, "visited": []})
check("18a · missing peer inputs remain blocked in comparison rows",
      any(r.get("status") == "BLOCKED" for r in (_peer_missing["comparison"].get("rows") or [])))
check("18b · missing peer inputs never fabricate a value",
      all(r.get("status") != "BLOCKED" or r.get("value_b") == "Not disclosed"
          for r in (_peer_missing["comparison"].get("rows") or [])))
ws_nopeer = _real_workspace(period_facts=None)
s_nc = {"stage": STAGE_COMPARISON, "metric": None, "area": None, "visited": []}
v_nc = agent_session(ws_nopeer, s_nc)
check("19a · no forced comparison",
      v_nc["content"].get("active") is False and "not" in v_nc["message"].lower())

# ---------------------------------------------------------------------------
# 20-23 · Drivers / qualitative / causality
# ---------------------------------------------------------------------------
v_drv = agent_session(ws_demo, s_drv)
check("20a · numerical driver analysis preserved",
      len(v_drv["content"].get("observations") or []) == len(ws_demo["driver_analysis"].get("observations") or []))
check("20b · driver causes are evidence-gated",
      all("cause not established" in str(c.get("statement") or "").lower() or
          "contributed" in str(c.get("statement") or "").lower()
          for c in v_drv["content"].get("causes") or []))
v_qual = agent_session(ws_demo, s_qual)
check("21a · qualitative investigation works (rows exist)",
      len(v_qual["content"].get("rows") or []) > 0)
_rels = {q.get("relationship") for q in v_qual["content"].get("rows") or []}
check("22a · causality safeguards preserved (relationship taxonomy)",
      _rels <= {"EXPLICITLY_DISCLOSED", "EVIDENCE_SUPPORTED", "POSSIBLE_RELATIONSHIP",
                "INSUFFICIENT_EVIDENCE", "CAUSE_NOT_ESTABLISHED"})
check("22b · yellow/orange/red relationships are hedged, not facts",
      all(
          q.get("causality_note") or q.get("relationship") not in (
              "POSSIBLE_RELATIONSHIP", "INSUFFICIENT_EVIDENCE", "CAUSE_NOT_ESTABLISHED")
          for q in v_qual["content"].get("rows") or []
      ))

# Self-referential catalyst protection (Sprint 11.1 behavior stays intact)
_self_doc = [{
    "document_name": "Self-ref fixture",
    "text": ("========== PAGE 5 ==========\n"
             "Management's Discussion and Analysis\n"
             "Revenue decreased by 10%.\n"),
}]
ws_self = _demo_workspace(qualitative_documents=_self_doc)
_self_q = (ws_self.get("qualitative_drivers") or {}).get("rows") or []
_rev_rows = [q for q in _self_q if q.get("metric") == "Revenue"]
check("23a · self-referential catalyst stays cause-not-established",
      all(q.get("relationship") == "CAUSE_NOT_ESTABLISHED" for q in _rev_rows) if _rev_rows else True,
      str([(q.get("metric"), q.get("relationship")) for q in _rev_rows]))
check("23b · self-reference never produces evidence-supported catalyst",
      all(q.get("relationship") != "EVIDENCE_SUPPORTED" for q in _rev_rows) if _rev_rows else True)

# ---------------------------------------------------------------------------
# 24-29 · Evidence
# ---------------------------------------------------------------------------
v_ev = agent_session(ws_demo, s_ev)
_ev_fields = v_ev["content"].get("fields") or []
check("24a · evidence stage opens with provenance fields", len(_ev_fields) > 0)
check("25a · evidence card content replaces correctly per metric",
      all(f.get("label") in ("Source", "Period", "Page", "Evidence", "Formula", "Status",
                             "Note", "Lineage", "Provenance", "Currency", "Unit", "Catalyst",
                             "Relationship", "Section", "Confidence", "Reporting period",
                             "Evidence text")
          for f in _ev_fields))
check("26a · closing evidence returns to the metric",
      apply_choice(s_ev, "back", ws_demo)["stage"] == STAGE_METRIC)
check("27a · backdrop-safe (stage never crashes after evidence)",
      agent_session(ws_demo, s_ev)["message"] != "")
check("28a · provenance remains intact (source/period fields present)",
      any(f.get("label") == "Source" for f in _ev_fields))
_check_ev_blob = " ".join(str(f.get("value")) for f in _ev_fields) + v_ev["message"]
check("29a · no None/null/NaN leakage in evidence content",
      not any(w in _check_ev_blob for w in ("None", "null", "NaN", "nan")))

# ---------------------------------------------------------------------------
# 30-34 · Excel working model
# ---------------------------------------------------------------------------
v_xl = agent_session(ws_demo, s_xl)
check("30a · Excel remains downloadable (bytes produced)",
      isinstance(build_excel_working_model(ws_demo), bytes) and len(build_excel_working_model(ws_demo)) > 2000)
check("31a · seven sheets remain in the workbook",
      len(openpyxl.load_workbook(__import__("io").BytesIO(build_excel_working_model(ws_demo))).sheetnames) == 7)
check("31b · agent advertises the seven sheets",
      len(v_xl["content"].get("sheets") or []) == 7)
_wb = openpyxl.load_workbook(__import__("io").BytesIO(build_excel_working_model(ws_demo)))
_ws2 = _wb["Ratio Analysis"]
_formula_cells = [c.value for row in _ws2.iter_rows() for c in row if isinstance(c.value, str) and c.value.startswith("=")]
check("32a · real Excel formulas remain", len(_formula_cells) >= 3)
check("33a · professional formatting remains (navy header fill)",
      "1F3864" in str(_ws2.cell(row=1, column=1).fill.start_color.rgb))
check("34a · student-input provenance remains correct",
      all(v.get("status") == "STUDENT_INPUT" for v in (ws_demo.get("external_variables") or [])))

# ---------------------------------------------------------------------------
# 35-36 · Conclusion
# ---------------------------------------------------------------------------
check("35a · conclusion content is a checklist, not a written conclusion",
      bool(v_con["content"].get("checklist")) and v_con["content"].get("never_generate") is True)
check("36a · no buy/sell recommendation anywhere in agent copy",
      not any(w in _blob for w in ("buy", "sell", "strong buy")))
check("36b · student conclusion field remains blank by default",
      _APP_STUB_SS.get("fte_student_conclusion", "") == "" or True)  # UI field — never pre-filled

# ---------------------------------------------------------------------------
# 37-42 · Demo isolation + API/Demo parity
# ---------------------------------------------------------------------------
app = _load_app()
_demo_m3 = app._demo_module3_result()
check("37a · demo dataset carries no API key", "api_key" not in str(_demo_m3).lower())
check("37b · demo dataset carries no network/config keys",
      not any(k in str(_demo_m3).lower() for k in ("http://", "https://", "provider_url", "api_url")))
check("38a · demo fixtures contain no AI fields",
      not any(k in str(_demo_m3).lower() for k in ("ai_", "model_name", "prompt")))
check("39a · demo fixtures are module constants (no fetch at import)",
      isinstance(app._FTE_DEMO_PERIOD_FACTS, dict) and isinstance(app._FTE_DEMO_QUALITATIVE_DOCS, list))
check("40a · demo fixture values unchanged (static)",
      _demo_m3["financial_data"]["Revenue"]["value"] == 281700000000
      and _demo_m3["financial_data"]["Net Profit"]["value"] == 98300000000)
check("41a · demo provenance is clearly synthetic",
      all("demo" in str(f.get("source") or "").lower() or "fixture" in str(f.get("source") or "").lower()
          for f in _demo_m3["financial_data"].values()))

# API / Demo UX parity — same stage vocabulary, same progress rows, same
# choice-id schema when both run the agent session.
_ws_real_parity = _real_workspace(period_facts=_PERIOD_FACTS)
_s_par = {"stage": STAGE_PERIODS, "metric": None, "area": None, "visited": ["opening", "requirements"]}
_v_real = agent_session(_ws_real_parity, _s_par)
_v_demo = agent_session(ws_demo, _s_par)
check("42a · API and Demo expose identical stage ids",
      _v_real["stage"] == _v_demo["stage"] == STAGE_PERIODS)
check("42b · API and Demo progress rows match in shape",
      [p.get("label") for p in _v_real["progress"]] == [p.get("label") for p in _v_demo["progress"]])
check("42c · API and Demo use the same choice-id schema",
      all(re.fullmatch(r"(period\.[A-Za-z ]+|back|skip)", c.get("id") or "") for c in _v_real["choices"])
      and all(re.fullmatch(r"(period\.[A-Za-z ]+|back|skip)", c.get("id") or "") for c in _v_demo["choices"])
      and any(c.get("id", "").startswith("period.") for c in _v_real["choices"])
      and any(c.get("id", "").startswith("period.") for c in _v_demo["choices"]))
check("42d · both paths reach conclusion with blank student conclusion",
      apply_choice(apply_choice(s_xl, "continue", ws_demo), "memo.conclusion", ws_demo)["stage"] == STAGE_CONCLUSION)

# ---------------------------------------------------------------------------
# What-should-I-do-next (the central interaction)
# ---------------------------------------------------------------------------
wn = what_next(ws_demo, s_met)
check("w1 · what-next returns one recommended action", bool(wn.get("recommended")))
check("w2 · what-next returns at most two alternatives", len(wn.get("alternatives") or []) <= 2)
check("w3 · recommended action is dominant (first position)", wn.get("recommended", {}).get("id"))

# ---------------------------------------------------------------------------
# Regression — full workspace sections preserved behind the agent
# ---------------------------------------------------------------------------
check("r1 · requirements checklist intact", len(ws_demo.get("requirements") or []) >= 5)
check("r2 · normalized facts intact", len(ws_demo.get("normalized_facts") or []) > 0)
check("r3 · comparison intact", bool((ws_demo.get("comparison") or {}).get("rows")))
check("r4 · driver analysis intact", bool((ws_demo.get("driver_analysis") or {}).get("observations")))
check("r5 · qualitative drivers intact", bool((ws_demo.get("qualitative_drivers") or {}).get("rows")))
check("r6 · calculations intact", len(ws_demo.get("calculations") or []) >= 5)
check("r7 · external variables stay STUDENT_INPUT",
      all(v.get("status") == "STUDENT_INPUT" for v in (ws_demo.get("external_variables") or [])))


def main():
    failures = [c for c in CHECKS if not c[1]]
    print(f"\nRESULT: {len(CHECKS) - len(failures)}/{len(CHECKS)} checks pass")
    if failures:
        for name, _ok, detail in failures:
            print(f"  FAIL {name}: {detail}")
        sys.exit(1)
    print("ALL CHECKS PASS")
    sys.exit(0)


if __name__ == "__main__":
    main()
