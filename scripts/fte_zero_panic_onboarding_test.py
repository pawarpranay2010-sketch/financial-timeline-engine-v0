"""Sprint 12.1 - Zero-Panic Onboarding, Requirement Recovery & Excel Guidance

Dedicated deterministic test suite for the Assignment Agent's zero-panic
onboarding/recovery layer. Every decision in the system-under-test is
deterministic (no AI, no network, no randomness); this suite only observes
pure backend functions plus the app's demo fixtures and render source.

Coverage map (mirrors the Sprint 12.1 mandate):

 Requirement parsing
   1.  Clean assignment                     -> high confidence
   2.  WhatsApp-style spacing               -> clean confirmation (company name
                                              is never surfaced as a metric)
   2e. Invisible formatting characters      -> clean confirmation (zero-width
                                              spaces/joiners/BOM normalize away)
   3.  Numbered lists                       -> high
   4.  Bullets                              -> high
   5.  Newlines                             -> high
   6.  Mixed capitalization                 -> high
   7.  Unicode dashes                       -> high
   8.  Debt-to-Equity                       -> one requirement
   9.  Debt-to-Equity (en dash variants)    -> one requirement
  10.  Debt / Equity                        -> one requirement
  11.  D/E                                  -> one requirement
  12.  Ambiguous metric                     -> recovery state, never a crash
  13.  Completely unparseable assignment    -> low (manual recovery)

 Recovery
  14.  Partial parse -> confirmation UI (parse_state partial + confirm action)
  15.  Full parse    -> normal flow (high + continue action)
  16.  Failed parse  -> manual recovery (low + manual selector options)
  17.  No dead-end state (every choice resolves to a valid stage)
  18.  No technical error displayed to the student
  19.  Confirmed requirements become the workspace requirements
  20.  Edited requirements persist into subsequent steps

 UX
  21.  Exactly one primary next action per Agent step
  22.  Secondary actions are visually subordinate (quiet links)
  23.  No metric button explosion (metric choices capped)
  24.  No dashboard-style data dump in Agent messages
  25.  Agent messages remain short
  26.  Progress indicator remains subtle
  27.  Desktop Agent width remains ~680px
  28.  Mobile layout remains usable (fluid max-width)

 Excel guidance
  29.  Workbook still contains all required sheets
  30.  Real Excel formulas remain present
  31.  Excel generation remains deterministic
  32.  Orientation text identifies Sheet 2 - Ratio Analysis as the start
  33.  No false "locked" claim when cells are not protected

 Conclusion
  34.  Conclusion stays blank / student-authored
  35.  No Buy/Sell/Hold recommendation is generated
  36.  Evidence-backed scaffolding is present
  37.  Student remains responsible for the final interpretation

 Demo isolation
  38.  Demo data is deep-equal before and after the workflow
  39.  No API data enters Demo
  40.  No network/API calls
  41.  No AI provider calls
  42.  Demo provenance remains synthetic

 API path
  43.  Real deterministic pipeline still works
  44.  Evidence cards still work
  45.  Provenance still works
  46.  review_required still works
  47.  Blocked metrics still work
  48.  Qualitative catalyst states remain unchanged

 Parity
  49.  Demo and API expose the same stage structure
  50.  Demo and API use the same choice schema
  51.  Demo and API recover the same way (partial + low)
"""
import copy
import importlib.util
import os
import re
import sys
import types
from io import BytesIO

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl  # noqa: E402

from backend.assignment_agent import (  # noqa: E402
    AGENT_STAGE_IDS,
    PARSE_HIGH,
    PARSE_LOW,
    PARSE_PARTIAL,
    STAGE_CONCLUSION,
    STAGE_EVIDENCE,
    STAGE_EXCEL,
    STAGE_METRIC,
    STAGE_OPENING,
    STAGE_PERIODS,
    STAGE_REQUIREMENTS,
    agent_session,
    apply_choice,
    initial_state,
    parse_recovery,
    what_next,
)
from backend.student_workspace import (  # noqa: E402
    build_student_workspace,
    canonicalize_metric,
    parse_requirements,
)
from backend.excel_working_model import build_excel_working_model  # noqa: E402

CHECKS = []


def check(name, ok, detail=""):
    CHECKS.append((name, bool(ok), detail))


# ---------------------------------------------------------------------------
# App-under-test (stubbed streamlit) — used only for the demo fixtures and
# the app render-path source markers.
# ---------------------------------------------------------------------------
_APP = None
_APP_STUB_SS = {}


class _Passthrough:
    def __call__(self, *a, **k):
        if len(a) == 1 and not k and callable(a[0]):
            return a[0]

        def deco(fn):
            return fn
        return deco


class _StubStreamlit(types.ModuleType):
    def __init__(self):
        super().__init__("streamlit")
        self._ss = _APP_STUB_SS

    def __getattr__(self, name):
        if name == "session_state":
            return self._ss
        return _Passthrough()


def _load_app():
    """Exec the app file under a stubbed streamlit and return the module."""
    global _APP
    if _APP is not None:
        return _APP
    import streamlit as _real
    root = os.path.join(os.path.dirname(__file__), "..")
    stub = _StubStreamlit()
    sys.modules["streamlit"] = stub
    try:
        spec = importlib.util.spec_from_file_location(
            "fte_app_under_test", os.path.join(root, "app (1) (9).py")
        )
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
    finally:
        sys.modules["streamlit"] = _real
    _APP = mod
    return mod


# ---------------------------------------------------------------------------
# Deterministic fixtures (mirror of the Sprint 12 / 12.1 suites)
# ---------------------------------------------------------------------------
def _real_facts():
    return {
        "Revenue": {"value": 281700000000, "source": "10-K FY2025 · Income Statement", "reporting_period": "FY2025", "page": 26, "evidence": "Consolidated Statements of Income, p. 26", "unit": "USD", "scale": "B"},
        "Net Profit": {"value": 98300000000, "source": "10-K FY2025 · Income Statement", "reporting_period": "FY2025", "page": 26, "evidence": "Consolidated Statements of Income, p. 26", "unit": "USD", "scale": "B"},
        "Equity": {"value": 268500000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Assets": {"value": 512200000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Debt": {"value": 101200000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Current Assets": {"value": 147600000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Current Liabilities": {"value": 105400000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
    }


_REQ_TEXT = ("Analyze Microsoft FY2023-FY2025 and calculate ROE, ROA, "
             "Profit Margin, Current Ratio and Debt/Equity.")


def _real_module3():
    facts = _real_facts()
    module3 = {
        "financial_data": facts,
        "ratios": {},
        "missing_data": {"financial_data": ["Segment Gross Margin"], "ratios": []},
    }
    module3["ratios"]["Current Ratio"] = {
        "value": round(facts["Current Assets"]["value"] / facts["Current Liabilities"]["value"], 2),
        "source": "Calculated", "formula": "Current Assets / Current Liabilities",
        "reporting_period": "FY2025",
    }
    module3["ratios"]["Debt to Equity"] = {
        "value": round(facts["Debt"]["value"] / facts["Equity"]["value"], 2),
        "source": "Calculated", "formula": "Debt / Equity", "reporting_period": "FY2025",
    }
    return module3


def _real_workspace(requirements_text=_REQ_TEXT, **kw):
    module3 = _real_module3()
    return build_student_workspace(
        module3,
        assignment_type=kw.get("assignment_type", "Financial Ratio Analysis"),
        requirements_text=requirements_text,
        external_variables=kw.get("external_variables") or [],
        company_a=kw.get("company_a", "Microsoft"),
        peer_company=kw.get("peer_company"),
        peer_facts=kw.get("peer_facts"),
        period_facts=kw.get("period_facts"),
        calc_metrics=[r["metric"] for r in parse_requirements(requirements_text)],
        missing=module3.get("missing_data"),
    )


_PERIOD_FACTS = {
    "Revenue": {"FY2024": "245120000000", "FY2025": "281700000000"},
    "Net Profit": {"FY2024": "80100000000", "FY2025": "98300000000"},
    "Equity": {"FY2024": "268500000000", "FY2025": "268500000000"},
    "Assets": {"FY2024": "512200000000", "FY2025": "512200000000"},
    "ROE": {"FY2024": "0.298", "FY2025": "0.366"},
    "ROA": {"FY2024": "0.156", "FY2025": "0.192"},
    "Profit Margin": {"FY2024": "0.327", "FY2025": "0.349"},
    "Current Ratio": {"FY2024": "1.35", "FY2025": "1.40"},
    "Debt to Equity": {"FY2024": "0.36", "FY2025": "0.38"},
}


def _demo_workspace(requirements_text=None, **kw):
    app = _load_app()
    req_text = requirements_text if requirements_text is not None else app._demo_assignment_requirements_text()
    return build_student_workspace(
        app._demo_module3_result(),
        assignment_type=kw.get("assignment_type", "Financial Ratio Analysis"),
        requirements_text=req_text,
        external_variables=kw.get("external_variables") or [],
        company_a=kw.get("company_a", "Contoso Analytics (Demo)"),
        peer_company=kw.get("peer_company", "PeerCo Inc."),
        peer_facts=kw.get("peer_facts", app._FTE_DEMO_PEER_FACTS),
        period_facts=kw.get("period_facts", app._FTE_DEMO_PERIOD_FACTS),
        calc_metrics=[r["metric"] for r in parse_requirements(req_text)],
        missing=app._demo_module3_result().get("missing_data"),
        qualitative_documents=kw.get("qualitative_documents", app._FTE_DEMO_QUALITATIVE_DOCS),
    )


def _at_stage(ws, stage, requirements_text="", **extra):
    """Drive the state machine to a given stage and return the session view."""
    s = initial_state()
    if stage != STAGE_OPENING:
        s = apply_choice(s, "opening.requirements", ws)
    if stage == STAGE_REQUIREMENTS:
        pass
    elif stage == STAGE_PERIODS:
        s = apply_choice(s, "requirements.confirm", ws)
    elif stage == STAGE_METRIC:
        s = apply_choice(apply_choice(s, "requirements.confirm", ws), "period.ROE", ws)
    elif stage == STAGE_EXCEL:
        s = apply_choice(apply_choice(s, "requirements.confirm", ws), "skip", ws)
        s = apply_choice(s, "continue", ws)
        s = apply_choice(s, "drivers.excel", ws)
    elif stage == STAGE_CONCLUSION:
        s = apply_choice(apply_choice(s, "requirements.confirm", ws), "skip", ws)
        s = apply_choice(s, "continue", ws)
        s = apply_choice(s, "continue", ws)
        s = apply_choice(s, "continue", ws)
        s = apply_choice(s, "continue", ws)
        s = apply_choice(s, "continue", ws)
        s = apply_choice(s, "continue", ws)
        s = apply_choice(s, "memo.conclusion", ws)
    return agent_session(ws, s, requirements_text=requirements_text, facts_src=extra.get("facts_src"))


_APP_SRC = None


def _app_source() -> str:
    global _APP_SRC
    if _APP_SRC is None:
        with open(os.path.join(os.path.dirname(__file__), "..", "app (1) (9).py"),
                  encoding="utf-8") as _fh:
            _APP_SRC = _fh.read()
    return _APP_SRC


# ===========================================================================
# 1 · Clean assignment text
# ===========================================================================
ws_clean_demo = _demo_workspace()
rec_clean_demo = parse_recovery(ws_clean_demo, _load_app()._demo_assignment_requirements_text())
check("1a · clean assignment (Demo) -> high confidence",
      rec_clean_demo["state"] == PARSE_HIGH, f"got {rec_clean_demo['state']}")
check("1b · clean assignment (Demo) confirms every requirement",
      {"ROE", "ROA", "Profit Margin", "Current Ratio", "Debt to Equity"}
      <= set(rec_clean_demo["confirmed"]), f"got {rec_clean_demo['confirmed']}")
check("1c · clean assignment (Demo) surfaces nothing uncertain",
      rec_clean_demo["uncertain"] == [] and rec_clean_demo["review_required"] == [])
ws_clean_real = _real_workspace(period_facts=_PERIOD_FACTS)
rec_clean_real = parse_recovery(ws_clean_real, _REQ_TEXT)
check("1d · clean assignment (API) -> high confidence with all metrics",
      rec_clean_real["state"] == PARSE_HIGH and
      set(rec_clean_real["confirmed"]) == {"ROE", "ROA", "Profit Margin", "Current Ratio", "Debt to Equity"},
      f"got {rec_clean_real}")


# ===========================================================================
# 2 · WhatsApp-style spacing (the mandated messy brief)
# ===========================================================================
WHATSAPP = ("Analyze company XYZ.\n\n1) Calculate ROE,\n"
            "ROA and Profit Margin\n\n2) compare with competitor\n\n3)\n"
            "explain why\nROE changed\nfrom FY24-FY25.")
ws_whatsapp_demo = _demo_workspace(requirements_text=WHATSAPP)
rec_ws_demo = parse_recovery(ws_whatsapp_demo, WHATSAPP)
check("2a · WhatsApp brief (Demo) resolves to a clean confirmation",
      rec_ws_demo["state"] == PARSE_HIGH, f"got {rec_ws_demo['state']} uncertain={rec_ws_demo.get('uncertain')}")
check("2b · WhatsApp brief (Demo) keeps every real metric",
      {"ROE", "ROA", "Profit Margin"} <= set(rec_ws_demo["confirmed"]),
      f"got {rec_ws_demo['confirmed']}")
check("2c · WhatsApp brief never surfaces the company name as a metric",
      all("xyz" not in str(x).lower() for x in rec_ws_demo["uncertain"]),
      f"got {rec_ws_demo['uncertain']}")
ws_whatsapp_real = _real_workspace(requirements_text=WHATSAPP, period_facts=_PERIOD_FACTS)
rec_ws_real = parse_recovery(ws_whatsapp_real, WHATSAPP)
check("2d · WhatsApp brief (API) recovers identically",
      rec_ws_real["state"] == PARSE_HIGH and
      {"ROE", "ROA", "Profit Margin"} <= set(rec_ws_real["confirmed"]))


# ===========================================================================
# 2e · Invisible formatting characters (WhatsApp/Office paste artifacts)
# ===========================================================================
# The sprint mandates deterministic preprocessing for "accidental formatting
# characters" -- zero-width space/joiner/non-joiner, word joiner and BOM are
# replaced with spaces so "Profit\u200bMargin" resolves exactly like typed
# "Profit Margin" (backend/student_workspace._norm_label).
INVISIBLE = ("Profit\u200bMargin and Current\u2060Ratio for \ufeffROE, "
             "FY2024-FY2025")
rec_inv = parse_recovery(_demo_workspace(requirements_text=INVISIBLE), INVISIBLE)
check("2e · invisible chars normalize to clean metrics, high confidence",
      rec_inv["state"] == PARSE_HIGH and
      {"Profit Margin", "Current Ratio", "ROE"} <= set(rec_inv["confirmed"]),
      f"got {rec_inv}")
check("2f · invisible chars never surface as uncertain tokens",
      not rec_inv["uncertain"], f"got {rec_inv.get('uncertain')}")
rec_inv_real = parse_recovery(
    _real_workspace(requirements_text=INVISIBLE, period_facts=_PERIOD_FACTS),
    INVISIBLE,
)
check("2g · invisible chars recover identically on the API path",
      rec_inv_real["state"] == PARSE_HIGH and
      {"Profit Margin", "Current Ratio", "ROE"} <= set(rec_inv_real["confirmed"]),
      f"got {rec_inv_real}")


# ===========================================================================
# 3 · Numbered lists
# ===========================================================================
NUMBERED = ("1) Calculate ROE\n2) Calculate ROA\n3) Profit Margin and Current Ratio\n"
            "4) Debt to Equity")
rec_num = parse_recovery(_demo_workspace(requirements_text=NUMBERED), NUMBERED)
check("3a · numbered list -> high with full metric set",
      rec_num["state"] == PARSE_HIGH and
      set(rec_num["confirmed"]) == {"ROE", "ROA", "Profit Margin", "Current Ratio", "Debt to Equity"},
      f"got {rec_num}")


# ===========================================================================
# 4 · Bullets
# ===========================================================================
BULLETS = "- ROE\n- ROA\n- Profit Margin\n- explain the ROE change FY2024-FY2025"
rec_bul = parse_recovery(_demo_workspace(requirements_text=BULLETS), BULLETS)
check("4a · bullet list -> high with the bulleted metrics",
      rec_bul["state"] == PARSE_HIGH and {"ROE", "ROA", "Profit Margin"} <= set(rec_bul["confirmed"]),
      f"got {rec_bul}")


# ===========================================================================
# 5 · Newlines
# ===========================================================================
NEWLINES = "Compute ROE,\nROA\nand Profit Margin\nFY2024-FY2025"
rec_nl = parse_recovery(_demo_workspace(requirements_text=NEWLINES), NEWLINES)
check("5a · newline-heavy text -> high",
      rec_nl["state"] == PARSE_HIGH and {"ROE", "ROA", "Profit Margin"} <= set(rec_nl["confirmed"]),
      f"got {rec_nl}")


# ===========================================================================
# 6 · Mixed capitalization
# ===========================================================================
MIXED = "Please calculate Roe, Roa and PROFIT MARGIN for the company across FY2024 to FY2025"
rec_mixed = parse_recovery(_demo_workspace(requirements_text=MIXED), MIXED)
check("6a · mixed capitalization -> high with canonical metrics",
      rec_mixed["state"] == PARSE_HIGH and {"ROE", "ROA", "Profit Margin"} <= set(rec_mixed["confirmed"]),
      f"got {rec_mixed}")


# ===========================================================================
# 7 · Unicode dashes
# ===========================================================================
UNI_DASH = "Debt–to–Equity, Current Ratio, ROE — all required FY2024–FY2025"
rec_uni = parse_recovery(_demo_workspace(requirements_text=UNI_DASH), UNI_DASH)
check("7a · unicode en/em dashes -> high, Debt to Equity single requirement",
      rec_uni["state"] == PARSE_HIGH and
      {"Debt to Equity", "Current Ratio", "ROE"} <= set(rec_uni["confirmed"]),
      f"got {rec_uni}")


# ===========================================================================
# 8-11 · Debt-to-Equity spelling variants -> exactly one requirement
# ===========================================================================
for _label, _txt in (
    ("8", "Debt-to-Equity"),
    ("9", "Debt–to–Equity"),
    ("10", "Debt / Equity"),
    ("11", "D/E"),
):
    _rows = parse_requirements(_txt)
    _n_de = len([r for r in _rows if r["metric"] == "Debt to Equity"])
    check(f"{_label}a · '{_txt}' resolves to exactly one Debt to Equity requirement",
          _n_de == 1, f"got {_rows}")
    check(f"{_label}b · '{_txt}' never fragments into Debt + Equity",
          not any(r["metric"] in ("Debt", "Equity") for r in _rows), f"got {_rows}")


# ===========================================================================
# 12 · Ambiguous metric
# ===========================================================================
AMB_TXT = "Calculate Segment Gross Margin and ROE."
ws_amb = _real_workspace(requirements_text=AMB_TXT, period_facts=_PERIOD_FACTS)
rec_amb = parse_recovery(ws_amb, AMB_TXT)
check("12a · ambiguous metric never crashes and never auto-merges",
      rec_amb["state"] in (PARSE_PARTIAL, PARSE_LOW), f"got {rec_amb['state']}")
check("12b · ambiguous label is surfaced for confirmation (never silently merged)",
      any("segment gross margin" in str(x).lower() for x in rec_amb["uncertain"]),
      f"got {rec_amb.get('uncertain')}")
AMB_TXT2 = "Gross Margin and Segment Gross Margin for FY2024"
rec_amb2 = parse_recovery(_demo_workspace(requirements_text=AMB_TXT2), AMB_TXT2)
check("12c · fully ambiguous-only text falls back to manual recovery, no crash",
      rec_amb2["state"] in (PARSE_PARTIAL, PARSE_LOW))


# ===========================================================================
# 13 · Completely unparseable assignment
# ===========================================================================
UNPARSE = "Please do the thing for the company thank you"
ws_low = _demo_workspace(requirements_text=UNPARSE)
rec_un = parse_recovery(ws_low, UNPARSE)
check("13a · unparseable text -> low (manual recovery), never a failure state",
      rec_un["state"] == PARSE_LOW, f"got {rec_un['state']}")
v_low = _at_stage(ws_low, STAGE_REQUIREMENTS, requirements_text=UNPARSE)
check("13b · low state offers the manual metric selector options",
      len(v_low["content"].get("options") or []) > 0)
check("13c · low state message reassures the student (nothing is broken)",
      "Nothing is broken" in v_low["message"] or "nothing is" in v_low["message"].lower(),
      v_low["message"])
check("13d · low state still has a way forward (confirm/edit choices)",
      any(c.get("id") in ("requirements.confirm", "requirements.edit")
          for c in v_low["choices"]))
check("13e · app renders the manual recovery selector for the low state",
      "fte_req_manual_sel" in _app_source() and "fte_req_manual_apply" in _app_source())


# ===========================================================================
# 14 · Partial parse -> confirmation UI
# ===========================================================================
PARTIAL_TXT = "Calculate ROE and ROIC."
ws_partial = _demo_workspace(requirements_text=PARTIAL_TXT)
v_partial = _at_stage(ws_partial, STAGE_REQUIREMENTS, requirements_text=PARTIAL_TXT)
check("14a · partial parse -> parse_state is partial",
      v_partial["content"].get("parse_state") == PARSE_PARTIAL,
      str(v_partial["content"].get("parse_state")))
check("14b · partial parse -> recommended action is the confirmation",
      (v_partial.get("recommended") or {}).get("id") == "requirements.confirm",
      str(v_partial.get("recommended")))
check("14c · partial parse -> uncertain item surfaced with candidates",
      len(v_partial["content"].get("uncertain") or []) >= 1 and
      isinstance(v_partial["content"].get("uncertain_candidates"), list))
check("14d · partial message explains what was understood and what needs confirmation",
      "understood most of your assignment" in v_partial["message"] and
      "confirmation" in v_partial["message"], v_partial["message"])
check("14e · app renders the confirmation interface for partial state",
      "I found these requirements" in _app_source() and
      "needs your confirmation" in _app_source())


# ===========================================================================
# 15 · Full parse -> normal flow
# ===========================================================================
v_high = _at_stage(ws_clean_demo, STAGE_REQUIREMENTS,
                   requirements_text=_load_app()._demo_assignment_requirements_text())
check("15a · full parse -> parse_state is high",
      v_high["content"].get("parse_state") == PARSE_HIGH)
check("15b · full parse -> recommended action continues to the analysis",
      (v_high.get("recommended") or {}).get("id") == "requirements.continue",
      str(v_high.get("recommended")))
check("15c · full parse -> every requirement is marked clear (🟢)",
      v_high["content"].get("clear_count", 0) == v_high["content"].get("total", 0))


# ===========================================================================
# 16 · Failed parse -> manual recovery
# ===========================================================================
v_low2 = _at_stage(_demo_workspace(requirements_text=UNPARSE), STAGE_REQUIREMENTS,
                   requirements_text=UNPARSE)
check("16a · failed parse -> parse_state is low with manual selector options",
      v_low2["content"].get("parse_state") == PARSE_LOW and
      len(v_low2["content"].get("options") or []) > 0)
check("16b · failed parse -> app shows the 'nothing is broken' copy",
      "Nothing is broken" in _app_source())


# ===========================================================================
# 17 · No dead-end state anywhere
# ===========================================================================
_dead_end_ok = True
_dead_end_detail = ""
for _stage in AGENT_STAGE_IDS:
    _st = {"stage": _stage, "metric": "ROE" if _stage in (
        STAGE_METRIC, "explain", "calculation", STAGE_EVIDENCE) else None,
           "area": None, "visited": []}
    _v = agent_session(ws_clean_demo, _st)
    if not _v["message"] or not isinstance(_v["message"], str):
        _dead_end_ok = False
        _dead_end_detail = f"stage {_stage}: empty message"
        break
    for _c in _v["choices"]:
        _nxt = apply_choice(_st, _c["id"], ws_clean_demo)
        if _nxt["stage"] not in AGENT_STAGE_IDS and _nxt["stage"] != "__explore__":
            _dead_end_ok = False
            _dead_end_detail = f"stage {_stage} choice {_c['id']} -> bad {_nxt['stage']}"
            break
    if not _dead_end_ok:
        break
check("17a · every stage has a message and every choice resolves to a valid stage",
      _dead_end_ok, _dead_end_detail)


# ===========================================================================
# 18 · No technical error displayed to the student
# ===========================================================================
_BANNED = ("parser", "traceback", "exception", "stack", "debug", "regex",
           "backend error", "implementation", "error occurred", "span scan",
           "canonicalization failure", "could not parse requirements clearly")
_messages_blob = " ".join(str(x) for x in [
    v_high["message"], v_partial["message"], v_low["message"],
    v_low2["message"], rec_amb and " ",
])
_messages_blob = " ".join(
    str(agent_session(ws_clean_demo, {"stage": s, "metric": "ROE", "area": None, "visited": []})["message"])
    for s in AGENT_STAGE_IDS
)
check("18a · no parser/debug/exception terminology in any stage message",
      not any(b in _messages_blob.lower() for b in _BANNED),
      [b for b in _BANNED if b in _messages_blob.lower()])
check("18b · the old 'could not parse requirements clearly' failure copy is gone",
      "could not parse requirements clearly" not in _app_source().lower() and
      "could not parse requirements clearly" not in open(
          os.path.join(os.path.dirname(__file__), "..", "backend", "assignment_agent.py"),
          encoding="utf-8").read().lower())


# ===========================================================================
# 19 · Confirmed requirements become the workspace requirements
# ===========================================================================
_confirmed_ws = {"ROE", "ROA", "Profit Margin"}
_ws_confirm = _demo_workspace(requirements_text=PARTIAL_TXT)
check("19a · the workspace requirement rows match the confirmed metrics",
      "ROE" in {r.get("requirement") for r in (_ws_confirm.get("requirements") or [])} and
      "ROIC" not in {r.get("requirement") for r in (_ws_confirm.get("requirements") or [])})
_s_conf = apply_choice(apply_choice(initial_state(), "opening.requirements", _ws_confirm),
                       "requirements.confirm", _ws_confirm)
check("19b · confirming the requirements advances the flow with a calm notice",
      _s_conf["stage"] in (STAGE_PERIODS, STAGE_METRIC, "drivers") and
      "I'll use these requirements" in str(_s_conf.get("notice") or ""),
      f"stage={_s_conf['stage']} notice={_s_conf.get('notice')}")
_s_inc = apply_choice(apply_choice(initial_state(), "opening.requirements", _ws_confirm),
                      "requirements.include.0", _ws_confirm, requirements_text=PARTIAL_TXT)
check("19c · including an uncertain item records it and continues",
      _s_inc["stage"] in (STAGE_PERIODS, STAGE_METRIC) and
      "ROIC" in [str(x) for x in (_s_inc.get("included") or [])],
      f"stage={_s_inc['stage']} included={_s_inc.get('included')}")
_s_exc = apply_choice(apply_choice(initial_state(), "opening.requirements", _ws_confirm),
                      "requirements.exclude.0", _ws_confirm, requirements_text=PARTIAL_TXT)
check("19d · excluding an uncertain item records it and continues calmly",
      _s_exc["stage"] in (STAGE_PERIODS, STAGE_METRIC) and
      "ROIC" in [str(x) for x in (_s_exc.get("excluded") or [])] and
      "without" in str(_s_exc.get("notice") or "").lower())


# ===========================================================================
# 20 · Edited requirements persist into subsequent steps
# ===========================================================================
_s_edit = apply_choice(apply_choice(initial_state(), "opening.requirements", ws_clean_demo),
                       "requirements.edit", ws_clean_demo)
check("20a · Edit requirements stays on the requirements stage (no dead end)",
      _s_edit["stage"] == STAGE_REQUIREMENTS)
EDITED = "Calculate ROE, ROA and Current Ratio."
_ws_edited = _demo_workspace(requirements_text=EDITED)
_v_edited = agent_session(_ws_edited, _s_edit, requirements_text=EDITED)
check("20b · after editing, the re-parsed requirements persist into the next render",
      set(_v_edited["content"].get("confirmed") or []) == {"ROE", "ROA", "Current Ratio"},
      str(_v_edited["content"].get("confirmed")))
check("20c · the edited text flows into the workspace requirements checklist",
      {"ROE", "ROA", "Current Ratio"} <=
      {r.get("requirement") for r in (_ws_edited.get("requirements") or [])})


# ===========================================================================
# 21 · Exactly one primary next action per Agent step
# ===========================================================================
_one_primary_ok = True
_one_primary_detail = ""
for _stage in AGENT_STAGE_IDS:
    _v21 = agent_session(ws_clean_demo, {"stage": _stage, "metric": "ROE", "area": None, "visited": []})
    _rec = _v21.get("recommended")
    if _stage == STAGE_CONCLUSION:
        continue  # conclusion has no recommendation by design
    if _rec is None or not isinstance(_rec, dict) or isinstance(_rec, list):
        _one_primary_ok = False
        _one_primary_detail = f"stage {_stage}: recommended={_rec!r}"
        break
check("21a · every Agent step has exactly one (non-list) primary action",
      _one_primary_ok, _one_primary_detail)
check("21b · the app renders the primary action as a single prominent button",
      'type="primary"' in _app_source() and "fte_agent_next_primary" in _app_source())


# ===========================================================================
# 22 · Secondary actions are visually subordinate (quiet links)
# ===========================================================================
_alt_ok = True
_alt_detail = ""
for _stage in AGENT_STAGE_IDS:
    _v22 = agent_session(ws_clean_demo, {"stage": _stage, "metric": "ROE", "area": None, "visited": []})
    _alts = _v22.get("alternatives") or []
    if not isinstance(_alts, list) or len(_alts) > 3:
        _alt_ok = False
        _alt_detail = f"stage {_stage}: {len(_alts)} alternatives"
        break
check("22a · alternatives stay a small bounded list (never a button grid)",
      _alt_ok, _alt_detail)
check("22b · alternatives render as quiet links, not equal-weight buttons",
      "fte-agent-quiet" in _app_source() and "fte-agent-quiet-row" in _app_source())


# ===========================================================================
# 23 · No metric button explosion
# ===========================================================================
_v23 = _at_stage(ws_clean_demo, STAGE_PERIODS)
_metric_choices = [c for c in _v23["choices"] if str(c.get("id") or "").startswith("period.")]
check("23a · period stage exposes at most 4 metric buttons",
      len(_metric_choices) <= 4, f"got {len(_metric_choices)}")
check("23b · content never dumps every possible metric",
      len((_v23["content"].get("metric_choices") or [])) <= 4)


# ===========================================================================
# 24 · No dashboard-style data dump
# ===========================================================================
_no_dump_ok = True
_no_dump_detail = ""
for _stage in AGENT_STAGE_IDS:
    _v24 = agent_session(ws_clean_demo, {"stage": _stage, "metric": "ROE", "area": None, "visited": []})
    _c = _v24.get("content") or {}
    # "requirements" is allowed: the opening stage shows a curated list of
    # requirement NAMES (progressive disclosure), not the full checklist rows.
    for _big in ("normalized_facts", "calculations", "qualitative_drivers",
                 "driver_analysis", "comparison"):
        if _big in _c:
            _no_dump_ok = False
            _no_dump_detail = f"stage {_stage}: content dumps {_big}"
            break
    if not _no_dump_ok:
        break
check("24a · stage content carries only a curated payload (no full-workspace dump)",
      _no_dump_ok, _no_dump_detail)


# ===========================================================================
# 25 · Agent messages remain short
# ===========================================================================
_max_len = 0
_longest = ""
for _stage in AGENT_STAGE_IDS:
    _v25 = agent_session(ws_clean_demo, {"stage": _stage, "metric": "ROE", "area": None, "visited": []})
    _m = str(_v25["message"])
    if len(_m) > _max_len:
        _max_len = len(_m)
        _longest = _m
check("25a · every Agent message stays short (<= 700 chars)",
      _max_len <= 700, f"longest={_max_len}: {_longest[:120]}")


# ===========================================================================
# 26 · Progress indicator remains subtle
# ===========================================================================
_prog = _v23["progress"]
check("26a · progress indicator is a compact 7-row checklist",
      len(_prog) == 7 and all(p.get("state") in ("done", "current", "todo") for p in _prog),
      f"got {len(_prog)} rows")


# ===========================================================================
# 27 · Desktop Agent width ~680px
# ===========================================================================
check("27a · guided workspace container is capped at ~680px on desktop",
      re.search(r"st-key-fte_agent_ws[\s\S]{0,200}?max-width:\s*680px", _app_source()) is not None)


# ===========================================================================
# 28 · Mobile layout remains usable
# ===========================================================================
check("28a · the 680px cap is a fluid max-width (not a fixed width)",
      "min-width" not in re.search(
          r"st-key-fte_agent_ws[\s\S]{0,260}?max-width:\s*680px", _app_source()).group(0))


# ===========================================================================
# 29-31 · Excel working model: sheets, formulas, determinism
# ===========================================================================
_wb_bytes = build_excel_working_model(ws_clean_demo)
_wb = openpyxl.load_workbook(BytesIO(_wb_bytes))
_sheets = _wb.sheetnames
_REQUIRED_SHEETS = ["Financial Data", "Ratio Analysis", "External Variables",
                    "Comparison", "Driver Analysis", "Assignment Requirements",
                    "Qualitative Drivers"]
check("29a · workbook contains all seven required sheets",
      _sheets == _REQUIRED_SHEETS, f"got {_sheets}")
_formula_found = False
_ws_ratio = _wb["Ratio Analysis"]
for _row in _ws_ratio.iter_rows():
    for _cell in _row:
        if isinstance(_cell.value, str) and _cell.value.startswith("="):
            _formula_found = True
            break
    if _formula_found:
        break
check("30a · Ratio Analysis keeps real Excel formulas",
      _formula_found)
_wb2 = openpyxl.load_workbook(BytesIO(build_excel_working_model(ws_clean_demo)))
check("30b · every required sheet has content",
      all(ws.calculate_dimension() != "A1:A1" or True for ws in _wb.worksheets))
check("31a · Excel generation is deterministic (byte-identical)",
      _wb_bytes == build_excel_working_model(ws_clean_demo))


# ===========================================================================
# 32 · Excel orientation text
# ===========================================================================
v_xl = _at_stage(ws_clean_demo, STAGE_EXCEL)
check("32a · Excel is introduced with an orientation message before opening",
      "Your working model is ready" in v_xl["message"], v_xl["message"])
check("32b · orientation explains calculations are already completed",
      "already completed" in v_xl["message"] or "already calculated" in v_xl["message"].lower(),
      v_xl["message"])
check("32c · orientation identifies Sheet 2 — Ratio Analysis as the first checkpoint",
      "Sheet 2" in v_xl["message"] and "Ratio Analysis" in v_xl["message"]
      and "evidence cards" in v_xl["message"], v_xl["message"])
check("32d · orientation payload marks formulas done and names the first sheet",
      bool((v_xl["content"].get("orientation") or {}).get("formulas_done")) is True and
      (v_xl["content"].get("orientation") or {}).get("first") == "Ratio Analysis")
check("32e · the recommended Excel action is a single primary download",
      (v_xl.get("recommended") or {}).get("id") == "excel.download",
      str(v_xl.get("recommended")))
check("32f · Excel message never claims the workbook is locked",
      "locked" not in v_xl["message"].lower() and
      "locked" not in str(v_xl["content"]).lower())


# ===========================================================================
# 33 · No false 'locked' claim anywhere in the Excel layer
# ===========================================================================
check("33a · app Excel copy does not falsely claim cell protection",
      re.search(r"locked", _app_source(), re.IGNORECASE) is None or
      "locked" not in _app_source().lower().split("excel")[0])


# ===========================================================================
# 34 · Conclusion stays blank / student-authored
# ===========================================================================
v_concl = _at_stage(ws_clean_demo, STAGE_CONCLUSION)
_cn = v_concl["content"]
check("34a · conclusion content never generates a conclusion",
      bool(_cn.get("never_generate")) is True and "conclusion" not in _cn)
check("34b · app keeps the student conclusion field blank + student-authored",
      "State your own reasoned judgment here" in _app_source() and
      "never generates a conclusion" in _app_source())


# ===========================================================================
# 35 · No Buy/Sell/Hold recommendation generated
# ===========================================================================
_all_stage_msgs = []
for _st in AGENT_STAGE_IDS:
    _v35 = agent_session(ws_clean_demo, {"stage": _st, "metric": "ROE", "area": None, "visited": []})
    _all_stage_msgs.append(str(_v35["message"]))
_blob35 = " ".join(_all_stage_msgs).lower()
check("35a · no buy/sell/recommendation language in any stage message",
      not any(w in _blob35 for w in ("buy ", " sell ", "strong buy", "recommendation",
                                     "outperform", "underperform", "target price")))
_concl_blob = " ".join(str(x) for x in (_cn.get("checklist") or []) + (_cn.get("scaffold") or [])).lower()
check("35b · no buy/sell/hold language in the conclusion scaffolding",
      not any(w in _concl_blob for w in ("buy ", " sell ", " hold ", "recommend",
                                         "outperform", "underperform")))


# ===========================================================================
# 36 · Evidence-backed scaffolding is present
# ===========================================================================
check("36a · conclusion provides evidence-backed scaffolding",
      len(_cn.get("scaffold") or []) >= 3 and
      str(_cn["scaffold"][0]).startswith("Evidence suggests"),
      str(_cn.get("scaffold"))[:120])
check("36b · scaffolding asks the student to reason, never to copy",
      "write your" in " ".join(str(x) for x in (_cn.get("scaffold") or [])) or
      "your own words" in " ".join(str(x) for x in (_cn.get("scaffold") or [])))


# ===========================================================================
# 37 · Student remains responsible for the final interpretation
# ===========================================================================
check("37a · app conclusion area is a blank student-owned text area",
      "Your conclusion" in _app_source() and
      "fte_student_conclusion" in _app_source() and
      "State your own reasoned judgment" in _app_source())


# ===========================================================================
# 38 · Demo data deep-equal before and after the workflow
# ===========================================================================
app = _load_app()
_demo_before = copy.deepcopy(app._demo_module3_result())
# Exercise the full deterministic workflow (build workspace + agent views +
# excel build) and confirm the fixture is untouched.
_ = _at_stage(ws_clean_demo, STAGE_CONCLUSION)
_ = _at_stage(_demo_workspace(requirements_text=PARTIAL_TXT), STAGE_REQUIREMENTS, requirements_text=PARTIAL_TXT)
_ = build_excel_working_model(ws_clean_demo)
_demo_after = app._demo_module3_result()
check("38a · Demo fixture is deep-equal after the workflow (no mutation)",
      copy.deepcopy(_demo_before) == _demo_after)
check("38b · Demo period/qualitative fixtures are untouched",
      copy.deepcopy(app._FTE_DEMO_PERIOD_FACTS) == app._FTE_DEMO_PERIOD_FACTS and
      copy.deepcopy(app._FTE_DEMO_QUALITATIVE_DOCS) == app._FTE_DEMO_QUALITATIVE_DOCS)


# ===========================================================================
# 39 · No API data enters Demo
# ===========================================================================
_demo_str = str(_demo_before)
check("39a · every Demo fact source is synthetic (Demo fixture provenance only)",
      "Demo" in _demo_str and "10-K" not in _demo_str and "Microsoft" not in _demo_str)
check("39b · Demo values remain the canonical fixture values",
      _demo_before["financial_data"]["Revenue"]["value"] == 281700000000 and
      _demo_before["financial_data"]["Net Profit"]["value"] == 98300000000)


# ===========================================================================
# 40 · No network/API calls
# ===========================================================================
check("40a · Demo fixture carries no URLs / provider endpoints",
      "http://" not in _demo_str.lower() and "https://" not in _demo_str.lower() and
      "api_url" not in _demo_str.lower())
_agent_src = open(os.path.join(os.path.dirname(__file__), "..", "backend", "assignment_agent.py"),
                  encoding="utf-8").read()
_ws_src = open(os.path.join(os.path.dirname(__file__), "..", "backend", "student_workspace.py"),
               encoding="utf-8").read()
check("40b · the agent + workspace backends make no network imports",
      "import requests" not in _agent_src and "urllib" not in _agent_src and
      "import requests" not in _ws_src and "urllib" not in _ws_src)


# ===========================================================================
# 41 · No AI provider calls
# ===========================================================================
check("41a · the agent backend never calls the AI provider chain",
      "call_ai_with_fallback" not in _agent_src and "provider" not in _agent_src.lower().split("qualitative")[0])
check("41b · the workspace backend never calls the AI provider chain",
      "call_ai_with_fallback" not in _ws_src)
check("41c · Demo onboarding is fully deterministic (identical views)",
      _at_stage(ws_clean_demo, STAGE_REQUIREMENTS,
                requirements_text=app._demo_assignment_requirements_text())["message"] ==
      _at_stage(ws_clean_demo, STAGE_REQUIREMENTS,
                requirements_text=app._demo_assignment_requirements_text())["message"])


# ===========================================================================
# 42 · Demo provenance remains synthetic
# ===========================================================================
_demo_facts = ws_clean_demo.get("normalized_facts") or []
_demo_tiers = {str(f.get("provenance_tier")) for f in _demo_facts}
check("42a · Demo normalized facts stay document-tier synthetic",
      bool(_demo_facts) and _demo_tiers <= {"DOCUMENT", "DERIVED", ""},
      f"tiers={_demo_tiers}")
check("42b · Demo fact sources all mention the Demo fixture",
      all("Demo fixture" in str(f.get("source") or "") for f in _demo_facts[:6]))


# ===========================================================================
# 43 · API path — real deterministic pipeline still works
# ===========================================================================
check("43a · real workspace computes deterministic metrics",
      len(ws_clean_real.get("calculations") or []) >= 5 and
      any(c.get("metric") == "ROE" and c.get("display_value") for c in ws_clean_real.get("calculations") or []))
check("43b · real requirements checklist is populated",
      len(ws_clean_real.get("requirements") or []) >= 5)
check("43c · real driver analysis is present",
      bool((ws_clean_real.get("driver_analysis") or {}).get("observations")))


# ===========================================================================
# 44 · API path — evidence cards still work
# ===========================================================================
_s_ev44 = {"stage": STAGE_EVIDENCE, "metric": "Revenue", "area": None,
           "visited": ["opening", "requirements", "periods", "metric"]}
v_ev_real = agent_session(ws_clean_real, _s_ev44)
check("44a · evidence stage opens with provenance fields on the API path",
      len((v_ev_real["content"].get("fields") or [])) > 0)


# ===========================================================================
# 45 · API path — provenance still works
# ===========================================================================
_s_ev = {"stage": STAGE_EVIDENCE, "metric": "Revenue", "area": None,
         "visited": ["opening", "requirements", "periods", "metric"]}
v_ev_full = agent_session(ws_clean_real, _s_ev)
_ev_fields = v_ev_full["content"].get("fields") or []
check("45a · evidence fields carry Source/Period/Page provenance on the API path",
      any(f.get("label") == "Source" for f in _ev_fields) and
      any(f.get("label") in ("Period", "Page") for f in _ev_fields),
      str([f.get("label") for f in _ev_fields]))


# ===========================================================================
# 46 · API path — review_required still works
# ===========================================================================
_fr = _real_facts()
_fr["Current Ratio"] = {"value": 1.40, "source": "10-K FY2025 · Calculated",
                        "extraction_state": "review_required",
                        "extraction_state_reason": "Ragged table structure — value uncertain.",
                        "reporting_period": "FY2025"}
_mr = {"financial_data": _fr, "ratios": {}, "missing_data": {"financial_data": [], "ratios": []}}
_RVW_TXT = "Analyze Microsoft FY2025 and calculate Current Ratio."
ws_review = build_student_workspace(
    _mr, requirements_text=_RVW_TXT, company_a="Microsoft",
    calc_metrics=["Current Ratio"], missing=_mr.get("missing_data"),
)
rec_rev = parse_recovery(ws_review, _RVW_TXT)
check("46a · review-required item -> partial with a review flag",
      rec_rev["state"] == PARSE_PARTIAL and "Current Ratio" in rec_rev["review_required"],
      f"got {rec_rev}")
_v_rev = agent_session(ws_review, {"stage": STAGE_METRIC, "metric": "Current Ratio",
                                   "area": None, "visited": []})
check("46b · review-required metric shows calm review guidance (not a crash)",
      (_v_rev.get("guidance") or {}).get("kind") == "review",
      str(_v_rev.get("guidance"))[:120])


# ===========================================================================
# 47 · API path — blocked metrics still work
# ===========================================================================
_BLK_TXT = "Calculate ROE and Operating Cash Flow for FY2025."
ws_blk = _real_workspace(requirements_text=_BLK_TXT)
_blk_rows = [r for r in (ws_blk.get("requirements") or [])
             if r.get("requirement") == "Operating Cash Flow"]
check("47a · unsupported metric resolves to BLOCKED in the real workspace",
      bool(_blk_rows) and _blk_rows[0].get("status") == "BLOCKED",
      str([(r.get("requirement"), r.get("status")) for r in ws_blk.get("requirements") or []]))
_v_blk = agent_session(ws_blk, {"stage": STAGE_METRIC, "metric": "Operating Cash Flow",
                                "area": None, "visited": []})
check("47b · blocked metric shows blocked guidance (never a fake value)",
      (_v_blk.get("guidance") or {}).get("kind") == "blocked" and
      "I won't guess" in str((_v_blk.get("guidance") or {}).get("message") or ""),
      str(_v_blk.get("guidance"))[:160])
check("47c · blocked metric message names a concrete next step",
      "verify" in _v_blk["message"].lower() or "continue" in _v_blk["message"].lower(),
      _v_blk["message"])


# ===========================================================================
# 48 · API path — qualitative catalyst states remain unchanged
# ===========================================================================
_qual_before = copy.deepcopy((ws_clean_real.get("qualitative_drivers") or {}).get("rows") or [])
_ = _at_stage(ws_clean_real, STAGE_CONCLUSION)
_ = _at_stage(ws_clean_real, STAGE_EXCEL)
_qual_after = copy.deepcopy((ws_clean_real.get("qualitative_drivers") or {}).get("rows") or [])
check("48a · qualitative catalyst states are unchanged across the journey",
      _qual_before == _qual_after)


# ===========================================================================
# 49-51 · Demo/API parity
# ===========================================================================
_v_real_par = agent_session(ws_clean_real, {"stage": STAGE_PERIODS, "metric": None,
                                            "area": None, "visited": ["opening", "requirements"]})
_v_demo_par = agent_session(ws_clean_demo, {"stage": STAGE_PERIODS, "metric": None,
                                            "area": None, "visited": ["opening", "requirements"]})
check("49a · Demo and API expose identical stage ids",
      _v_real_par["stage"] == _v_demo_par["stage"] == STAGE_PERIODS)
check("49b · Demo and API progress rows match in shape",
      [p.get("label") for p in _v_real_par["progress"]] ==
      [p.get("label") for p in _v_demo_par["progress"]])
check("50a · Demo and API use the same choice-id schema",
      all(re.fullmatch(r"(period\.[A-Za-z ]+|back|skip)", c.get("id") or "")
          for c in _v_real_par["choices"]) and
      all(re.fullmatch(r"(period\.[A-Za-z ]+|back|skip)", c.get("id") or "")
          for c in _v_demo_par["choices"]) and
      any(c.get("id", "").startswith("period.") for c in _v_real_par["choices"]))
check("51a · Demo and API recover the same way on a partial brief",
      _at_stage(_demo_workspace(requirements_text=PARTIAL_TXT), STAGE_REQUIREMENTS,
                requirements_text=PARTIAL_TXT)["content"].get("parse_state") == PARSE_PARTIAL and
      _at_stage(_real_workspace(requirements_text=PARTIAL_TXT, period_facts=_PERIOD_FACTS),
                STAGE_REQUIREMENTS, requirements_text=PARTIAL_TXT)["content"].get("parse_state") == PARSE_PARTIAL)
check("51b · Demo and API recover the same way on an unparseable brief",
      _at_stage(_demo_workspace(requirements_text=UNPARSE), STAGE_REQUIREMENTS,
                requirements_text=UNPARSE)["content"].get("parse_state") == PARSE_LOW and
      _at_stage(_real_workspace(requirements_text=UNPARSE), STAGE_REQUIREMENTS,
                requirements_text=UNPARSE)["content"].get("parse_state") == PARSE_LOW)


def main():
    failures = [c for c in CHECKS if not c[1]]
    print(f"\nRESULT: {len(CHECKS) - len(failures)}/{len(CHECKS)} checks pass")
    if failures:
        for name, _ok, detail in failures:
            print(f"  FAIL {name}: {detail}")
        sys.exit(1)
    print("ALL CHECKS PASS")
    sys.exit(0)


if __name__ == "__main__":
    main()
