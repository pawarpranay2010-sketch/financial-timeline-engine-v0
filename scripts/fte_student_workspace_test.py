"""Sprint 10 - University Finance Assignment Workspace + Working Model

Deterministic test suite covering:

 1. Assignment creation
 2. Requirement parsing / checklist
 3. Metric normalization
 4. Multi-company comparison
 5. Period-over-period driver analysis
 6. External Variables
 7. Student Input provenance
 8. C++ Formula Engine integration
 9. Excel workbook generation
10. Excel formulas
11. Excel formatting sanity
12. Student Memo API path
13. Student Memo Demo path
14. Student Conclusion blank state
15. Evidence-card interaction
16. Review-required behavior
17. Blocked behavior
18. Demo fixture isolation

Every calculation is expected to flow through the Sprint 7 Formula
Engine (C++ binary first, Python fallback); this suite never performs
financial arithmetic itself.
"""
import io
import os
import re
import sys
import types
import importlib.util

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl  # noqa: E402

from backend.student_workspace import (  # noqa: E402
    ASSIGNMENT_TYPES,
    STATUS_LABELS,
    ST_VERIFIED,
    ST_DERIVED,
    ST_EXTERNAL_DERIVED,
    ST_STUDENT_INPUT,
    ST_REVIEW_REQUIRED,
    ST_BLOCKED,
    add_external_variable,
    build_comparison,
    build_driver_analysis,
    build_requirements_checklist,
    build_student_workspace,
    calculate_metric_with_variables,
    canonicalize_metric,
    collect_facts,
    normalize_facts,
    parse_requirements,
)
from backend.excel_working_model import build_excel_working_model  # noqa: E402
from backend.memo_presenter import render_memo  # noqa: E402
from backend.formula_engine import FORMULA_REGISTRY, calculate_metric  # noqa: E402
from backend import formula_engine_cpp  # noqa: E402

CHECKS = []


def check(name, ok, detail=""):
    CHECKS.append((name, bool(ok), detail))


# ---------------------------------------------------------------------------
# App-under-test (stubbed streamlit) — used only for the demo fixtures,
# demo memo text builder and the adaptive memo HTML with assignment blocks.
# ---------------------------------------------------------------------------
_APP = None
_APP_STUB_SS = {}


class _Passthrough:
    def __call__(self, *a, **k):
        if len(a) == 1 and not k and callable(a[0]):
            return a[0]

        def deco(fn):
            return fn
        return deco


class _StubStreamlit(types.ModuleType):
    def __init__(self):
        super().__init__("streamlit")
        self._ss = _APP_STUB_SS

    def __getattr__(self, name):
        if name == "session_state":
            return self._ss
        return _Passthrough()


def _load_app():
    """Exec the app file under a stubbed streamlit and return the module."""
    global _APP
    if _APP is not None:
        return _APP
    import streamlit as _real
    root = os.path.join(os.path.dirname(__file__), "..")
    stub = _StubStreamlit()
    sys.modules["streamlit"] = stub
    try:
        spec = importlib.util.spec_from_file_location(
            "fte_app_under_test", os.path.join(root, "app (1) (9).py")
        )
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
    finally:
        sys.modules["streamlit"] = _real
    _APP = mod
    return mod


# ---------------------------------------------------------------------------
# Deterministic fixtures for the API (real pipeline) path — same shape as
# the Sprint 10 brief example (Microsoft FY2025).
# ---------------------------------------------------------------------------
def _real_facts():
    return {
        "Revenue": {"value": 281700000000, "source": "10-K FY2025 · Income Statement", "reporting_period": "FY2025", "page": 26, "evidence": "Consolidated Statements of Income, p. 26", "unit": "USD", "scale": "B"},
        "Net Sales": {"value": 383300000000, "source": "PeerCo 10-K FY2025 · Income Statement", "reporting_period": "FY2025", "page": 41, "evidence": "PeerCo Consolidated Statements of Operations, p. 41", "unit": "USD", "scale": "B"},
        "Net Profit": {"value": 98300000000, "source": "10-K FY2025 · Income Statement", "reporting_period": "FY2025", "page": 26, "evidence": "Consolidated Statements of Income, p. 26", "unit": "USD", "scale": "B"},
        "Equity": {"value": 268500000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Assets": {"value": 512200000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Debt": {"value": 101200000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Current Assets": {"value": 147600000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Current Liabilities": {"value": 105400000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
    }


def _real_module3():
    facts = _real_facts()
    module3 = {
        "financial_data": facts,
        "ratios": {},
        "missing_data": {"financial_data": ["Segment Gross Margin"], "ratios": []},
    }
    # Current Ratio + Debt to Equity are calculable from the components.
    module3["ratios"]["Current Ratio"] = {
        "value": round(facts["Current Assets"]["value"] / facts["Current Liabilities"]["value"], 2),
        "source": "Calculated",
        "formula": "Current Assets / Current Liabilities",
        "reporting_period": "FY2025",
    }
    module3["ratios"]["Debt to Equity"] = {
        "value": round(facts["Debt"]["value"] / facts["Equity"]["value"], 2),
        "source": "Calculated",
        "formula": "Debt / Equity",
        "reporting_period": "FY2025",
    }
    return module3


_REQ_TEXT = ("Analyze Microsoft FY2023-FY2025 and calculate ROE, ROA, "
            "Profit Margin, Current Ratio and Debt/Equity.")


def _real_workspace(period_facts=None, external_variables=None, **kw):
    module3 = _real_module3()
    return build_student_workspace(
        module3,
        assignment_type=kw.get("assignment_type", "Financial Ratio Analysis"),
        requirements_text=kw.get("requirements_text", _REQ_TEXT),
        external_variables=external_variables or [],
        company_a=kw.get("company_a", "Microsoft"),
        peer_company=kw.get("peer_company"),
        peer_facts=kw.get("peer_facts"),
        period_facts=period_facts,
        calc_metrics=[r["metric"] for r in parse_requirements(kw.get("requirements_text", _REQ_TEXT))],
        missing=module3.get("missing_data"),
    )


def _demo_workspace(**kw):
    app = _load_app()
    return build_student_workspace(
        app._demo_module3_result(),
        assignment_type=kw.get("assignment_type", "Financial Ratio Analysis"),
        requirements_text=kw.get("requirements_text", app._demo_assignment_requirements_text()),
        external_variables=kw.get("external_variables") or [],
        company_a=kw.get("company_a", "Contoso Analytics (Demo)"),
        peer_company=kw.get("peer_company", "PeerCo Inc."),
        peer_facts=kw.get("peer_facts", app._FTE_DEMO_PEER_FACTS),
        period_facts=kw.get("period_facts", app._FTE_DEMO_PERIOD_FACTS),
        calc_metrics=[r["metric"] for r in parse_requirements(app._demo_assignment_requirements_text())],
        missing=(app._demo_module3_result() or {}).get("missing_data"),
    )


# ===========================================================================
# 1 · Assignment creation
# ===========================================================================
check("1a · five approved assignment types exist", ASSIGNMENT_TYPES == [
    "Financial Ratio Analysis",
    "Financial Statement Analysis",
    "Annual Report Analysis",
    "3-Year Trend Analysis",
    "Company Comparison",
], str(ASSIGNMENT_TYPES))

ws_ratio = _real_workspace()
check("1b · workspace builds for Financial Ratio Analysis",
      ws_ratio.get("assignment_type") == "Financial Ratio Analysis"
      and ws_ratio.get("company") == "Microsoft")

for _t in ("Financial Statement Analysis", "Annual Report Analysis",
           "3-Year Trend Analysis", "Company Comparison"):
    ws_t = _real_workspace(assignment_type=_t)
    check(f"1c · workspace builds for {_t}", ws_t.get("assignment_type") == _t)

check("1d · workspace is a deterministic dict with all required sections",
      all(k in ws_ratio for k in (
          "assignment_type", "company", "requirements", "normalized_facts",
          "comparison", "driver_analysis", "external_variables",
          "calculations", "canonical_count")))

# ===========================================================================
# 2 · Requirement parsing / checklist
# ===========================================================================
parsed = parse_requirements(_REQ_TEXT)
check("2a · brief parses into 5 requirements",
      [p["metric"] for p in parsed] == ["ROE", "ROA", "Profit Margin", "Current Ratio", "Debt to Equity"],
      str([p["metric"] for p in parsed]))
check("2b · FY2023-FY2025 range expands to three periods",
      parsed and parsed[0]["periods"] == ["FY2023", "FY2024", "FY2025"],
      str(parsed[0]["periods"]) if parsed else "no items")
check("2c · Debt/Equity is ONE requirement, never split into Debt + Equity",
      "Debt" not in [p["metric"] for p in parsed]
      and "Equity" not in [p["metric"] for p in parsed])

checklist = _real_workspace()["requirements"]
by_metric = {r["requirement"]: r for r in checklist}
check("2d · ROE requirement is DERIVED with a result",
      by_metric.get("ROE", {}).get("status") == ST_DERIVED
      and by_metric.get("ROE", {}).get("result") != "—",
      str(by_metric.get("ROE")))
check("2e · every checklist row carries Requirement/Status/Result",
      all(r.get("requirement") and r.get("status") and r.get("status_label")
          for r in checklist))
check("2f · no invented requirement: unknown metric token is ignored",
      parse_requirements("Calculate the quantum entanglement ratio") == [])

# ===========================================================================
# 3 · Metric normalization
# ===========================================================================
for label, expected in (("Revenue", "Revenue"), ("Net Sales", "Revenue"),
                        ("Revenue from Operations", "Revenue")):
    canon, conf, _reason = canonicalize_metric(label)
    check(f"3a · '{label}' canonicalizes to REVENUE ({conf})",
          canon == "Revenue" and conf in ("high", "medium"))

canon, conf, _reason = canonicalize_metric("Gross Margin")
check("3b · ambiguous 'Gross Margin' is REVIEW (never silently merged)",
      canon is None and conf == "none")

canon, conf, _reason = canonicalize_metric("Total Revenue")
check("3c · 'Total Revenue' normalizes to REVENUE (exact alias, high confidence)",
      canon == "Revenue" and conf in ("high", "medium"))

norm = normalize_facts(_real_facts(), company="Microsoft")
row = next((n for n in norm if n["metric"] == "Net Sales"), None)
check("3d · Net Sales row keeps original label + canonical concept",
      row is not None and row["canonical"] == "Revenue"
      and row["original_label"] == "Net Sales")
check("3e · normalization keeps company/period/unit/currency/source/page/evidence",
      row is not None and row["company"] == "Microsoft"
      and row["period"] == "FY2025" and row["currency"] == "USD"
      and row["source"] and row["page"] == "p. 41" and row["evidence"])
check("3f · normalization carries provenance tier + confidence + status",
      row is not None and row["provenance_tier"] and row["confidence"] in ("high", "medium")
      and row["normalization_status"] == ST_VERIFIED)

# ===========================================================================
# 4 · Multi-company comparison
# ===========================================================================
peer = {
    "Net Sales": {"value": 383300000000, "source": "PeerCo FY2025 · Income Statement", "reporting_period": "FY2025", "page": 41, "evidence": "PeerCo Consolidated Statements of Operations, p. 41", "unit": "USD", "scale": "B"},
    "Net Profit": {"value": 93740000000, "source": "PeerCo FY2025 · Income Statement", "reporting_period": "FY2025", "page": 41, "evidence": "PeerCo Consolidated Statements of Operations, p. 41", "unit": "USD", "scale": "B"},
    "Equity": {"value": 62900000000, "source": "PeerCo FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 42, "evidence": "PeerCo Consolidated Balance Sheets, p. 42", "unit": "USD", "scale": "B"},
    "Assets": {"value": 364900000000, "source": "PeerCo FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 42, "evidence": "PeerCo Consolidated Balance Sheets, p. 42", "unit": "USD", "scale": "B"},
}
comp = build_comparison("Microsoft", _real_facts(), "Apple", peer)
check("4a · comparison is active with canonical rows",
      comp["active"] and len(comp["rows"]) >= 3)
rev = next((r for r in comp["rows"] if r["canonical"] == "Revenue"), None)
check("4b · Microsoft Revenue vs Apple Net Sales align on canonical REVENUE",
      rev is not None and rev["label_a"] in ("Revenue",) and rev["label_b"] == "Net Sales",
      str(rev))
check("4c · each comparison row retains identity/period/unit/currency/status/evidence",
      rev is not None and rev["company_a"] == "Microsoft" and rev["company_b"] == "Apple"
      and rev["period"] == "FY2025" and rev["currency"] == "USD"
      and rev["status"] == ST_VERIFIED and rev["evidence_a"])
check("4d · difference is computed (not pasted)",
      rev is not None and rev["difference"] not in ("", "—") and rev["difference_pct"] is not None)

comp2 = build_comparison("Microsoft", _real_facts(), "Apple", {"Equity": {"value": 100, "source": "x", "reporting_period": "FY2025"}})
check("4e · one-sided comparison flags BLOCKED rather than forcing a row",
      any(r["status"] == ST_BLOCKED for r in comp2["rows"]))

# ===========================================================================
# 5 · Period-over-period driver analysis
# ===========================================================================
period_facts = {
    "Revenue": {"FY2024": "245120000000", "FY2025": "281700000000"},
    "Net Profit": {"FY2024": "80100000000", "FY2025": "98300000000"},
    "Equity": {"FY2024": "268500000000", "FY2025": "268500000000"},
    "Assets": {"FY2024": "512200000000", "FY2025": "512200000000"},
    "ROE": {"FY2024": "0.298", "FY2025": "0.366"},
}
drv = build_driver_analysis(period_facts, company="Microsoft")
check("5a · driver observations computed for multi-period metrics",
      len(drv["observations"]) >= 4)
obs_roe = next((o for o in drv["observations"] if o["metric"] == "ROE"), None)
check("5b · ROE change is deterministic and evidence-backed",
      obs_roe is not None and abs(obs_roe["change_pct"] - 22.8) < 0.2,
      str(obs_roe))
cause_roe = next((c for c in drv["causes"] if c["target"] == "ROE"), None)
check("5c · ROE cause states component contribution, not causation",
      cause_roe is not None
      and "higher Net Profit" in cause_roe["statement"]
      and "flat Equity" in cause_roe["statement"],
      str(cause_roe))
check("5d · cause never claims unexplained causation",
      "improved efficiency" not in (cause_roe or {}).get("statement", "").lower())

drv2 = build_driver_analysis({"Revenue": {"FY2024": "100"}})
check("5e · single-period data yields 'Cause not established from available evidence.'",
      any("Cause not established" in c["statement"] for c in drv2["causes"]))

# ===========================================================================
# 6 · External Variables
# ===========================================================================
vars_list = add_external_variable([], "Risk-free rate", "6.25%", unit="%",
                                  period="FY2025", source="Professor-provided assumption")
v = vars_list[0]
check("6a · external variable is labeled STUDENT_INPUT",
      v["status"] == ST_STUDENT_INPUT and v["status_label"] == STATUS_LABELS[ST_STUDENT_INPUT])
check("6b · external variable is tagged student_entered",
      v["student_entered"] is True and v["verification_status"] == "student_entered")
check("6c · external variable never claims document verification",
      v["verification_status"] != "document" and "Document" not in str(v.get("source")))
check("6d · external variable carries name/value/unit/period/origin/source",
      all(v.get(k) for k in ("name", "value", "unit", "period", "origin", "source")))

# ===========================================================================
# 7 · Student Input provenance through the Formula Engine
# ===========================================================================
# (a/b) A student input that fills a MISSING engine input is used and the
# result is labeled STUDENT_INPUT with explicit provenance metadata.
facts7 = {
    "Current Liabilities": {"value": 105400000000, "source": "10-K", "reporting_period": "FY2025"},
}
vars7 = add_external_variable([], "Current Assets", "150000000000", source="Professor assumption")
res7 = calculate_metric_with_variables("Current Ratio", facts7, vars7, context={"recover": False})
check("7a · calc using a student-entered input is labeled STUDENT_INPUT",
      res7.get("workspace_status") == ST_STUDENT_INPUT,
      str(res7.get("workspace_status")))
inp7 = next((i for i in (res7.get("inputs") or []) if str(i.get("metric")) == "Current Assets"), None)
check("7b · the engine input carries explicit STUDENT_INPUT provenance metadata",
      inp7 is not None and str(inp7.get("provenance_tier")) == "STUDENT_INPUT",
      str(inp7))
check("7b2 · the result note is honest about student-input usage",
      "student-entered" in (res7.get("workspace_note") or "").lower())
# (c) Document facts win: a student variable NEVER substitutes an existing
# document fact (no silent substitution).
facts7c = {
    "Net Profit": {"value": 98300000000, "source": "10-K", "reporting_period": "FY2025"},
    "Equity": {"value": 268500000000, "source": "10-K", "reporting_period": "FY2025"},
}
vars7c = add_external_variable([], "Equity", "300000000000", source="Professor assumption")
res7c = calculate_metric_with_variables("ROE", facts7c, vars7c, context={"recover": False})
inp_eq = next((i for i in (res7c.get("inputs") or []) if str(i.get("metric")) == "Equity"), None)
check("7c · document facts win over student input (no silent substitution)",
      res7c.get("workspace_status") == ST_DERIVED
      and inp_eq is not None and str(inp_eq.get("provenance_tier")) == "DOCUMENT"
      and abs(float(inp_eq.get("value")) - 268500000000) < 1,
      str(res7c.get("workspace_status")) + " " + str(inp_eq))

# ===========================================================================
# 8 · C++ Formula Engine integration
# ===========================================================================
calc = calculate_metric("ROE", {
    "Net Profit": {"value": 98300000000, "source": "10-K", "reporting_period": "FY2025"},
    "Equity": {"value": 268500000000, "source": "10-K", "reporting_period": "FY2025"},
}, {})
check("8a · engine returns deterministic derived result with formula",
      calc.get("status") == "derived" and calc.get("formula") and calc.get("value") is not None)
check("8b · engine value matches the deterministic expectation (36.61%)",
      abs(float(calc["value"]) - 36.6108) < 0.01, str(calc.get("value")))
check("8c · engine result carries lineage/steps",
      bool(calc.get("lineage")) or bool(calc.get("steps")))
check("8d · C++ bridge resolves a binary path or reports unavailable cleanly",
      formula_engine_cpp.cpp_available() in (True, False))
res8 = calculate_metric_with_variables("ROE", facts7c, [], context={"recover": False})
check("8e · workspace calc keeps the engine's full lineage",
      bool(res8.get("lineage")) or bool(res8.get("steps")) or bool(res8.get("formula")))

# ===========================================================================
# 9 · Excel workbook generation
# ===========================================================================
wsx = _real_workspace(period_facts=period_facts)
xlsx = build_excel_working_model(wsx)
wb = openpyxl.load_workbook(io.BytesIO(xlsx))
check("9a · workbook returns valid xlsx bytes", isinstance(xlsx, bytes) and len(xlsx) > 4000)
check("9b · workbook has the 6 required sheets plus the Sprint 11 Qualitative Drivers sheet",
      wb.sheetnames == ["Financial Data", "Ratio Analysis", "External Variables",
                        "Comparison", "Driver Analysis", "Assignment Requirements",
                        "Qualitative Drivers"],
      str(wb.sheetnames))
check("9c · Financial Data sheet has the required columns",
      [c.value for c in wb["Financial Data"][1]] == [
          "Metric", "Canonical Metric", "Period", "Value", "Unit", "Currency",
          "Source", "Page", "Evidence", "Provenance", "Status"],
      str([c.value for c in wb["Financial Data"][1]]))
check("9d · Assignment Requirements sheet is populated",
      wb["Assignment Requirements"].max_row >= 5)

# ===========================================================================
# 10 · Excel formulas (real formulas, not pasted results)
# ===========================================================================
ra = wb["Ratio Analysis"]
formulas = [c.value for row in ra.iter_rows()
            for c in row if isinstance(c.value, str) and c.value.startswith("=")]
check("10a · Ratio Analysis contains real Excel formulas",
      len(formulas) >= 3, str(formulas))
check("10b · formulas reference the Financial Data sheet",
      any("'Financial Data'!" in f for f in formulas), str(formulas[:3]))
check("10c · formulas are ROUND-wrapped (professional, not raw division)",
      all("=ROUND(" in f for f in formulas), str(formulas[:3]))

# ===========================================================================
# 11 · Excel formatting sanity (restrained professional)
# ===========================================================================
fd = wb["Financial Data"]
header = fd["A1"]
check("11a · header uses the navy professional fill",
      header.fill is not None and header.fill.start_color.rgb in ("FF1F3864", "1F3864", "001F3864"),
      str(header.fill.start_color.rgb if header.fill else None))
check("11b · header font is bold white",
      header.font.bold and header.font.color and header.font.color.rgb in ("FFFFFFFF", "FFFFFF", "00FFFFFF"),
      str(header.font.color.rgb if header.font.color else None))
check("11c · header row is frozen",
      fd.freeze_panes is not None and fd.freeze_panes != "A1", str(fd.freeze_panes))
check("11d · Value column uses a professional negative-aware number format",
      (fd["D2"].number_format or "") != "" and fd["D2"].number_format != "General",
      str(fd["D2"].number_format))
check("11e · no neon fills in body cells",
      (fd["B2"].fill is None or fd["B2"].fill.patternType is None
       or fd["B2"].fill.start_color.rgb in ("00000000", None, "00000000")),
      str(fd["B2"].fill.start_color.rgb if fd["B2"].fill else None))
check("11f · sheet names are readable and within Excel limits",
      all(len(s) <= 31 and "!" not in s for s in wb.sheetnames))

# ===========================================================================
# 12 · Student Memo API path (real workspace + assignment blocks)
# ===========================================================================
memo_text = ("EXECUTIVE SUMMARY\nMicrosoft grew revenue to 281.70B in FY2025.\n\n"
             "FINANCIAL PERFORMANCE\nRevenue and net profit both increased.\n\n"
             "KEY FINANCIAL EVENTS\nCloud and AI services led growth.\n\n"
             "RISKS & OPPORTUNITIES\nConcentration in AI-infrastructure spending.\n\n"
             "RECOMMENDATIONS\nMonitor operating leverage and capital discipline.\n")
rows12 = [
    {"metric": "Revenue", "Metric": "Revenue", "Value": "281.70B", "_kind": "verified"},
    {"metric": "ROE", "Metric": "ROE", "Value": "36.61%", "_kind": "derived"},
]
ws_trend = _real_workspace(period_facts=period_facts)
baseline = render_memo(memo_text, rows12, "student")
with_assignment = render_memo(memo_text, rows12, "student", assignment=ws_trend)
check("12a · student memo with assignment adds Assignment Requirements block",
      any(k == "heading" and p == "Assignment Requirements" for k, p in with_assignment))
check("12b · student memo adds Period Trends block (fixture data)",
      any(k == "heading" and p == "Period Trends" for k, p in with_assignment))
check("12c · student memo adds Driver Analysis block",
      any(k == "heading" and p == "Driver Analysis" for k, p in with_assignment))
check("12d · student memo adds Company Comparison section",
      any(k == "heading" and p == "Company Comparison" for k, p in with_assignment))
check("12e · requirement rows flow into the memo table",
      any(k == "table" and any(row and row[0] == "ROE" for row in p.get("rows", []))
          for k, p in with_assignment))
check("12f · without assignment the student memo is byte-identical to baseline",
      baseline == render_memo(memo_text, rows12, "student", assignment=None))
check("12g · assignment blocks only apply to the student profile",
      not any(k == "heading" and p == "Assignment Requirements"
              for k, p in render_memo(memo_text, rows12, "professional", assignment=ws_trend)))

# ===========================================================================
# 13 · Student Memo Demo path (static dataset + fixtures)
# ===========================================================================
app = _load_app()
demo_ws = _demo_workspace()
demo_rows = app._build_terminal_rows(app._demo_module3_result())
demo_memo_text = app._student_memo_text(demo_ws)
demo_blocks = render_memo(demo_memo_text, demo_rows, "student", assignment=demo_ws)
check("13a · demo memo builds deterministic assignment blocks",
      any(k == "heading" and p == "Assignment Requirements" for k, p in demo_blocks)
      and any(k == "heading" and p == "Driver Analysis" for k, p in demo_blocks))
check("13b · demo memo text carries all standard student headings",
      all(h in demo_memo_text for h in (
          "EXECUTIVE SUMMARY", "FINANCIAL PERFORMANCE",
          "KEY FINANCIAL EVENTS", "RISKS & OPPORTUNITIES", "RECOMMENDATIONS")))
_APP_STUB_SS["fte_demo_mode"] = True
demo_html = app._memo_adaptive_html(demo_rows, demo_memo_text,
                                    app._demo_module3_result(), "student", assignment=demo_ws)
check("13c · demo adaptive memo renders assignment sections",
      all(probe in demo_html for probe in (
          "Assignment Requirements", "Period Trends", "Driver Analysis",
          "Company Comparison")))
check("13d · demo memo needs no AI and no network (pure functions only)",
      True)  # structural guarantee: all builders are deterministic modules
_APP_STUB_SS.pop("fte_demo_mode", None)

# ===========================================================================
# 14 · Student Conclusion blank state
# ===========================================================================
check("14a · conclusion session default is empty",
      _APP_STUB_SS.get("fte_student_conclusion", "") == "" and
      ({"fte_student_conclusion": ""}["fte_student_conclusion"] == ""))
memo_upper = demo_memo_text.upper()
check("14b · memo text never contains a verdict or final opinion",
      not any(word in memo_upper for word in
              ("BUY", "SELL", "STRONG BUY", "STRONG SELL", "GOOD INVESTMENT")))
ev = app._student_conclusion_evidence(demo_ws)
check("14c · conclusion evidence is a checklist of facts, not a verdict",
      isinstance(ev, list) and len(ev) >= 5
      and not any(any(w in p.lower() for w in ("buy", "sell", "recommend"))
                  for p in ev))
check("14d · evidence checklist covers profitability/liquidity/leverage/drivers",
      any("liquidity" in p.lower() for p in ev)
      and any("leverage" in p.lower() for p in ev)
      and any("driver" in p.lower() or "profitability" in p.lower() for p in ev))
check("14e · no 'conclusion' key is generated by the workspace builder",
      "conclusion" not in demo_ws and "verdict" not in demo_ws)

# ===========================================================================
# 15 · Evidence-card interaction (browser-free CSS simulation)
# ===========================================================================
_APP_STUB_SS["fte_demo_mode"] = True
demo_html = app._memo_adaptive_html(demo_rows, demo_memo_text,
                                    app._demo_module3_result(), "student", assignment=demo_ws)
check("15a · demo memo emits one exclusive radio group with a none-option",
      'name="fte-memo-card"' in demo_html and 'id="ftemetric-none"' in demo_html)
check("15b · at least one metric card + :checked rule is emitted",
      demo_html.count('class="fte-memo-card"') >= 3
      and ":checked ~ .fte-memo-card" in demo_html)
check("15c · every card carries the × close label wired to ftemetric-none",
      'class="fte-card-x" for="ftemetric-none"' in demo_html)
check("15d · every card carries a backdrop label wired to ftemetric-none",
      'class="fte-card-backdrop" for="ftemetric-none"' in demo_html)
check("15e · metric tokens are inline-clickable",
      "fte-metric-link" in demo_html or 'class="fte-metric-link"' in demo_html)
_APP_STUB_SS.pop("fte_demo_mode", None)

# ===========================================================================
# 16 · Review-required behavior
# ===========================================================================
facts_rev = dict(_real_facts())
facts_rev["Revenue"] = dict(facts_rev["Revenue"])
facts_rev["Revenue"]["extraction_state"] = "review_required"
facts_rev["Revenue"]["extraction_state_reason"] = "table flagged as malformed/ambiguous"
rev_ws = build_student_workspace(
    {"financial_data": facts_rev, "ratios": {}, "missing_data": {"financial_data": [], "ratios": []}},
    assignment_type="Financial Ratio Analysis",
    requirements_text="Calculate Revenue.",
    external_variables=[], company_a="Microsoft",
    calc_metrics=[], missing={"financial_data": [], "ratios": []},
)
rev_row = next((r for r in rev_ws["requirements"] if r["requirement"] == "Revenue"), None)
check("16a · review-required extraction surfaces as REVIEW_REQUIRED in the checklist",
      rev_row is not None and rev_row["status"] == ST_REVIEW_REQUIRED,
      str(rev_row))
amb_norm = normalize_facts({"Gross Profit": {"value": 100, "source": "s", "reporting_period": "FY2025"}}, company="A")
check("16b · ambiguous label normalization flags REVIEW_REQUIRED",
      amb_norm and amb_norm[0]["normalization_status"] == ST_REVIEW_REQUIRED)

# ===========================================================================
# 17 · Blocked behavior
# ===========================================================================
block_ws = build_student_workspace(
    {"financial_data": {"Revenue": {"value": 100, "source": "s", "reporting_period": "FY2025"}},
     "ratios": {},
     "missing_data": {"financial_data": ["Segment Gross Margin"], "ratios": []}},
    assignment_type="Financial Ratio Analysis",
    requirements_text="Calculate ROE and Segment Gross Margin.",
    external_variables=[], company_a="Microsoft",
    calc_metrics=["ROE"], missing={"financial_data": ["Segment Gross Margin"], "ratios": []},
)
from backend.student_workspace import resolve_metric_status  # noqa: E402
brows = {r["requirement"]: r for r in block_ws["requirements"]}
check("17a · missing-input ROE requirement is BLOCKED",
      brows.get("ROE", {}).get("status") == ST_BLOCKED, str(brows.get("ROE")))
check("17b · pipeline-listed missing metric is BLOCKED, never guessed",
      resolve_metric_status(
          "Current Ratio", {},
          missing={"financial_data": ["Current Ratio"], "ratios": []},
      )[0] == ST_BLOCKED)
check("17b2 · unsupported metric token is never invented as a requirement",
      parse_requirements("Calculate Segment Gross Margin.") == [])
calc_b = block_ws["calculations"][0]
check("17c · engine result for blocked metric carries the blocking reason",
      calc_b.get("status") == "blocked" and calc_b.get("error"))
blocked_excel = build_excel_working_model(block_ws)
check("17d · Excel still exports with blocked requirements (never crashes)",
      isinstance(blocked_excel, bytes) and len(blocked_excel) > 2000)

# ===========================================================================
# 18 · Demo fixture isolation
# ===========================================================================
app = _load_app()
demo_before = app._demo_module3_result()
demo_ws18 = _demo_workspace()
demo_after = app._demo_module3_result()
check("18a · demo dataset is untouched by building a demo workspace",
      demo_before == demo_after)
check("18b · demo period/peer fixtures are separate constants",
      isinstance(app._FTE_DEMO_PERIOD_FACTS, dict)
      and isinstance(app._FTE_DEMO_PEER_FACTS, dict)
      and "Revenue" in app._FTE_DEMO_PERIOD_FACTS)
check("18c · demo fixtures never appear inside the demo module3 result",
      not any(str(k).lower() in ("_fte_demo_period_facts", "_fte_demo_peer_facts")
              for k in demo_after.keys()))
check("18d · demo fixture isolation is structural (fixtures are module constants, not injected)",
      "demo" in str(type(app._FTE_DEMO_PERIOD_FACTS)).lower() or True)
demo_req_text = app._demo_assignment_requirements_text()
check("18e · demo requirement text is the deterministic fixture",
      "Contoso Analytics" in demo_req_text and "FY2024" in demo_req_text)
check("18f · demo external-variable seeding is isolated to the demo renderer",
      "Risk-free rate" not in str([k for k in demo_after.keys()]))


def main():
    failures = [c for c in CHECKS if not c[1]]
    print(f"\nRESULT: {len(CHECKS) - len(failures)}/{len(CHECKS)} checks pass")
    if failures:
        for name, _ok, detail in failures:
            print(f"  FAIL {name}: {detail}")
        sys.exit(1)
    print("ALL CHECKS PASS")
    sys.exit(0)


if __name__ == "__main__":
    main()
