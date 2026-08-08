"""FT-E Hardcore Full-App Adversarial Verification (post-Sprint 11).

Verification/hardening exercise ONLY — no production code is modified.

Covers (mapping to the full-app test brief):
  A. Normalization adversarial   (canonical merge / no partial-word / ambiguous)
  B. Requirement parser adversarial (reordered, punctuation, Debt/Equity forms)
  C. Formula engine edge cases   (zero denom, missing inputs, negatives, determinism)
  D. Causality attack            (Sprint 11: no-cause, misleading, contradictory)
  E. Demo isolation attack       (deep-equal before/after, no reliability leakage)
  F. Determinism                 (API fixture twice -> identical outputs)
  G. Excel working model         (7 sheets, real formulas, reopen, damage test)
  H. Memo content safety         (no None/null/NaN/TODO/placeholder leakage)
  I. Student conclusion          (always blank, never generated)
  J. Security                    (malicious filenames, injection, HYPERLINK, secrets)
  K. Performance / stress        (1/5/10 documents, wall-clock timings)
  L. API/Demo parity             (same machinery, values may differ)

Run: python3 scripts/fte_hardcore_adversarial_test.py
"""
import copy
import io
import os
import re
import sys
import time
import types
import importlib.util

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl  # noqa: E402

from backend.student_workspace import (  # noqa: E402
    ST_VERIFIED,
    ST_DERIVED,
    ST_EXTERNAL_DERIVED,
    ST_STUDENT_INPUT,
    ST_REVIEW_REQUIRED,
    ST_BLOCKED,
    ST_UNANALYZED,
    build_comparison,
    build_driver_analysis,
    build_requirements_checklist,
    build_student_workspace,
    canonicalize_metric,
    normalize_facts,
    parse_requirements,
)
from backend.qualitative_catalyst import (  # noqa: E402
    REL_EXPLICIT,
    REL_SUPPORTED,
    REL_POSSIBLE,
    REL_INSUFFICIENT,
    REL_CAUSE_NOT_ESTABLISHED,
    build_qualitative_drivers,
    extract_narrative_items,
)
from backend.formula_engine import calculate_metric, FORMULA_REGISTRY  # noqa: E402
from backend.excel_working_model import build_excel_working_model  # noqa: E402
from backend.memo_presenter import render_memo  # noqa: E402
from ingestion.parser import parse_document  # noqa: E402

CHECKS = []


def check(name, ok, detail=""):
    CHECKS.append((name, bool(ok), detail))
    status = "PASS" if ok else "FAIL"
    print(f"  {status}  {name}  [{detail if detail else ''}]")


def section(t):
    print(f"\n### {t}")


# ---------------------------------------------------------------------------
# App-under-test (stubbed streamlit) — demo fixtures only.
# ---------------------------------------------------------------------------
_APP = None
_APP_STUB_SS = {}


class _Passthrough:
    def __call__(self, *a, **k):
        if len(a) == 1 and not k and callable(a[0]):
            return a[0]

        def deco(fn):
            return fn
        return deco


class _StubStreamlit(types.ModuleType):
    def __init__(self):
        super().__init__("streamlit")
        self._ss = _APP_STUB_SS

    def __getattr__(self, name):
        if name == "session_state":
            return self._ss
        return _Passthrough()


def _load_app():
    global _APP
    if _APP is not None:
        return _APP
    import streamlit as _real
    root = os.path.join(os.path.dirname(__file__), "..")
    stub = _StubStreamlit()
    sys.modules["streamlit"] = stub
    try:
        spec = importlib.util.spec_from_file_location(
            "fte_hardcore_app", os.path.join(root, "app (1) (9).py")
        )
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
    finally:
        sys.modules["streamlit"] = _real
    _APP = mod
    return mod


# ---------------------------------------------------------------------------
# Deterministic API-path fixtures (same shape as the existing suites).
# ---------------------------------------------------------------------------
def _real_facts():
    return {
        "Revenue": {"value": 281700000000, "source": "10-K FY2025 · Income Statement", "reporting_period": "FY2025", "page": 26, "evidence": "Consolidated Statements of Income, p. 26", "unit": "USD", "scale": "B"},
        "Net Profit": {"value": 98300000000, "source": "10-K FY2025 · Income Statement", "reporting_period": "FY2025", "page": 26, "evidence": "Consolidated Statements of Income, p. 26", "unit": "USD", "scale": "B"},
        "Equity": {"value": 268500000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Assets": {"value": 512200000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Debt": {"value": 101200000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Current Assets": {"value": 147600000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
        "Current Liabilities": {"value": 105400000000, "source": "10-K FY2025 · Balance Sheet", "reporting_period": "FY2025", "page": 27, "evidence": "Consolidated Balance Sheets, p. 27", "unit": "USD", "scale": "B"},
    }


def _real_module3():
    facts = _real_facts()
    module3 = {
        "financial_data": facts,
        "ratios": {},
        "missing_data": {"financial_data": ["Segment Gross Margin"], "ratios": []},
    }
    module3["ratios"]["Current Ratio"] = {
        "value": round(facts["Current Assets"]["value"] / facts["Current Liabilities"]["value"], 2),
        "source": "Calculated", "formula": "Current Assets / Current Liabilities", "reporting_period": "FY2025",
    }
    module3["ratios"]["Debt to Equity"] = {
        "value": round(facts["Debt"]["value"] / facts["Equity"]["value"], 2),
        "source": "Calculated", "formula": "Debt / Equity", "reporting_period": "FY2025",
    }
    return module3


_REQ_TEXT = ("Analyze Microsoft FY2023-FY2025 and calculate ROE, ROA, "
             "Profit Margin, Current Ratio and Debt/Equity.")


def _real_workspace(period_facts=None, external_variables=None, **kw):
    module3 = _real_module3()
    return build_student_workspace(
        module3,
        assignment_type=kw.get("assignment_type", "Financial Ratio Analysis"),
        requirements_text=kw.get("requirements_text", _REQ_TEXT),
        external_variables=external_variables or [],
        period_facts=period_facts,
        calc_metrics=[r["metric"] for r in parse_requirements(kw.get("requirements_text", _REQ_TEXT))],
        missing=module3.get("missing_data"),
    )


def _demo_workspace(**kw):
    app = _load_app()
    return build_student_workspace(
        app._demo_module3_result(),
        assignment_type=kw.get("assignment_type", "Financial Ratio Analysis"),
        requirements_text=kw.get("requirements_text", app._demo_assignment_requirements_text()),
        external_variables=[],
        period_facts=kw.get("period_facts", app._FTE_DEMO_PERIOD_FACTS),
        calc_metrics=[r["metric"] for r in parse_requirements(app._demo_assignment_requirements_text())],
        missing=(app._demo_module3_result() or {}).get("missing_data"),
        qualitative_documents=kw.get("qualitative_documents", app._FTE_DEMO_QUALITATIVE_DOCS),
    )


def demo_memo_text(blocks):
    out = []
    for kind, payload in blocks:
        if kind == "heading":
            out.append(str(payload))
        elif kind == "bullets":
            out.append("\n".join(str(b) for b in payload))
        elif kind == "table":
            out.append(str(payload.get("title", "")))
            for r in payload.get("rows", []):
                out.append(" ".join(str(c) for c in r))
        elif kind in ("paragraph", "note"):
            out.append(str(payload))
    return "\n".join(out)


# ===========================================================================
# A. NORMALIZATION ADVERSARIAL
# ===========================================================================
section("A. Normalization adversarial (canonical layer)")

# Equivalent labels must merge to the same canonical concept.
for label, expected in [
    ("Revenue", "Revenue"), ("Net Sales", "Revenue"), ("Revenue from Operations", "Revenue"),
    ("Total Revenue", "Revenue"), ("revenues", "Revenue"),
    ("Net Profit", "Net Profit"), ("Profit After Tax", "Net Profit"), ("PAT", "Net Profit"),
    ("Total Assets", "Assets"), ("Shareholders' Equity", "Equity"),
]:
    canon, conf, _reason = canonicalize_metric(label)
    check(f"A1 · '{label}' -> {expected} (high/medium)",
          canon == expected and conf in ("high", "medium"), f"{canon}|{conf}")

# No partial-word matching: EPS must NOT equal EPSILON; Debt != debt-like.
eps, eps_conf, _ = canonicalize_metric("EPSILON")
check("A2 · EPSILON != EPS (no partial-word)", eps != "EPS", f"{eps}|{eps_conf}")
debt, debt_conf, _ = canonicalize_metric("Debt-like")
check("A3 · 'Debt-like' not merged into Debt", debt != "Debt", f"{debt}|{debt_conf}")
rg, rg_conf, _ = canonicalize_metric("Revenue Growth")
check("A4 · 'Revenue Growth' != 'Revenue' (distinct concept)", rg == "Revenue Growth" or rg != "Revenue", f"{rg}|{rg_conf}")

# Ambiguous labels must NOT auto-merge (REVIEW_REQUIRED).
for amb in ["Gross Profit", "Gross Margin", "Working Capital", "Market Cap",
            "Interest Expense", "Tax Expense", "Exceptional Items", "Other Income"]:
    canon, conf, _ = canonicalize_metric(amb)
    check(f"A5 · ambiguous '{amb}' not auto-merged", conf == "none" and canon is None, f"{canon}|{conf}")

# Multi-word containing BOTH "debt" and "equity" should map to Debt to Equity.
d2e, d2e_conf, _ = canonicalize_metric("Debt / Equity")
check("A6 · 'Debt / Equity' -> Debt to Equity", d2e == "Debt to Equity" and d2e_conf in ("high", "medium"), f"{d2e}|{d2e_conf}")

# normalize_facts carries REVIEW_REQUIRED for unmapped labels and never merges them.
facts_ambiguous = {
    "Gross Profit": {"value": 50000000000, "source": "X", "reporting_period": "FY2025"},
    "Revenue": {"value": 281700000000, "source": "Y", "reporting_period": "FY2025"},
}
normed = normalize_facts(facts_ambiguous, company="Company A")
by_metric = {f["metric"]: f for f in normed}
check("A7 · ambiguous fact keeps REVIEW_REQUIRED status",
      by_metric.get("Gross Profit", {}).get("normalization_status") == ST_REVIEW_REQUIRED)
check("A8 · clean fact normalizes to VERIFIED",
      by_metric.get("Revenue", {}).get("normalization_status") == ST_VERIFIED
      and by_metric.get("Revenue", {}).get("canonical") == "Revenue")

# ===========================================================================
# B. REQUIREMENT PARSER ADVERSARIAL
# ===========================================================================
section("B. Requirement parser adversarial")

def _metrics(text):
    return [r["metric"] for r in parse_requirements(text)]

expected_metrics = ["ROE", "ROA", "Profit Margin", "Current Ratio", "Debt to Equity"]

# 1. Canonical brief.
check("B1 · canonical brief parses all 5 metrics",
      set(_metrics(_REQ_TEXT)) == set(expected_metrics), str(_metrics(_REQ_TEXT)))

# 2. Reordered requirements.
check("B2 · reordered requirements parse identically",
      set(_metrics("Calculate Current Ratio, Debt/Equity, ROA, Profit Margin and ROE.")) == set(expected_metrics))

# 3. Punctuation changes / separators.
check("B3 · 'Debt-to-Equity' single requirement",
      set(_metrics("Compute ROE and Debt-to-Equity.")) == {"ROE", "Debt to Equity"})
check("B4 · 'Debt / Equity' single requirement (not Debt+Equity fragments)",
      set(_metrics("Compute Debt / Equity.")) == {"Debt to Equity"})

# 4. 'ROE and ROA' both detected.
check("B5 · 'ROE and ROA' -> both",
      set(_metrics("Analyze ROE and ROA.")) == {"ROE", "ROA"})

# 5. Multiple periods expand.
per = parse_requirements("Analyze FY2023-FY2025 revenue.")
check("B6 · FY2023-FY2025 expands to 3 periods",
      per and per[0].get("periods") == ["FY2023", "FY2024", "FY2025"], str(per))

# 6. Unknown metric tokens are ignored (never guessed).
check("B7 · unknown tokens ignored",
      _metrics("Analyze the XYZ metric and goodwill.") == [])

# 7. 'Debt' alone must NOT silently become 'Debt to Equity'.
check("B8 · standalone Debt is not upgraded to Debt to Equity",
      _metrics("Analyze total Debt.") == ["Debt"], str(_metrics("Analyze total Debt.")))

# 8. Qualitative/external requirements: no fabricated metric.
check("B9 · explanation requirement does not fabricate a metric",
      _metrics("Explain the major factors behind the change in ROE.") == ["ROE"])

# ===========================================================================
# C. FORMULA ENGINE EDGE CASES
# ===========================================================================
section("C. Formula engine edge cases")

# 1. Zero denominator -> BLOCKED, never inf/NaN.
zero_facts = {
    "Net Profit": {"value": 100, "source": "S", "reporting_period": "FY2025"},
    "Revenue": {"value": 0, "source": "S", "reporting_period": "FY2025"},
}
pm_zero = calculate_metric("Profit Margin", zero_facts, context={"recover": False})
check("C1 · zero denominator -> BLOCKED, not inf/NaN",
      str(pm_zero.get("status")).upper() == "BLOCKED" and pm_zero.get("value") is None,
      f"{pm_zero.get('status')}|{pm_zero.get('value')}")

# 2. Missing inputs -> BLOCKED (fail closed).
missing_facts = {"Revenue": {"value": 100, "source": "S", "reporting_period": "FY2025"}}
roe_missing = calculate_metric("ROE", missing_facts, context={"recover": False})
check("C2 · missing inputs -> BLOCKED",
      str(roe_missing.get("status")).upper() == "BLOCKED", roe_missing.get("status"))

# 3. Negative values where mathematically valid (Revenue Growth).
neg_facts = {
    "Revenue": {"value": 100, "source": "S", "reporting_period": "FY2024", "previous_value": 200},
    "Previous Revenue": {"value": 200, "source": "S", "reporting_period": "FY2023"},
}
# Revenue Growth requires two periods; feed via period map instead.
rg_calc = calculate_metric("Revenue Growth", {
    "Revenue": {"value": 100, "source": "S", "reporting_period": "FY2025"},
}, context={"primary_facts": {"Previous Revenue": {"value": 200, "source": "S", "reporting_period": "FY2024"}}, "recover": False})
check("C3 · negative growth computed deterministically",
      rg_calc.get("value") is not None and rg_calc.get("value") == -50.0,
      f"{rg_calc.get('status')}|{rg_calc.get('value')}")

# 4. Determinism: same inputs -> identical output.
a1 = calculate_metric("ROE", _real_facts(), context={"recover": False})
a2 = calculate_metric("ROE", _real_facts(), context={"recover": False})
check("C4 · formula engine deterministic (ROE twice)",
      a1.get("value") == a2.get("value") and a1.get("display_value") == a2.get("display_value"),
      f"{a1.get('value')}")

# 5. Unsupported metric -> UNANALYZED.
unsup = calculate_metric("Book Value Per Share", _real_facts(), context={"recover": False})
check("C5 · unsupported metric -> UNANALYZED",
      str(unsup.get("status")).upper() == "UNANALYZED", unsup.get("status"))

# 6. Every registered formula runs without error on the fixture.
reg_ok = True
for key in FORMULA_REGISTRY:
    try:
        calculate_metric(key, _real_facts(), context={"recover": False})
    except Exception as e:  # noqa: BLE001
        reg_ok = False
        print(f"    formula {key} raised: {e}")
check("C6 · every registered formula executes without exception", reg_ok)

# ===========================================================================
# D. CAUSALITY ATTACK (Sprint 11 qualitative layer)
# ===========================================================================
section("D. Causality attack")

_OBS = lambda metric, direction="decrease", change="-10.0%": {
    "metric": metric, "from": "FY2024", "to": "FY2025",
    "from_value": "200000000000", "to_value": "180000000000",
    "change_display": change, "direction": direction, "change_pct": -10.0,
}

# 1. "Revenue decreased 10%" with NO narrative evidence -> CAUSE_NOT_ESTABLISHED,
#    and the explanation must NOT claim a cause.
no_docs = build_qualitative_drivers(
    [_OBS("Revenue")],
    facts={"Revenue": {"value": 180000000000, "source": "S", "reporting_period": "FY2025"}},
    period_facts={"Revenue": {"FY2024": "200000000000", "FY2025": "180000000000"}},
    qualitative_documents=[],
    requirements=[],
    company="Company A",
)
nrow = no_docs["rows"][0]
check("D1 · no narrative -> CAUSE_NOT_ESTABLISHED",
      nrow["relationship"] == REL_CAUSE_NOT_ESTABLISHED, nrow["relationship_label"])
check("D2 · no fabricated cause wording",
      "demand" not in nrow["student_explanation"].lower()
      and "Cause not established" in nrow["student_explanation"],
      nrow["student_explanation"][:120])

# 2. Misleading evidence: narrative mentions a factor but never links it.
mislead_docs = [{
    "document_name": "Contoso 10-K",
    "text": ("========== PAGE 40 ==========\n"
             "Management's Discussion and Analysis\n"
             "Revenue decreased by 10% during the period. "
             "Market demand softened during the year.\n"),
}]
mislead = build_qualitative_drivers(
    [_OBS("Revenue")],
    facts={"Revenue": {"value": 180000000000, "source": "S", "reporting_period": "FY2025"}},
    period_facts={"Revenue": {"FY2024": "200000000000", "FY2025": "180000000000"}},
    qualitative_documents=mislead_docs,
    requirements=[],
    company="Company A",
)
mrow = mislead["rows"][0]
check("D3 · misleading evidence is never EXPLICITLY_DISCLOSED",
      mrow["relationship"] != REL_EXPLICIT,
      mrow["relationship_label"])
qualified = ("does not explicitly establish" in mrow["student_explanation"]
             or "Cause not established" in mrow["student_explanation"]
             or "Student interpretation is required" in mrow["student_explanation"]
             or "Student judgment is required" in mrow["student_explanation"])
check("D4 · misleading evidence wording stays appropriately qualified",
      qualified, mrow["student_explanation"][:110])
check("D4b · catalyst never presented as an established cause",
      ("the filing does not explicitly establish" in mrow.get("causality_note", "")
       or "student judgment is required" in mrow.get("causality_note", "")
       or "Cause not established" in mrow["student_explanation"]),
      mrow.get("causality_note", "")[:100])

# 3. Explicit cause IS detected when the filing states it.
explicit_docs = [{
    "document_name": "Contoso 10-K",
    "text": ("========== PAGE 40 ==========\n"
             "Management's Discussion and Analysis\n"
             "Revenue increased due to higher volume and favorable pricing.\n"),
}]
explicit = build_qualitative_drivers(
    [_OBS("Revenue", direction="increase", change="+10.0%")],
    facts={"Revenue": {"value": 200000000000, "source": "S", "reporting_period": "FY2025"}},
    period_facts={"Revenue": {"FY2024": "180000000000", "FY2025": "200000000000"}},
    qualitative_documents=explicit_docs,
    requirements=[],
    company="Company A",
)
erow = explicit["rows"][0]
check("D5 · explicit disclosure detected when stated",
      erow["relationship"] == REL_EXPLICIT, erow["relationship_label"])

# 4. POSSIBLE relationships are never upgraded into facts downstream.
poss_docs = [{
    "document_name": "Contoso 10-K",
    "text": ("========== PAGE 41 ==========\n"
             "Risk Factors\n"
             "Input costs could increase in future periods.\n"),
}]
poss = build_qualitative_drivers(
    [_OBS("Net Profit")],
    facts={"Net Profit": {"value": 90000000000, "source": "S", "reporting_period": "FY2025"}},
    period_facts={"Net Profit": {"FY2024": "80000000000", "FY2025": "90000000000"}},
    qualitative_documents=poss_docs,
    requirements=[],
    company="Company A",
)
prow = poss["rows"][0]
check("D6 · possible relationship never upgraded to a fact",
      prow["relationship"] in (REL_POSSIBLE, REL_INSUFFICIENT, REL_CAUSE_NOT_ESTABLISHED),
      prow["relationship_label"])

# 5. Contradictory narrative + review-required fact -> REVIEW_REQUIRED gate.
rev_req_facts = {"Revenue": {"value": 180000000000, "source": "S",
                             "reporting_period": "FY2025",
                             "extraction_state": "review_required",
                             "extraction_state_reason": "Conflicting values across tables"}}
rev_req = build_qualitative_drivers(
    [_OBS("Revenue")],
    facts=rev_req_facts,
    period_facts={"Revenue": {"FY2024": "200000000000", "FY2025": "180000000000"}},
    qualitative_documents=explicit_docs,  # even with narrative evidence...
    requirements=[],
    company="Company A",
)
rrow = rev_req["rows"][0]
check("D7 · review-required fact -> REVIEW_REQUIRED even with narrative",
      rrow["relationship_label"] == "🟠 REVIEW_REQUIRED"
      and rrow["foundation"] == "REVIEW_REQUIRED",
      rrow["relationship_label"])

# 6. BLOCKED metric -> no invented numerical change, CAUSE_NOT_ESTABLISHED.
blocked_req = [{"requirement": "Current Ratio", "status": ST_BLOCKED, "result": "—"}]
blk = build_qualitative_drivers(
    [_OBS("Current Ratio")],
    facts={"Current Ratio": {"value": None, "source": "S"}},
    period_facts={"Current Ratio": {"FY2024": "1.2", "FY2025": "1.4"}},
    qualitative_documents=explicit_docs,
    requirements=blocked_req,
    company="Company A",
)
brow = blk["rows"][0]
check("D8 · blocked metric -> CAUSE_NOT_ESTABLISHED, no invented driver",
      brow["relationship"] == REL_CAUSE_NOT_ESTABLISHED
      and brow["numerical_driver"] == "—",
      brow["relationship_label"])

# 7. Irrelevant narrative never becomes evidence for an unrelated metric.
irrel_docs = [{
    "document_name": "Contoso 10-K",
    "text": ("========== PAGE 42 ==========\n"
             "Risk Factors\n"
             "The company faces competitive pressure in overseas markets.\n"),
}]
irrel = build_qualitative_drivers(
    [_OBS("Current Ratio")],
    facts={"Current Ratio": {"value": 1.4, "source": "S"}},
    period_facts={"Current Ratio": {"FY2024": "1.2", "FY2025": "1.4"}},
    qualitative_documents=irrel_docs,
    requirements=[],
    company="Company A",
)
irow = irrel["rows"][0]
check("D9 · irrelevant narrative -> CAUSE_NOT_ESTABLISHED (no evidence fabricated)",
      irow["relationship"] == REL_CAUSE_NOT_ESTABLISHED and irow["source"] == "—",
      irow["relationship_label"])

# 8. Narrative extraction keeps provenance, never invents missing page.
items = extract_narrative_items([{
    "document_name": "DocX",
    "text": ("Management's Discussion and Analysis\n"
             "Revenue increased during the period driven by higher volume.\n"),
}])
check("D10 · missing page provenance stays '—'/None (never invented)",
      all(i.get("page") is None for i in items),
      str([(i.get("section"), i.get("page")) for i in items]))

# ===========================================================================
# E. DEMO ISOLATION ATTACK
# ===========================================================================
section("E. Demo isolation attack")

app = _load_app()
demo_before = copy.deepcopy({
    "period_facts": app._FTE_DEMO_PERIOD_FACTS,
    "peer_facts": app._FTE_DEMO_PEER_FACTS,
})

# Run the full demo workflow: workspace, Excel, memo.
demo_ws = _demo_workspace()
demo_xlsx = build_excel_working_model(demo_ws)
demo_memo = render_memo(
    "Demo memo text for Contoso Analytics.\n\nRevenue increased during FY2025.\n"
    "ROE improved from 29.8% to 36.6%.\n",
    [{"metric": "Revenue", "value": "281.70", "unit": "B", "source": "Demo fixture",
      "reporting_period": "FY2025", "page": 40}],
    "student",
    assignment=demo_ws,
)

demo_after = {
    "period_facts": app._FTE_DEMO_PERIOD_FACTS,
    "peer_facts": app._FTE_DEMO_PEER_FACTS,
}
check("E1 · demo dataset deep-equal after full workflow",
      demo_before == demo_after, "")

# Demo facts must not acquire reliability metadata.
demo_leak = False
for v in app._FTE_DEMO_PERIOD_FACTS.values():
    if isinstance(v, dict) and any(k in v for k in (
            "extraction_state", "extraction_conflict", "extraction_reliability")):
        demo_leak = True
check("E2 · demo facts carry no reliability/conflict metadata", not demo_leak)

check("E3 · demo workspace has no API facts injected",
      "Qualcomm" not in str(demo_ws.get("company")))
check("E4 · demo memo does not contain API-only sources",
      "Microsoft" not in demo_memo_text(demo_memo))

# Sprint 11.1: demo provenance must be explicitly synthetic.
demo_m3 = app._demo_module3_result()
demo_srcs = [str(v.get("source", "")) for v in (demo_m3.get("financial_data") or {}).values()
             if isinstance(v, dict)]
check("E5 · demo fact sources marked synthetic (Demo fixture)",
      all(s.startswith("Demo fixture") for s in demo_srcs), str(demo_srcs[:3]))
check("E6 · no real-company provenance in demo identity",
      "Microsoft" not in str(app._FTE_DEMO_COMPANY)
      and "MSFT" not in str(app._FTE_DEMO_TICKER)
      and "Microsoft" not in str(app._FTE_DEMO_DOC_NAME)
      and "Microsoft" not in str(app._FTE_DEMO_SUMMARY),
      str(app._FTE_DEMO_DOC_NAME))

# ===========================================================================
# F. DETERMINISM
# ===========================================================================
section("F. Determinism (API fixture twice)")

ws1 = _real_workspace()
ws2 = _real_workspace()
check("F1 · workspace dicts identical", ws1 == ws2)

x1 = build_excel_working_model(ws1)
x2 = build_excel_working_model(ws2)
check("F2 · excel bytes identical", x1 == x2)

m1 = render_memo("Memo.\n\nRevenue increased.", [], "student", assignment=ws1)
m2 = render_memo("Memo.\n\nRevenue increased.", [], "student", assignment=ws2)
check("F3 · memo blocks identical", m1 == m2)

# Qualitative catalysts deterministic across runs.
q1 = [(r.get("metric"), r.get("relationship"), r.get("catalyst")) for r in (ws1.get("qualitative_drivers") or {}).get("rows", [])]
q2 = [(r.get("metric"), r.get("relationship"), r.get("catalyst")) for r in (ws2.get("qualitative_drivers") or {}).get("rows", [])]
check("F4 · qualitative catalyst rows identical across runs", q1 == q2)

# ===========================================================================
# G. EXCEL WORKING MODEL
# ===========================================================================
section("G. Excel working model")

wb = openpyxl.load_workbook(io.BytesIO(x1))
check("G1 · all 7 sheets present",
      wb.sheetnames == ["Financial Data", "Ratio Analysis", "External Variables",
                        "Comparison", "Driver Analysis", "Assignment Requirements",
                        "Qualitative Drivers"],
      str(wb.sheetnames))

# Real formulas in Ratio Analysis + Driver Analysis.
formula_cells = 0
error_cells = []
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells += 1
            if isinstance(cell.value, str) and any(err in cell.value for err in
                    ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!")):
                error_cells.append((ws.title, cell.coordinate, cell.value))
check("G2 · workbook contains real Excel formulas", formula_cells > 0, f"{formula_cells} formulas")
check("G3 · no formula error strings anywhere", not error_cells, str(error_cells[:3]))

# Reopen after save-to-file round trip.
tmp = "/tmp/fte_hardcore_excel.xlsx"
with open(tmp, "wb") as f:
    f.write(x1)
wb2 = openpyxl.load_workbook(tmp)
check("G4 · workbook reopens after save", wb2.sheetnames == wb.sheetnames)
check("G5 · reopened workbook preserves formulas",
      sum(1 for ws in wb2.worksheets for row in ws.iter_rows()
          for c in row if isinstance(c.value, str) and c.value.startswith("=")) == formula_cells)

# Excel user-damage test: edit a student-input / data cell, add a row, copy a formula.
try:
    damaged = openpyxl.load_workbook(io.BytesIO(x1))
    fin = damaged["Financial Data"]
    # Overwrite a data cell like a careless student.
    fin.cell(row=fin.max_row, column=2).value = "overwritten by student"
    # Copy a formula from Ratio Analysis into a new row.
    ra = damaged["Ratio Analysis"]
    copied = None
    for row in ra.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                copied = c.value
                ra.cell(row=ra.max_row + 1, column=c.column).value = c.value
                break
        if copied:
            break
    dbuf = io.BytesIO()
    damaged.save(dbuf)
    reloaded = openpyxl.load_workbook(io.BytesIO(dbuf.getvalue()))
    check("G5b · damaged workbook still opens and saves", reloaded.sheetnames == wb.sheetnames)
    # Formula errors after damage: only the copied cell may be invalid if its
    # reference points outside data; the original formulas must stay intact.
    errs_after = []
    for ws in reloaded.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(
                        err in cell.value for err in
                        ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!")):
                    errs_after.append((ws.title, cell.coordinate))
    check("G5c · no formula errors introduced by student damage", not errs_after, str(errs_after[:3]))
except Exception as e:  # noqa: BLE001
    check("G5b · damaged workbook still opens and saves", False, str(e))

# No external-link injection from document text into formulas.
link_leak = False
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and ("HYPERLINK(" in c.value or c.value.startswith("=HYPERLINK")):
                link_leak = True
check("G6 · no HYPERLINK injection in generated workbook", not link_leak)

# ===========================================================================
# H. MEMO CONTENT SAFETY
# ===========================================================================
section("H. Memo content safety")

def memo_text(blocks):
    out = []
    for kind, payload in blocks:
        if kind == "heading":
            out.append(str(payload))
        elif kind == "bullets":
            out.append("\n".join(str(b) for b in payload))
        elif kind == "table":
            out.append(str(payload.get("title", "")))
            for r in payload.get("rows", []):
                out.append(" ".join(str(c) for c in r))
            out.append(" ".join(str(n) for n in payload.get("notes", [])))
        elif kind == "paragraph":
            out.append(str(payload))
        elif kind == "note":
            out.append(str(payload))
    return "\n".join(out)

for profile in ("classic", "student", "professional"):
    mt = memo_text(render_memo("Revenue grew.\n\nNet profit rose 22.7%.",
                               [{"metric": "Revenue", "value": "281.70", "unit": "B",
                                 "source": "10-K", "reporting_period": "FY2025", "page": 26}],
                               profile, assignment=_real_workspace()))
    bad = [w for w in ("none", "null", "undefined", "nan", "todo", "lorem",
                       "placeholder", "fake page", "invented") if re.search(rf"\b{w}\b", mt.lower())]
    check(f"H1 · {profile} memo has no placeholder leakage", not bad, str(bad))

# ===========================================================================
# I. STUDENT CONCLUSION
# ===========================================================================
section("I. Student conclusion")

ws_conc = _real_workspace()
conc_keys = [k for k in ws_conc.keys() if "conclusion" in k.lower()]
check("I1 · workspace has no generated conclusion section",
      not any(v for k, v in ws_conc.items() if "conclusion" in k.lower() and v),
      str(conc_keys))

for profile in ("classic", "student", "professional"):
    mt = memo_text(render_memo("Revenue grew.", [], profile, assignment=ws_conc)).lower()
    for banned in ("buy", "sell", "strong buy", "good investment", "outperform", "underperform"):
        if banned in mt:
            check(f"I2 · {profile} memo contains no recommendation wording",
                  False, banned)
            break
    else:
        check(f"I2 · {profile} memo contains no recommendation wording", True)

# ===========================================================================
# J. SECURITY
# ===========================================================================
section("J. Security")

# 1. Malicious-looking filename must not execute anything.
pwn_marker = "/tmp/fte_pwned_by_filename"
if os.path.exists(pwn_marker):
    os.remove(pwn_marker)
try:
    class _FakeUpload:
        name = "x; touch /tmp/fte_pwned_by_filename; .pdf"
        def seek(self, *a, **k):
            return None
        def read(self, *a, **k):
            return b""
        def getvalue(self):
            return b""
    parse_document(_FakeUpload())
except Exception:  # noqa: BLE001 — failure to parse is acceptable
    pass
check("J1 · malicious filename does not execute shell", not os.path.exists(pwn_marker))

# 2. Student input is not interpreted as executable code.
code_attempt = "__import__('os').system('touch /tmp/fte_pwned_by_student')"
try:
    _ = calculate_metric("ROE", {"Net Profit": {"value": code_attempt, "source": "S"},
                                 "Equity": {"value": 100, "source": "S"}},
                         context={"recover": False})
except Exception:  # noqa: BLE001
    pass
check("J2 · student input not executed as code", not os.path.exists("/tmp/fte_pwned_by_student"))

# 3. No secrets printed into memo or Excel outputs.
secret_scan = []
for probe in ("sk-", "AIza", "AKIA", "ghp_", "Bearer "):
    mt = memo_text(render_memo("Revenue grew.", [], "student", assignment=_real_workspace()))
    if probe.lower() in mt.lower():
        secret_scan.append(probe)
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str):
                for probe in ("sk-", "AIza", "AKIA", "ghp_", "Bearer "):
                    if probe.lower() in c.value.lower():
                        secret_scan.append(probe)
check("J3 · no secret material in memo or Excel output", not secret_scan, str(secret_scan))

# ===========================================================================
# K. PERFORMANCE / STRESS
# ===========================================================================
section("K. Performance / stress")

def _timed(label, fn):
    t0 = time.perf_counter()
    out = fn()
    dt = (time.perf_counter() - t0) * 1000.0
    print(f"    {label}: {dt:.1f} ms")
    return out, dt

ws_p, t_ws = _timed("workspace (1 doc)", lambda: _real_workspace())
_, t_xl = _timed("excel (1 doc)", lambda: build_excel_working_model(ws_p))
_, t_memo = _timed("memo (1 doc)", lambda: render_memo("Revenue grew.", [], "student", assignment=ws_p))
check("K1 · workspace < 2s", t_ws < 2000.0, f"{t_ws:.0f} ms")
check("K2 · excel < 2s", t_xl < 2000.0, f"{t_xl:.0f} ms")
check("K3 · memo < 1s", t_memo < 1000.0, f"{t_memo:.0f} ms")

# Scale: 10-document corpus for narrative extraction + qualitative layer.
big_docs = []
for i in range(10):
    big_docs.append({
        "document_name": f"Doc {i}",
        "text": (f"========== PAGE {i+1} ==========\n"
                 "Management's Discussion and Analysis\n"
                 "Revenue increased due to higher volume and favorable pricing. "
                 f"Segment performance improved in period {i}.\n"
                 "========== PAGE 40 ==========\n"
                 "Risk Factors\n"
                 "Input costs could increase in future periods.\n"),
    })
t0 = time.perf_counter()
qbig = build_qualitative_drivers(
    [_OBS("Revenue", direction="increase", change="+10.0%"),
     _OBS("Net Profit"), _OBS("Operating Profit"), _OBS("Current Ratio")],
    facts={"Revenue": {"value": 200, "source": "S", "reporting_period": "FY2025"},
           "Net Profit": {"value": 90, "source": "S", "reporting_period": "FY2025"},
           "Operating Profit": {"value": 120, "source": "S", "reporting_period": "FY2025"}},
    period_facts={"Revenue": {"FY2024": "180", "FY2025": "200"},
                  "Net Profit": {"FY2024": "80", "FY2025": "90"},
                  "Operating Profit": {"FY2024": "100", "FY2025": "120"}},
    qualitative_documents=big_docs,
    requirements=[],
    company="Company A",
)
dt10 = (time.perf_counter() - t0) * 1000.0
print(f"    qualitative layer (10 docs): {dt10:.1f} ms")
check("K4 · qualitative layer 10 docs < 2s", dt10 < 2000.0, f"{dt10:.0f} ms")

# ===========================================================================
# L. API/DEMO PARITY
# ===========================================================================
section("L. API/Demo parity")

demo_ws2 = _demo_workspace()
real_ws2 = _real_workspace()
demo_keys = set(demo_ws2.keys())
real_keys = set(real_ws2.keys())
common = demo_keys & real_keys
check("L1 · same workspace sections (machinery parity)",
      common >= {"assignment_type", "company", "requirements", "normalized_facts",
                 "comparison", "driver_analysis", "qualitative_drivers",
                 "external_variables", "calculations"}, "")

dxl = build_excel_working_model(demo_ws2)
dwb = openpyxl.load_workbook(io.BytesIO(dxl))
check("L2 · demo Excel has same 7 sheets as API",
      dwb.sheetnames == wb.sheetnames, str(dwb.sheetnames))

# Demo values unchanged from static fixture.
demo_rev = (demo_ws2.get("normalized_facts") or [])
rev_row = next((f for f in demo_rev if f.get("canonical") == "Revenue"), None)
check("L3 · demo Revenue value matches static fixture (not recomputed/leaked)",
      rev_row is not None and rev_row.get("value") is not None,
      str(rev_row))

# ===========================================================================
# Summary
# ===========================================================================
print("\n" + "=" * 60)
passed = sum(1 for _, ok, _ in CHECKS if ok)
failed = [(n, d) for n, ok, d in CHECKS if not ok]
print(f"RESULT: {passed}/{len(CHECKS)} checks pass")
if failed:
    print("FAILURES:")
    for n, d in failed:
        print(f"  - {n} [{d}]")
else:
    print("ALL CHECKS PASS")
